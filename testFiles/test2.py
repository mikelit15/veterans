import openpyxl
import os
import shutil
import re
import sys
from openpyxl.worksheet.hyperlink import Hyperlink
sys.path.append(r'C:\workspace\veterans\microsoftOCR')
import microsoftOCR


def process_cemetery(cem, cemeteryName, basePath, initialCount, start, startFlag, redac, redac2):
    uppercase_alphabet = [chr(i) for i in range(ord('A'), ord('Z') + 1)]
    firstFileFlag = True
    for letter in uppercase_alphabet:
        name_path = os.path.join(basePath, cemeteryName, letter)
        try:
            print(f"Processing {cemeteryName} - {letter}")
            initialCount, startFlag = clean(cem, letter, name_path, os.path.join(basePath, cemeteryName), \
                initialCount, firstFileFlag, start, startFlag, redac2)
            firstFileFlag = False
        except FileNotFoundError:
            continue
    return initialCount, startFlag


def clean(cem, letter, namePath, cemPath, initialCount, isFirstFile, start, startFlag, redac2):
    pdfFiles = sorted(os.listdir(namePath))
    letters = sorted([folder for folder in os.listdir(cemPath) if os.path.isdir(os.path.join(cemPath, folder))])
    try:
        currentLetterIndex = letters.index(letter)
    except ValueError:
        return
    folderBeforeIndex = max(0, currentLetterIndex - 1)
    folderBefore = letters[folderBeforeIndex]
    pdfFilesBefore = sorted([file for file in os.listdir(os.path.join(cemPath, folderBefore)) if file.lower().endswith('.pdf')])
    maxCounter = 0
    for file in pdfFilesBefore:
        match = re.search(r'\d+', file)
        if match:
            number = int(match.group())
            maxCounter = max(maxCounter, number)
    counter = maxCounter + 1
    for x in pdfFiles:
        if counter != start and startFlag:
            isFirstFile = False
            if "a" in x[-5:]:
                pass
            else:
                counter += 1
            continue
        startFlag = False
        if isFirstFile:
            counter = initialCount 
        if redac2:
            if "a redacted.pdf" in x:
                counter -= 1
                os.remove(os.path.join(namePath, x))
                print(f"Removed {x}")
            elif "b redacted.pdf" in x:
                counter -= 1
                os.remove(os.path.join(namePath, x))
                print(f"Removed {x}")
            else:
                newName = f"{cemetery}{letter}{counter:05d}{redac2}.pdf"
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
        else:
            if "a.pdf" in x:
                newName = f"{cemetery}{letter}{counter:05d}a.pdf"
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
                counter -= 1
            elif "b.pdf"  in x:
                newName = f"{cemetery}{letter}{counter:05d}b.pdf"
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
            else:    
                newName = f"{cemetery}{letter}{counter:05d}.pdf"
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
        counter += 1
        isFirstFile = False
    return counter, startFlag


def cleanImages(cem, goodID, redac, redac2):
    baseCemPath = fr"\\ucclerk\pgmdoc\Veterans\test{redac}"
    cemeterys = [d for d in os.listdir(baseCemPath) if os.path.isdir(os.path.join(baseCemPath, d))]
    initialCount = 1
    start = goodID - 200
    startFlag = True
    for cemetery in cemeterys:
        if cemetery in ["Jewish", "Misc"]:
            subCemPath = os.path.join(baseCemPath, cemetery)
            subCemeteries = [d for d in os.listdir(subCemPath) if os.path.isdir(os.path.join(subCemPath, d))]
            for subCemetery in subCemeteries:
                print(f"Processing {cemetery} - {subCemetery}")
                initialCount, startFlag = process_cemetery(cem, subCemetery, subCemPath, initialCount, start, startFlag, redac, redac2)
        else:
            print(f"Processing {cemetery}")
            initialCount, startFlag = process_cemetery(cem, cemetery, baseCemPath, initialCount, start, startFlag, redac, redac2)
      

def adjustImageName(cemetery, goodID, badID, goodRow):
    baseCemPath = r"\\ucclerk\pgmdoc\Veterans\test"
    goodIDFound = False
    badIDFound = False
    goodIDFilePath = ""
    badIDFilePath = ""
    for dirpath, dirnames, filenames in os.walk(baseCemPath):
        for filename in filenames:
            if f"{goodID:05d}" in filename:
                goodIDFilePath = os.path.join(dirpath, filename)
                goodIDFound = True
            elif f"{badID:05d}" in filename:
                badIDFilePath = os.path.join(dirpath, filename)
                badIDFound = True
            if goodIDFound and badIDFound:
                break
        if goodIDFound and badIDFound:
            break
    if goodIDFound and badIDFound:
        goodIDPrefix = os.path.basename(os.path.dirname(goodIDFilePath))
        badIDPrefix = os.path.basename(os.path.dirname(badIDFilePath))
        newBadIDFilename = os.path.basename(badIDFilePath).replace(f"{badID:05d}", f"{goodID:05d}b").replace(badIDPrefix, goodIDPrefix)
        newBadIDFilePath = os.path.join(os.path.dirname(goodIDFilePath), newBadIDFilename)
        if not "a.pdf" in os.path.basename(goodIDFilePath):
            newGoodIDFilename = os.path.basename(goodIDFilePath).replace(".pdf", "a.pdf")
            newGoodIDFilePath = os.path.join(os.path.dirname(goodIDFilePath), newGoodIDFilename)
            os.rename(goodIDFilePath, newGoodIDFilePath)
            print(f"{os.path.basename(goodIDFilePath)} renamed to {newGoodIDFilename}")
        shutil.move(badIDFilePath, newBadIDFilePath)
        print(f"{os.path.basename(badIDFilePath)} moved and renamed to {newBadIDFilename}")
    else:
        print("GoodID or BadID file not found. Please check the IDs and directories.")
    start_column = 'A'
    end_column = 'O'
    start_column_index = openpyxl.utils.column_index_from_string(start_column)
    end_column_index = openpyxl.utils.column_index_from_string(end_column)
    for col_index in range(start_column_index, end_column_index + 1):
        worksheet.cell(row=goodRow, column=col_index).value = None
    worksheet[f'O{goodRow}'].hyperlink = None
    print(f"Record {goodID} data from row {goodID + 1} cleared successfully.")


def cleanDelete(cemetery, badID, badRow):
    baseRedacPath = r"\\ucclerk\pgmdoc\Veterans\test - Redacted"
    for dirpath, dirnames, filenames in os.walk(baseRedacPath):
        for filename in filenames:
            if f"{badID:05d}" in filename:
                file_path = os.path.join(dirpath, filename)
                os.remove(file_path)
                print(f"{filename} deleted successfully.")
    worksheet = workbook[cemetery]
    worksheet.delete_rows(badRow)
    for row in range(badRow, worksheet.max_row + 1):
        worksheet[f'A{row}'].value = worksheet[f'A{row}'].value - 1
        print(f"{worksheet[f'A{row}'].value + 1} changed to {worksheet[f'A{row}'].value}")
        next_row = row + 1
        cell_ref = f'O{row}'
        next_cell_ref = f'O{next_row}'
        if next_row <= worksheet.max_row and worksheet[next_cell_ref].hyperlink:
            temp_hyperlink = worksheet[next_cell_ref].hyperlink
            worksheet[cell_ref].hyperlink = Hyperlink(ref=cell_ref, target=temp_hyperlink.target, display=temp_hyperlink.display)
            worksheet[cell_ref].value = worksheet[next_cell_ref].value
            print(f"Hyperlink and value from row {next_row} moved to row {row}.")
    if worksheet[f'O{worksheet.max_row - 1}'].hyperlink:
        last_row = worksheet.max_row
        print(f"Adjusting the last hyperlink in row {last_row}.")
        second_last_hyperlink = worksheet[f'O{last_row - 1}'].hyperlink
        new_target = re.sub(r'(\d+)(?=%\d+)', lambda m: f"{int(m.group(1)) + 1:05d}", second_last_hyperlink.target)
        worksheet[f'O{last_row}'].hyperlink = Hyperlink(ref=f'O{last_row}', target=new_target, display=second_last_hyperlink.display)
        print(f"Updated hyperlink in row {last_row} to {new_target}.")


def cleanHyperlinks(length, startIndex):
    for x in range(2, worksheet.max_row + 1):
        worksheet[f'A{x}'].value = temp2
        temp2 = temp2 + 1
    for x in range(length):
        worksheet[f'O{worksheet.max_row - x}'].value = None
        worksheet[f'O{worksheet.max_row - x}'].hyperlink = None
    for row in range(badRow, worksheet.max_row + 1):
        worksheet[f'A{row}'].value = worksheet[f'A{row}'].value - 1
        print(f"{worksheet[f'A{row}'].value + 1} changed to {worksheet[f'A{row}'].value}")
        next_row = row + 1
        cell_ref = f'O{row}'
        next_cell_ref = f'O{next_row}'
        if next_row <= worksheet.max_row and worksheet[next_cell_ref].hyperlink:
            temp_hyperlink = worksheet[next_cell_ref].hyperlink
            worksheet[cell_ref].hyperlink = Hyperlink(ref=cell_ref, target=temp_hyperlink.target, display=temp_hyperlink.display)
            worksheet[cell_ref].value = worksheet[next_cell_ref].value
            print(f"Hyperlink and value from row {next_row} moved to row {row}.")
    if worksheet[f'O{worksheet.max_row - 1}'].hyperlink:
        last_row = worksheet.max_row
        print(f"Adjusting the last hyperlink in row {last_row}.")
        second_last_hyperlink = worksheet[f'O{last_row - 1}'].hyperlink
        new_target = re.sub(r'(\d+)(?=%\d+)', lambda m: f"{int(m.group(1)) + 1:05d}", second_last_hyperlink.target)
        worksheet[f'O{last_row}'].hyperlink = Hyperlink(ref=f'O{last_row}', target=new_target, display=second_last_hyperlink.display)
        print(f"Updated hyperlink in row {last_row} to {new_target}.")
    
    
cemetery = "Fairview"
goodIDs = [6600, 6620, 6640, 6660, 6680] 
badIDs = [6605, 6625, 6645, 6665, 6685]
excelFilePath = r"\\ucclerk\pgmdoc\Veterans\Veterans2.xlsx"
workbook = openpyxl.load_workbook(excelFilePath)
worksheet = workbook[cemetery]
for x in range(0, len(goodIDs)):
    for row in range(1, worksheet.max_row + 1):
        if worksheet[f'A{row}'].value == goodIDs[x]:
            goodRow = row
        if worksheet[f'A{row}'].value == badIDs[x]:
            badRow = row
    adjustImageName(cemetery, goodIDs[x], badIDs[x], goodRow)
    microsoftOCR.main(True, cemetery, worksheet[f'B{goodRow}'].value[0])
    cleanDelete(cemetery, badIDs[x], badRow)
    workbook.save(excelFilePath)
cleanImages(cemetery, 6600, "", "")
cleanImages(cemetery, 6400, " - Redacted", " redacted")
for row in range(1, worksheet.max_row + 1):
    if worksheet[f'A{row}'].value == goodIDs[x]:
        startIndex = row
cleanHyperlinks(len(goodIDs), startIndex)
workbook.save(excelFilePath)
