import re
import openpyxl 
from azure.core.credentials import AzureKeyCredential
from azure.ai.formrecognizer import DocumentAnalysisClient
from openpyxl.styles import Font
from collections import defaultdict
import traceback
import os
import PyPDF2
import fitz
import cv2
from reportlab.pdfgen import canvas
from PIL import Image
from nameparser import HumanName
import dateparser
from nameparser.config import CONSTANTS
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz


'''
Redacts sensitive information (e.g., social security or ID numbers) from a registration 
card by adding a black box over the specified area. The method of redaction varies based 
on the card format. The first page of the PDF is saved as a PNG, and the black box is 
added to the PNG. A new PDF is then created using the redacted PNG for the first page and 
the original second page.

@param file (str) - Path to the file to be processed
@param cemetery (str) - Name of the cemetery, used for categorizing and saving the 
                        file in the correct directory
@param letter (str) - Last name letter, used for categorizing and saving the 
                      file in the correct directory
@param nameCoords (list) - Bounding box coordinates for "NAME" key
@param serialCoords (list) - Bounding box coordinates for "SERIAL" key

@return new_pdf_file (str) - Path to the newly created redacted PDF file

@author Mike
'''
def redact(file, cemetery, letter, nameCoords, serialCoords):
    pdfDocument  = fitz.open(file)
    firstPage  = pdfDocument.load_page(0)
    first_page_image  = firstPage.get_pixmap(matrix=fitz.Matrix(600/72, 600/72))
    imageFile  = "temp.png"
    first_page_image.save(imageFile)
    img = cv2.imread(imageFile)
    if serialCoords:
        pt1 = (serialCoords[0][0]-75, serialCoords[0][1]-350)
        pt3 = (3525, serialCoords[1][1]+50)
    else:
        pt1 = (nameCoords[0][0]+2050, nameCoords[0][1]-350)
        pt3 = (3525, nameCoords[1][1]+50)
    cv2.rectangle(img, pt1, pt3, (0, 0, 0), thickness=cv2.FILLED)
    cv2.imwrite(imageFile, img)
    image = Image.open(imageFile)
    pdfFile = "temp.pdf"
    c = canvas.Canvas(pdfFile, pagesize=(image.width, image.height))
    c.drawImage(imageFile, 0, 0, width=image.width, height=image.height)
    c.save()
    image.close()

    fullLocation = r"\\ucclerk\pgmdoc\Veterans\Cemetery - Redacted"
    redactedLocation = f'{cemetery} - Redacted'
    fullLocation = os.path.join(fullLocation, redactedLocation)
    fullLocation = os.path.join(fullLocation, letter)
    redacted_pdf_document  = fitz.open(pdfFile)
    new_pdf_document  = fitz.open()
    new_pdf_document.insert_pdf(redacted_pdf_document, from_page=0, to_page=0)
    new_pdf_document.insert_pdf(pdfDocument, from_page=1, to_page=1)
    temp = file[-9:]
    if "a" in temp or "b" in temp:
        temp = file[-10:]
    new_pdf_file = f'{fullLocation}\\{cemetery}{letter}{temp.replace(".pdf", "")} redacted.pdf'
    new_pdf_document.save(new_pdf_file)
    new_pdf_document.close()
    redacted_pdf_document.close()
    pdfDocument.close()
    return new_pdf_file    


'''
Merges 'A' and 'B' redacted files into a single combined file, showing both versions 
of the card, both redacted.

@param pathA (str) - File path for file A (first document to be merged)
@param pathB (str) - File path for file B (second document to be merged)
@param cemetery (str) - Name of the cemetery, used for categorizing and saving the 
                        merged file in the correct directory
@param letter (str) - Last name letter, used for categorizing and saving the 
                      merged file in the correct directory

@author Mike
'''
def mergeImages(pathA, pathB, cemetery, letter):
    fullLocation = r"\\ucclerk\pgmdoc\Veterans\Cemetery - Redacted"
    redactedLocation = f'{cemetery} - Redacted'
    fullLocation = os.path.join(fullLocation, redactedLocation)
    fullLocation = os.path.join(fullLocation, letter)
    fileName = pathA.split(letter)
    merger = PyPDF2.PdfMerger()
    merger.append(pathA)
    merger.append(pathB)
    # with open(f'\\ucclerk\pgmdoc\Veterans\\{cemetery}{letter}{fileName[2].replace("a.pdf", ".pdf")}', 'wb') as out_pdf:
    #     merger.write(out_pdf)
    merger2 = PyPDF2.PdfMerger()
    merger2.append(f'{fullLocation}\\{cemetery}{letter}{fileName[-1].replace(".pdf", "")} redacted.pdf')
    merger2.append(f'{fullLocation}\\{cemetery}{letter}{fileName[-1].replace("a.pdf", "b")} redacted.pdf')
    with open(f'{fullLocation}\\{cemetery}{letter}{fileName[-1].replace("a.pdf", "")} redacted.pdf', 'wb') as out_pdf:
        merger2.write(out_pdf)
        
        
'''
Extracts field matches from the specified file using Microsoft Azure Textract, which 
performs an AI analysis to identify and extract text from the form. Uses a custom built
model that was trained to understand the specific form layout and variable placement per
field needed from extraction.

@param filePath (str) - File path for the file being processed 
@param id (int) - ID of the file, used for referencing the specific file that is 
                  currently being processed, printed in the terminal

@return result (AnalyzeResult) - The result text extraction of the fields that were found 
                                 by the custom model.

@author Mike
'''
def analyzeDocument(filePath, id):
    endpoint = os.environ["VISION_ENDPOINT"]
    key = os.environ["VISION_KEY"]
    document_analysis_client = DocumentAnalysisClient(endpoint=endpoint, credential=AzureKeyCredential(key))
    with open(filePath, "rb") as file:
        imgTest = file.read()
        bytesTest = bytearray(imgTest)
        print('Image loaded', id)
    poller = document_analysis_client.begin_analyze_document("Test21n", document=bytesTest)
    result = poller.result()
    return result


'''
Creates a dictionary storing the values for each field extracted from the document. 
Associates the values with the specific key names that are used throughout the code.

@param result (AnalyzeResult) - The result text extraction of the fields that were found 
                                by the custom model.

@return kvs (defaultdict(list)) - The result text extraction of the fields that were found 
                                  by the custom model.
@return nameCoords (list) - Bounding box coordinates for "NAME" key
@return serialCoords (list) - Bounding box coordinates for "SERIAL" key
@return war (string) - Secondary check for if war record info is placed outside of war 
                       record field.

@author Mike
'''
def extract_key_value_pairs(result):
    kvs = defaultdict(list)
    war = ""
    nameCoords = []
    serialCoords = []
    for idx, invoice in enumerate(result.documents):
        name = invoice.fields.get("Name")
        if name.content != None:
            kvs["NAME"].append(name.content)
        nameCoord = invoice.fields.get("nameCoord")
        if nameCoord.content != None:
            for region in nameCoord.bounding_regions:
                polygon = region.polygon
                nameCoords.append((int(polygon[0].x * 600), int(polygon[0].y * 600)))
                nameCoords.append((int(polygon[2].x * 600), int(polygon[2].y * 600)))
        serial = invoice.fields.get("Serial No.")
        if serial:
            for region in serial.bounding_regions:
                polygon = region.polygon
                serialCoords.append((int(polygon[0].x * 600), int(polygon[0].y * 600)))
                serialCoords.append((int(polygon[2].x * 600), int(polygon[2].y * 600)))
        born = invoice.fields.get("Born")
        if born.content != None:
            kvs["BORN"].append(born.content)
        death = invoice.fields.get("Date of Death")
        if death.content != None:
            kvs["DATE OF DEATH"].append(death.content)
        buried = invoice.fields.get("Buried")
        if buried.content != None:
            kvs["BURIED"].append(buried.content)
        century = invoice.fields.get("19")
        if century.content != None:
            kvs["19"].append(century.content)
        cemetery = invoice.fields.get("IN")
        if cemetery.content != None:
            kvs["IN"].append(cemetery.content)
        warRecord = invoice.fields.get("War Record")
        if warRecord.content != None:
            kvs["WAR RECORD"].append(warRecord.content)
        branch = invoice.fields.get("Branch of Service")
        if branch.content != None:
            kvs["BRANCH OF SERVICE"].append(branch.content)
        app = invoice.fields.get("App")
        if app.content != None:
            kvs["Application"].append(app.content)
        world = invoice.fields.get("War")
        if world.content != None:
            if "War" in world.content:
                kvs[""].append(world.content)
                war = world.content
        # care = invoice.fields.get("Care Assigned")
        # if care.content != None:
        #     extracted.append(True)
    return kvs, nameCoords, serialCoords, war


'''
Prints the key-value pairs in a readable format. Useful for debugging 
or presenting the extracted data.

@param kvs (defaultdict(list)) - Dictionary containing the key-value pairs to be printed

@author Mike
'''
def print_kvs(kvs):
    print("----Key-value pairs found in document----")
    for key, value in kvs.items():
        print(key, ":", value)
    print("\n")
    
    
'''
Searches for and retrieves the value associated with a specified key from the dictionary 
of key-value pairs. Compares strings for equality.

@param kvs (defaultdict(list)) - Dictionary containing key-value pairs
@param search_key (str) - The key for which the value needs to be retrieved for

@return value (str or None) - The value associated with the search_key, if found. 
                              Otherwise, None

@author Mike
''' 
def search_value(kvs, searchKey):
    for key, value in kvs.items():
        if key.strip().upper() == searchKey.upper():
            return value


'''
Searches for and retrieves the value associated with a specified key from the dictionary 
of key-value pairs. Uses REGEX instead of comparing strings.

@param kvs (defaultdict(list)) - Dictionary containing key-value pairs
@param search_key (str) - The key for which the value needs to be retrieved for

@return value (str or None) - The value associated with the search_key, if found. 
                            Otherwise, None

@author Mike
''' 
def search_value_regex(kvs, searchKey):
    searchPattern = r'\b' + re.escape(searchKey) + r'\b'
    for key, value in kvs.items():
        if re.search(searchPattern, key, re.IGNORECASE):
            return value


'''
Parses and formats a full name into its constituent parts: first name, middle name, 
last name, and suffix. The function handles various naming conventions and appends 
the formatted name components to a list.

@param finalVals (list) - List where the processed name components will be appended
@param value (str) - The extracted raw name to be processed

@author Mike
''' 
def nameRule(finalVals, value):
    value = value.replace("NAME", "").replace("Name", "").replace("name", "")\
        .replace("\n", " ").replace(".", " ")
    CONSTANTS.force_mixed_case_capitalization = True
    name = HumanName(value)
    name.capitalize()
    firstName = name.first
    middleName = name.middle
    lastName = name.last
    suffix = name.suffix
    title = name.title
    suffi = ["Jr", "Sr", "I", "II", "III", "IV", "V"]
    temp = value.replace("Jr", "").replace("Sr", "").replace("I", "").replace("II", "")\
        .replace("III", "").replace("IV", "").replace("V", "")
    if ("," in value and "." in firstName):
        if len(middleName) > 0:
            middleName = name.first
            firstName = name.middle
        else:
            lastName = name.last
            firstName = name.first
    elif ("." in firstName):
        if len(middleName) > 0:
            lastName = name.first
            firstName = name.middle
            middleName = name.last
        else:
            lastName = name.first
            firstName = name.last
    elif ("." in lastName):
        if len(middleName) == 0:
            lastName = name.last
            firstName = name.first
        else:
            lastName = name.first
            middleName = name.last
            firstName = name.middle
    elif len(suffix) > 0 and not middleName:
        suffix = suffix.replace(", ", "")
        middleName = suffix.replace("Sr", "").replace("Jr", "").replace("I", "").replace("II", "")\
        .replace("III", "").replace("IV", "").replace("V", "")
        suffix = suffix.replace(middleName, "")
    elif len(suffix) > 0 and middleName:
        suffix = suffix.replace(", ", "")
        suffix = suffix.replace(middleName, "")
    if value in suffi and "." not in temp and "," not in temp:
        if len(middleName) > 0:
            firstName = name.middle
            middleName = name.last
            lastName = name.first
        else:
            firstName = name.last
            lastName = name.first
    elif "." not in temp and "," not in temp:
        if len(middleName) > 0:
            firstName = name.middle
            middleName = name.last
            lastName = name.first
        else:
            lastName = name.first
            firstName = name.last
    if len(middleName) > 2:
        middleName = middleName.replace(".", "")
    dots = 0
    for x in firstName:
        if (x == "."):
            dots+=1
    if (dots == 2):
        firstName = firstName.upper()
    else:
        firstName = firstName.replace(".", "")
    if title and not suffix:
        suffix = title
    if middleName in suffi:
        suffi = middleName
        middleName = ""
    suffix = re.sub(r"[^a-zA-Z']", '', suffix)
    if suffix:
        if not "." in suffix:
            suffix += "."
    try:
        if middleName[-1] == "/":
            middleName = middleName[:-1]
    except IndexError:
        pass
    middleName = re.sub(r"[^a-zA-Z']", '', middleName)
    if len(middleName) == 1:
        middleName += "."
    if firstName.replace(".", "").lower() == "wm":
        firstName = "William"
    finalVals.append(re.sub(r"[^a-zA-Z']", '', lastName))
    finalVals.append(re.sub(r"[^a-zA-Z']", '', firstName))
    finalVals.append(middleName.replace("0", "O"))
    finalVals.append(suffix)

'''
Processes and formats birth date information. Handles partial and full dates, converting 
them into a consistent format.

@param birth (str) - Extracted raw full birth date
@param bYear (str) - Birth year, particularly if only the year is provided
@param birthYYFlag (bool) - Flag indicating if the birth year is provided as a two-digit 
                            number

@return birth (str) - Cleaned and formatted full birth date
@return bYear (str) - Birth year, seperated from full date if applicable
@return birthYYFlag (bool) - Updated flag indicating if the birth year is provided as 
                             a two-digit number

@author Mike
''' 
def parseBirth(birth, bYear, birthYYFlag):
    if birth.count('/') == 1:
        birth = birth.replace("/", " ")
        parseBirth(birth, bYear, birthYYFlag)
    pattern = r'([A-Za-z]+),\s*(\d{4})'
    match = re.match(pattern, birth.strip())
    if birth[-1:] == " ":
        birth = birth[:-1]
    if match:
        bYear = match.group(2)
        birth = ""
    elif birth.count('/') == 2:
        year = birth.split('/')[-1]
        year = year.replace(" ", "")
        if len(year) == 4:
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) == 2:
            birthYYFlag = True
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")[-2:]
            birth = birth.strftime("%m/%d/%Y")[:-2]
        elif len(year) == 6 or len(year) == 5:
            bYear = year[:4]
            birth = birth.replace(",", " ")
            birth = dateparser.parse(birth)
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) > 6:
            temp = year[:4]
            if temp.isnumeric():
                death_no_spaces = birth.replace(' ', '')
                startIndex = death_no_spaces.find(year[4:])
                num_spaces_before = birth[:startIndex].count(' ')
                adjusted_start_index = startIndex + num_spaces_before
                birth = birth[:adjusted_start_index].strip()
                birth = dateparser.parse(birth)
                birth = birth.strftime("%m/%d/%Y")
                bYear = temp
            else:
                birth = ""
        else:
            birth = ""
            bYear = ""
    elif "," in birth:
        year = birth.split(',')[-1]
        year = year.replace(" ", "")
        if len(year) == 4:
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) == 2:
            birthYYFlag = True
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")[-2:]
            birth = birth.strftime("%m/%d/%Y")[:-2]
        elif len(year) == 6 or len(year) == 5:
            bYear = year[-4:]
            birth = birth.replace(",", " ")
            birth = dateparser.parse(birth)
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) > 6:
            temp = year[:4]
            if temp.isnumeric():
                death_no_spaces = birth.replace(' ', '')
                startIndex = death_no_spaces.find(year[4:])
                num_spaces_before = birth[:startIndex].count(' ')
                adjusted_start_index = startIndex + num_spaces_before
                birth = birth[:adjusted_start_index].strip()
                birth = dateparser.parse(birth)
                birth = birth.strftime("%m/%d/%Y")
                bYear = temp
            else:
                birth = ""
        else:
            birth = ""
    elif " " in birth:
        year = birth.split(' ')[-1]
        year = year.replace(" ", "")
        if len(year) == 4 and year[0] == "1":
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) == 2:
            birthYYFlag = True
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")[-2:]
            birth = birth.strftime("%m/%d/%Y")[:-2]
        elif len(year) == 6 or len(year) == 5:
            bYear = year[-4:]
            birth = birth.replace(",", " ")
            birth = dateparser.parse(birth)
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) > 6:
            temp = year[:4]
            if temp.isnumeric():
                death_no_spaces = birth.replace(' ', '')
                startIndex = death_no_spaces.find(year[4:])
                num_spaces_before = birth[:startIndex].count(' ')
                adjusted_start_index = startIndex + num_spaces_before
                birth = birth[:adjusted_start_index].strip()
                birth = dateparser.parse(birth)
                birth = birth.strftime("%m/%d/%Y")
                bYear = temp
            else:
                birth = ""
        else:
            birth = ""
    return birth, bYear, birthYYFlag


'''
Processes and formats death date information. Handles partial and full dates, converting 
them into a consistent format.

@param death (str) - Extracted raw full death date 
@param dYear (str) - Death year, particularly if only the year is provided
@param deathYYFlag (bool) - Flag indicating if the death year is provided as a two-digit 
                            number
 
@return death (str) - Cleaned and formatted full death date
@return dYear (str) - Death year, seperated from full date if applicable
@return deathYYFlag (bool) - Updated flag indicating if the death year is provided as a 
                             two-digit number

@author Mike
''' 
def parseDeath(death, dYear, deathYYFlag):
    if death.count('/') == 1:
        death = death.replace("/", " ")
        parseDeath(death, dYear, deathYYFlag)
    pattern = r'([A-Za-z]+),\s*(\d{4})'
    match = re.match(pattern, death.strip())
    if death[-1:] == " ":
        death = death[:-1]
    if match:
        dYear = match.group(2)
        death = ""
    elif death.count('/') == 2:
        year3 = death.split('/')[-1]
        year3 = year3.replace(" ", "")
        if len(year3) == 4:
            death = dateparser.parse(death)
            dYear = death.strftime("%Y")
            death = death.strftime("%m/%d/%Y")
        elif len(year3) == 2:
            deathYYFlag = True
            death = dateparser.parse(death)
            dYear = death.strftime("%Y")[-2:]
            death = death.strftime("%m/%d/%Y")[:-2]
        elif len(year3) == 6 or len(year3) == 5:
            dYear = year3[-4:]
            death = death.replace(",", " ")
            death = dateparser.parse(death)
            try:
                death = death.strftime("%m/%d/%Y")
            except AttributeError:
                dYear = ""
                death = ""
        elif len(year3) > 6:
            temp = year3[:4]
            if temp.isnumeric():
                death_no_spaces = death.replace(' ', '')
                startIndex = death_no_spaces.find(year3[4:])
                num_spaces_before = death[:startIndex].count(' ')
                adjusted_start_index = startIndex + num_spaces_before
                death = death[:adjusted_start_index].strip()
                death = dateparser.parse(death)
                death = death.strftime("%m/%d/%Y")
                dYear = temp
            else:
                death = ""
        else:
            death = ""
    elif "," in death:
        year3 = death.split(',')[-1]
        year3 = year3.replace(" ", "")
        if len(year3) == 4:
            death = dateparser.parse(death)
            if death:
                dYear = death.strftime("%Y")
                death = death.strftime("%m/%d/%Y")
            else:
                dYear = year3
                death = ""
        elif len(year3) == 2:
            deathYYFlag = True     
            death = dateparser.parse(death)
            dYear = death.strftime("%Y")[-2:]
            death = death.strftime("%m/%d/%Y")[:-2]
        elif len(year3) == 6 or len(year3) == 5:
            dYear = year3[-4:]
            death = death.replace(",", " ")
            death = dateparser.parse(death)
            death = death.strftime("%m/%d/%Y")
        elif len(year3) > 6:
            temp = year3[:4]
            if temp.isnumeric():
                death_no_spaces = death.replace(' ', '')
                startIndex = death_no_spaces.find(year3[4:])
                num_spaces_before = death[:startIndex].count(' ')
                adjusted_start_index = startIndex + num_spaces_before
                death = death[:adjusted_start_index].strip()
                death = dateparser.parse(death)
                death = death.strftime("%m/%d/%Y")
                dYear = temp
            else:
                death = ""
        else:
            death = ""
    elif " " in death:
        year3 = death.split(' ')[-1]
        year3 = year3.replace(" ", "")
        if len(year3) == 4 and (year3[0] == "1" or year3[0] == "2"):
            death = dateparser.parse(death)
            dYear = death.strftime("%Y")
            death = death.strftime("%m/%d/%Y")
        elif len(year3) == 2:
            deathYYFlag = True     
            death = dateparser.parse(death)
            dYear = death.strftime("%Y")[-2:]
            death = death.strftime("%m/%d/%Y")[:-2]
        elif len(year3) == 6 or len(year3) == 5:
            dYear = year3[-4:]
            death = death.replace(",", " ")
            death = dateparser.parse(death)
            death = death.strftime("%m/%d/%Y")
        elif len(year3) > 6:
            temp = year3[:4]
            if temp.isnumeric():
                death_no_spaces = death.replace(' ', '')
                startIndex = death_no_spaces.find(year3[4:])
                num_spaces_before = death[:startIndex].count(' ')
                adjusted_start_index = startIndex + num_spaces_before
                death = death[:adjusted_start_index].strip()
                death = dateparser.parse(death)
                death = death.strftime("%m/%d/%Y")
                dYear = temp
            else:
                death = ""
        else:
            dYear = temp[-1]
            death = ""
    else:
        death = ""
        dYear = ""
    return death, dYear, deathYYFlag


'''
Processes and formats the burial year information. It adjusts the year based on the 
century and war involvement, and reconciles it with other date-related information.

@param value (str) - Extracted raw burial date
@param cent (str) - Value in "19" field, confirms death year is in 1900s and provides year
@param warsFlag (bool) - Flag indicating involvement in a 19th century war

@return buried4Year (str) - The full year of burial, 4 digits
@return buried2Year (str) - The partial year of burial, last 2 digits if full is not 
                            available

@author Mike
'''
def buriedRule(value, cent, warsFlag):
    buried = value.replace("I", "1").replace("-", "/")
    buried4Year = ""
    buried2Year = ""
    if len(cent) == 2:
        buried4Year = "19" + cent
    elif len(cent) == 4:
        buried4Year = cent
    elif len(cent) == 3:
        buried4Year = "19" + cent[-2:]
    elif buried:
        if dateparser.parse(buried, settings={'STRICT_PARSING': True}) != None:
            if buried.count('/') == 2:
                year5 = buried.split('/')[-1]
                if year5 == buried:
                    year5 = buried.split(',')[-1]
                    if year5[0] == " ":
                        year5 = year5[1:]
                    if len(year5) == 4:
                        buried4Year = year5
                    elif len(year5) == 2:
                        buried2Year = year5
                else:
                    if year5[-1:] == " ":
                        year5 = year5[:-1]
                    if len(year5) == 4:
                        buried4Year = year5
                    elif len(year5) == 2:
                        if warsFlag:
                            buried4Year = "19" + year5
                        else:
                            buried2Year = year5
            elif "," in buried:
                year5 = buried.split(',')[-1]
                if len(year5) == 4:
                    buried4Year = year5
                elif len(year5) == 2:
                    if warsFlag:
                        buried4Year = "19" + year5
                    else:
                        buried2Year = year5
            elif " " in buried:
                if buried.count('.') == 1:
                    buried = buried.replace(".", " ")
                year5 = buried.split(' ')
                if year5[-1] == " ":
                    year5 = year5[-2]
                else:
                    year5 = year5[-1]
                if len(year5) == 4:
                    buried4Year = year5
                elif len(year5) == 2:
                    if warsFlag:
                        buried4Year = "19" + year5
                    else:
                        buried2Year = year5
    return buried4Year, buried2Year

'''
Analyzes and formats date-related information based on various rules and conditions. 
It reconciles birth and death dates with other date-related information. Handles conditions
where missing birth and/or death dates and/or years. Attempts to correct 2-digit years
using cent date, else uses buried date, else uses war info, else uses "<" comparison and 
this comparison calls for the activation of warFlag signaling to highlight the row due to 
possible wrong century.

@param finalVals (list) - List where the processed date information will be appended
@param value (str) - Current value being processed (typically related to death date)
@param dob (str) - Raw date of birth
@param buried (str) - Raw date of burial
@param cent (str) -  Value in "19" field, confirms death year is in 1900s and provides year
@param war (str) - War during which the person served, if applicable
@param app (str) - The 4 digit year when the app was sent, used for correcting 2 digit years

@return warFlag (bool) - Flag indicating if a 2-digit year comparison was made to 
                         determine the century

@author Mike
''' 
def dateRule(finalVals, value, dob, buried, cent, war, app):
    warFlag = False
    warsFlag = False
    wars = ["World War 1", "World War 2", "Korean War", "Vietnam War", "Mexican Border War"]
    if war in wars:
        warsFlag = True
    birthYYFlag = False
    deathYYFlag = False
    bYear = ""
    dYear = ""
    buried4Year = ""
    buried2Year = ""
    cent = cent.replace(" ", "")
    appYear = ""
    if "." in app or "," in app:
        app = app.replace(".", ",")
        tempYear = app.split(",")[-1].replace(" ", "")
        while tempYear and not tempYear[-1].isnumeric():
            tempYear = tempYear[:-1]
        if len(tempYear) == 4:
            appYear = tempYear
    elif app:
        tempYear = app
        while tempYear and not tempYear[-1].isnumeric():
            tempYear = tempYear[:-1]
        tempYear = tempYear.split(" ")[-1]
        if len(tempYear) == 4:
            appYear = tempYear
    birth = dob
    if birth:
        if birth.count("/") == 2:
            birth = birth.replace(".", "")
        birth = birth.replace(":", " ").replace("I", "1").replace(".", " ")\
            .replace("&", "").replace("x", "").replace("\n", " ").replace(";", " ")\
            .replace("_", "")
        if "at" in birth.lower():
            birth = birth.lower().split("at")[0]
        while birth and not birth[-1].isnumeric():
            birth = birth[:-1]
        if "Age" in birth:
            match = re.search(r'\b\d{4}\b', birth)
            if match:
                bYear = match.group()
                birth = ""
            else:
                match = re.search(r'\b\d{2}\b', birth)
                if match:
                    birth = ""
        if "born" in birth.lower():
            birth = birth.lower().split("born")[1]
        if len(birth.replace(" ", "")) == 4:
            bYear = birth.replace(" ", "")
            birth = ""
    # temp = birth[:3].replace('7', '/')
    # birth = temp + birth[3:]
    death = value
    if death:
        if death.count("/") == 2:
            death = death.replace(".", "")
        death = death.replace(":", " ").replace("I", "1").replace(".", " ")\
            .replace("&", "").replace("x", "").replace("\n", " ").replace(";", " ")\
            .replace("_", "")
        if death[-1] == " ":
            death = death[:-1]
        if death[-1] == "/":
            death = death[:-1]
        if "-" in birth:
            birth = birth.replace("-", "/")
        if "-" in death:
            death = death.replace("-", "/")
        if death[0] == "/":
            death = death[1:]
        death = death.replace("Found", "").replace("on", "")
        if len(death.replace(" ", "")) == 4:
            dYear = death.replace(" ", "")
            death = ""
        if "death" in death.lower():
            death = death.lower().split("death")[1]
        while death and not death[-1].isnumeric():
            death = death[:-1]
    while buried and not buried[-1].isnumeric():
        buried = buried[:-1]
    # temp2 = death[:3].replace('7', '/')
    # death = temp2 + death[3:]
    try:
        buried4Year, buried2Year = buriedRule(buried.replace("_", ""), cent, warsFlag)
        if buried4Year.count("19") == 2:
            buried4Year = buried4Year.split("19")
            if all(item == "" for item in buried4Year):
                buried4Year = "1919"
            else:
                for x in buried4Year:
                    if x != "":
                        buried4Year = "19" + x
        else:
            if buried4Year[2:] == "19":
                buried4Year = buried4Year[2:] + buried4Year[:2]
            else:
                buried4Year = buried4Year
    except Exception:
        print("Buried didn't parse")
    if birth != "" and death != "":
        if dateparser.parse(birth, settings={'STRICT_PARSING': True}) != None:
            if dateparser.parse(death, settings={'STRICT_PARSING': True}) != None:
                birth, bYear, birthYYFlag = parseBirth(birth, bYear, birthYYFlag)
                death, dYear, deathYYFlag = parseDeath(death, dYear, deathYYFlag)
                if birthYYFlag and deathYYFlag:
                    if cent:
                        if bYear > dYear:
                            bYear = "18" + bYear
                            dYear = "19" + dYear
                        elif bYear < dYear:
                            bYear = "19" + bYear
                            dYear = "19" + dYear
                        birth = birth[:-2] + bYear
                        death = death[:-2] + dYear
                        finalVals.append(birth)
                        finalVals.append(int(bYear))
                        finalVals.append(death)
                        finalVals.append(int(dYear))
                    elif buried4Year:
                        if buried4Year[:2] == "17" and buried4Year[2:] > bYear:
                            bYear = "17" + bYear
                            dYear = buried4Year
                        elif buried4Year[:2] == "18" and buried4Year[2:] < bYear:
                            bYear = "17" + bYear
                            dYear = buried4Year
                        elif buried4Year[:2] == "18" and buried4Year[2:] > bYear:
                            bYear = "18" + bYear
                            dYear = buried4Year
                        elif buried4Year[:2] == "19" and buried4Year[2:] < bYear:
                            bYear = "18" + bYear
                            dYear = buried4Year
                        elif buried4Year[:2] == "19" and buried4Year[2:] > bYear:
                            bYear = "19" + bYear
                            dYear = buried4Year
                        elif buried4Year[:2] == "20" and buried4Year[2:] < bYear:
                            bYear = "19" + bYear
                            dYear = buried4Year
                        else:
                            finalVals.append("")
                            finalVals.append("")
                            finalVals.append("")
                            finalVals.append("")
                            return warFlag
                        birth = birth[:-2] + bYear
                        death = death[:-2] + dYear
                        finalVals.append(birth)
                        finalVals.append(int(bYear))
                        finalVals.append(death)
                        finalVals.append(int(dYear))
                    elif appYear:
                        if appYear[2:] < dYear:
                            if appYear[:2] == "20":
                                dYear = "19" + dYear
                            elif appYear[:2] == "19":
                                dYear = "18" + dYear
                            elif appYear[:2] == "18":
                                dYear = "17" + dYear
                        elif appYear[2:] >= dYear:
                            if appYear[:2] == "20":
                                dYear = "20" + dYear
                            elif appYear[:2] == "19":
                                dYear = "19" + dYear
                            elif appYear[:2] == "18":
                                dYear = "18" + dYear
                            elif appYear[:2] == "17":
                                dYear = "17" + dYear
                        if dYear[2:] < bYear:
                            if dYear[:2] == "20":
                                bYear = "19" + bYear
                            elif dYear[:2] == "19":
                                bYear = "18" + bYear
                            elif dYear[:2] == "18":
                                bYear = "17" + bYear
                            elif dYear[:2] == "17":
                                bYear = "16" + bYear
                        elif dYear[2:] > bYear:
                            if dYear[:2] == "20":
                                bYear = "20" + bYear
                            elif dYear[:2] == "19":
                                bYear = "19" + bYear
                            elif dYear[:2] == "18":
                                bYear = "18" + bYear
                            elif dYear[:2] == "17":
                                bYear = "17" + bYear
                        birth = birth[:-2] + bYear
                        death = death[:-2] + dYear
                        finalVals.append(birth)
                        finalVals.append(int(bYear))
                        finalVals.append(death)
                        finalVals.append(int(dYear))
                    elif war:
                        if war in wars and dYear[0] == "0":
                            bYear = "19" + bYear
                            birth = birth[:-2] + bYear
                            dYear = "20" + dYear
                            death = death[:-2] + dYear
                            finalVals.append(birth)
                            finalVals.append(int(bYear))
                            finalVals.append(death)
                            finalVals.append(int(dYear))
                        elif war in wars and bYear > dYear:
                            bYear = "18" + bYear
                            birth = birth[:-2] + bYear
                            dYear = "19" + dYear
                            death = death[:-2] + dYear
                            finalVals.append(birth)
                            finalVals.append(int(bYear))
                            finalVals.append(death)
                            finalVals.append(int(dYear))
                        elif (war in wars and bYear < dYear):
                            bYear = "19" + bYear
                            birth = birth[:-2] + bYear
                            dYear = "19" + dYear
                            death = death[:-2] + dYear
                            finalVals.append(birth)
                            finalVals.append(int(bYear))
                            finalVals.append(death)
                            finalVals.append(int(dYear))
                        elif war in wars or (war == "Spanish American War" and dYear < "98"):
                            if bYear > dYear:
                                bYear = "18" + bYear
                                birth = birth[:-2] + bYear
                                dYear = "19" + dYear
                                death = death[:-2] + dYear
                                finalVals.append(birth)
                                finalVals.append(int(bYear)) 
                                finalVals.append(death)
                                finalVals.append(int(dYear))
                            elif bYear < dYear:
                                bYear = "19" + bYear
                                birth = birth[:-2] + bYear
                                dYear = "19" + dYear
                                death = death[:-2] + dYear
                                finalVals.append(birth)
                                finalVals.append(int(bYear)) 
                                finalVals.append(death)
                                finalVals.append(int(dYear))
                        else:
                            finalVals.append("")
                            finalVals.append("") 
                            finalVals.append("")
                            finalVals.append("")
                    elif dYear < bYear:
                        bYear = "18" + bYear
                        birth = birth[:-2] + bYear
                        dYear = "19" + dYear
                        death = death[:-2] + dYear
                        finalVals.append(birth)
                        finalVals.append(int(bYear))
                        finalVals.append(death)
                        finalVals.append(int(dYear))
                        warFlag = True
                    elif dYear > bYear:
                        bYear = "19" + bYear
                        birth = birth[:-2] + bYear
                        dYear = "19" + dYear
                        death = death[:-2] + dYear
                        finalVals.append(birth)
                        finalVals.append(int(bYear))
                        finalVals.append(death)
                        finalVals.append(int(dYear))
                        warFlag = True
                    else:
                        finalVals.append("")
                        finalVals.append("") 
                        finalVals.append("")
                        finalVals.append("")
                    return warFlag
                elif not birthYYFlag and deathYYFlag:
                    if bYear[:2] == "19" and bYear[-2:] < dYear:
                        dYear = "19" + dYear
                    elif bYear[:2] == "19" and bYear[-2:] > dYear:
                        dYear = "20" + dYear
                    elif bYear[:2] == "18" and bYear[-2:] < dYear:
                        dYear = "18" + dYear
                    elif bYear[:2] == "18" and bYear[-2:] > dYear: 
                        dYear = "19" + dYear
                    elif bYear[:2] == "17" and bYear[-2:] < dYear:
                        dYear = "17" + dYear
                    elif bYear[:2] == "17" and bYear[-2:] > dYear: 
                        dYear = "18" + dYear
                    death = death[:-2] + dYear
                    if birth:
                        finalVals.append(birth)
                    else:
                        finalVals.append("")
                    if bYear:
                        finalVals.append(int(bYear))
                    else:
                        finalVals.append("")
                    if len(dYear) == 2:
                        dYear = "19" + dYear
                        finalVals.append(death[:-2] + dYear)
                        finalVals.append(int(dYear))
                    elif len(dYear) == 4:
                        finalVals.append(death)
                        finalVals.append(int(dYear))
                    else:
                        finalVals.append("")
                    return warFlag
                elif birthYYFlag and not deathYYFlag:
                    if dYear[:2] == "19" and dYear[2:] < bYear:
                        bYear = "18" + bYear
                    elif dYear[:2] == "19" and dYear[2:] > bYear:
                        bYear = "19" + bYear
                    elif dYear[:2] == "18" and dYear[2:] < bYear:
                        bYear = "19" + bYear
                    elif dYear[:2] == "18" and dYear[2:] > bYear: 
                        bYear = "18" + bYear
                    elif dYear[:2] == "20":
                        bYear = "19" + bYear
                    elif dYear[:2] == "17" and dYear[2:] > dYear:
                        bYear = "17" + bYear
                    elif dYear[:2] == "17" and dYear[2:] < dYear:
                        bYear = "18" + bYear
                    birth = birth[:-2] + bYear
                    finalVals.append(birth)
                    finalVals.append(int(bYear))
                    finalVals.append(death)
                    finalVals.append(int(dYear))
                    return warFlag
                else:
                    finalVals.append(birth)
                    finalVals.append(int(bYear))
                    finalVals.append(death)
                    finalVals.append(int(dYear))
                    return warFlag
            else:
                birth = dateparser.parse(birth)
                bYear = birth.strftime("%Y")
                birth = birth.strftime("%m/%d/%Y")
                death, dYear, deathYYFlag = parseDeath(death, dYear, deathYYFlag)
                if death:
                    if not deathYYFlag:
                        finalVals.append(birth)
                        if bYear:
                            finalVals.append(int(bYear))
                        else:
                            finalVals.append("")
                        finalVals.append(death)
                        if dYear:
                            finalVals.append(int(dYear))
                        else:
                            finalVals.append("")
                        return warFlag
                    else:
                        if bYear[:2] == "19" and bYear[-2:] > dYear:
                            dYear = "20" + dYear
                        elif bYear[:2] == "19" and bYear[-2:] < dYear:
                            dYear = "19" + dYear
                        elif bYear[:2] == "18" and bYear[-2:] > dYear:
                            dYear = "19" + dYear
                        elif bYear[:2] == "18" and bYear[-2:] < dYear: 
                            dYear = "18" + dYear
                        elif bYear[:2] == "17" and bYear[-2:] > dYear:
                            dYear = "18" + dYear
                        elif bYear[:2] == "17" and bYear[-2:] < dYear:
                            dYear = "17" + dYear
                        death = death[:-2] + dYear
                        finalVals.append(birth)
                        if bYear:
                            finalVals.append(int(bYear))
                        else:
                            finalVals.append("")
                        finalVals.append(death)
                        if dYear:
                            finalVals.append(int(dYear))
                        else:
                            finalVals.append("")
                        return warFlag
                elif dYear:
                    finalVals.append(birth)
                    if bYear:
                        finalVals.append(int(bYear))
                    else:
                        finalVals.append("")
                    finalVals.append("")
                    finalVals.append(int(dYear))
                    return warFlag
                elif cent:
                    if cent[2:] > bYear[2:]:
                        bYear = "19" + bYear[2:]
                    elif cent[2:] < bYear[2:]:
                        bYear = "18" + bYear[2:]
                    birth = birth[:-4] + bYear
                    finalVals.append(birth)
                    finalVals.append(int(bYear))
                    finalVals.append("")
                    finalVals.append("")
                    return warFlag
                elif appYear:
                    if bYear[2:] < appYear[2:]:
                        bYear = "19" + bYear[2:]
                    else:
                        bYear = "20" + bYear[2:]
                    birth = birth[:-4] + bYear
                    finalVals.append(birth)
                    finalVals.append(int(bYear))
                    finalVals.append("")
                    finalVals.append(int(appYear))
                else:
                    finalVals.append(birth)
                    if bYear:
                        finalVals.append(int(bYear))
                    else:
                        finalVals.append("")
                    finalVals.append("")
                    finalVals.append("")
                    return warFlag
        elif dateparser.parse(death, settings={'STRICT_PARSING': True}) != None:
            finalVals.append("")
            finalVals.append("")
            death = dateparser.parse(death)
            finalVals.append(death.strftime("%m/%d/%Y"))
            finalVals.append(int(death.strftime("%Y")))
            return warFlag
        else:
            birth, bYear, birthYYFlag = parseBirth(birth, bYear, birthYYFlag)
            death, dYear, deathYYFlag = parseDeath(death, dYear, deathYYFlag)
            if birthYYFlag and deathYYFlag:
                if cent:
                    if bYear > dYear:
                        bYear = "18" + bYear
                        dYear = "19" + dYear
                    elif bYear < dYear:
                        bYear = "19" + bYear
                        dYear = "19" + dYear
                    birth = birth[:-2] + bYear
                    death = death[:-2] + dYear
                    finalVals.append(birth)
                    finalVals.append(int(bYear))
                    finalVals.append(death)
                    finalVals.append(int(dYear))
                    return warFlag
                elif dYear < bYear:
                    finalVals.append("")
                    finalVals.append(int("18" + bYear))
                    finalVals.append("")
                    finalVals.append(int("19" + dYear))
                    warFlag = True
                    return warFlag
                elif dYear > bYear:
                    finalVals.append("")
                    finalVals.append(int("19" + bYear))
                    finalVals.append("")
                    finalVals.append(int("19" + dYear))
                    warFlag = True
                    return warFlag
            elif birthYYFlag and not deathYYFlag:
                if dYear[:2] == "19" and dYear[-2:] < bYear:
                    bYear = "18" + bYear
                elif dYear[:2] == "19" and dYear[-2:] > bYear:
                    bYear = "19" + bYear
                elif dYear[:2] == "18" and dYear[-2:] < bYear:
                    bYear = "19" + bYear
                elif dYear[:2] == "18" and dYear[-2:] > bYear: 
                    bYear = "18" + bYear
                elif dYear[:2] == "20":
                    bYear = "19" + bYear
                elif dYear[:2] == "17" and dYear[-2:] < bYear:
                    bYear = "18" + bYear
                elif dYear[:2] == "17" and dYear[-2:] > bYear:
                    bYear = "17" + bYear
                birth = birth[:-2] + bYear
                finalVals.append(birth)
                if bYear: 
                    finalVals.append(int(bYear))
                else:
                    finalVals.append("")
                finalVals.append(death)
                finalVals.append(int(dYear))
                return warFlag
            elif not birthYYFlag and deathYYFlag:
                if bYear[:2] == "19" and bYear[-2:] < dYear:
                    dYear = "20" + dYear
                elif bYear[:2] == "19" and bYear[-2:] > dYear:
                    dYear = "19" + dYear
                elif bYear[:2] == "18" and bYear[-2:] < dYear:
                    dYear = "19" + dYear
                elif bYear[:2] == "18" and bYear[-2:] > dYear: 
                    dYear = "18" + dYear
                elif bYear[:2] == "17" and bYear[-2:] < dYear:
                    dYear = "18" + dYear
                elif bYear[:2] == "17" and bYear[-2:] > dYear:
                    dYear = "17" + dYear
                death = death[:-2] + dYear
                finalVals.append(birth)
                finalVals.append(int(bYear))
                finalVals.append(death)
                finalVals.append(int(dYear))
                return warFlag
            else:
                finalVals.append("")
                finalVals.append("") 
                finalVals.append("")
                finalVals.append("") 
                return warFlag
    elif birth == "" and death != "":
        death, dYear, deathYYFlag = parseDeath(death, dYear, deathYYFlag)
        if not deathYYFlag:
            if bYear:
                finalVals.append("")
                finalVals.append(int(bYear))
            else:
                finalVals.append("")
                finalVals.append("")
            finalVals.append(death)
            if dYear:
                finalVals.append(int(dYear))
            else:
                finalVals.append("")
        else:
            if bYear:
                if bYear[:2] == "17" and bYear[-2:] < dYear:
                    dYear = "17" + dYear
                elif bYear[:2] == "17" and bYear[-2:] > dYear:
                    dYear = "18" + dYear
                elif bYear[:2] == "18" and bYear[-2:] < dYear:
                    dYear = "18" + dYear
                elif bYear[:2] == "18" and bYear[-2:] > dYear:
                    dYear = "19" + dYear
                elif bYear[:2] == "19" and bYear[-2:] < dYear:
                    dYear = "19" + dYear
                elif bYear[:2] == "19" and bYear[-2:] > dYear:
                    dYear = "20" + dYear
                finalVals.append("")
                finalVals.append(int(bYear))
                death = death[:-2] + dYear
                finalVals.append(death)
                finalVals.append(int(dYear))
            elif cent:
                finalVals.append("")
                finalVals.append("") 
                if len(cent) == 4:
                    dYear = cent
                elif len(cent) == 2:
                    dYear = "19" + dYear
                finalVals.append(death[:-2] + dYear)
                finalVals.append(int(dYear))
            elif war in wars or (war == "Spanish American War" and dYear[:2] < "98"):
                finalVals.append("")
                finalVals.append("") 
                dYear = "19" + dYear
                finalVals.append(death[:-2] + dYear)
                finalVals.append(int(dYear))
            else:
                finalVals.append("")
                finalVals.append("") 
                finalVals.append("")
                finalVals.append("") 
    elif birth != "" and death == "":
        birth, bYear, birthYYFlag = parseBirth(birth, bYear, birthYYFlag)
        if not birthYYFlag:
            if bYear:
                finalVals.append(birth)
                finalVals.append(int(bYear))
            else:
                finalVals.append("")
                finalVals.append("") 
            if dYear != "":
                finalVals.append("")
                finalVals.append(int(dYear))
                return warFlag
            elif buried4Year != "":
                dYear = buried4Year
                finalVals.append("")
                finalVals.append(int(dYear))
                return warFlag
            elif buried2Year != "":
                if bYear[:2] == "17" and bYear[2:] < buried2Year:
                    dYear = "17" + buried2Year
                elif bYear[:2] == "17" and bYear[2:] > buried2Year:
                    dYear = "18" + buried2Year
                elif bYear[:2] == "18" and bYear[2:] < buried2Year:
                    dYear = "18" + buried2Year
                elif bYear[:2] == "18" and bYear[2:] > buried2Year:
                    dYear = "19" + buried2Year
                elif bYear[:2] == "19" and bYear[2:] < buried2Year:
                    dYear = "19" + buried2Year
                elif bYear[:2] == "19" and bYear[2:] > buried2Year:
                    dYear = "20" + buried2Year
                finalVals.append("")
                finalVals.append(int(dYear))
                return warFlag
            elif cent:
                finalVals.append("")
                finalVals.append("") 
                finalVals.append("") 
                finalVals.append(int(cent))
            else:
                finalVals.append("")
                finalVals.append("") 
        elif buried4Year:
            if buried4Year[:2] == "19" and buried4Year[2:] > bYear:
                bYear = "19" + bYear
            elif buried4Year[:2] == "19" and buried4Year[2:] < bYear:
                bYear = "18" + bYear
            elif buried4Year[:2] == "20" and buried4Year[2:] < bYear:
                bYear = "19" + bYear
            elif buried4Year[:2] == "18" and buried4Year[2:] > bYear:
                bYear = "18" + bYear
            elif buried4Year[:2] == "18" and buried4Year[2:] < bYear:
                bYear = "17" + bYear
            birth = birth[:-2] + bYear
            finalVals.append(birth)
            finalVals.append(int(bYear))
            finalVals.append("")
            finalVals.append(int(buried4Year))
        else:
            if dYear:
                if dYear[:2] == "17" and dYear[-2:] > bYear:
                    bYear = "17" + bYear
                elif dYear[:2] == "18" and dYear[-2:] < bYear:
                    bYear = "17" + bYear
                elif dYear[:2] == "18" and dYear[-2:] > bYear:
                    bYear = "18" + bYear
                elif dYear[:2] == "19" and dYear[-2:] < bYear:
                    bYear = "18" + bYear
                elif dYear[:2] == "19" and dYear[-2:] > bYear:
                    bYear = "19" + bYear
                elif dYear[:2] == "20" and dYear[-2:] < bYear:
                    bYear = "19" + bYear
                birth = birth[:-2] + bYear
                finalVals.append(birth)
                finalVals.append(int(bYear))
                finalVals.append("")
                finalVals.append(int(dYear))
                return warFlag
            elif cent.isnumeric():
                if bYear > cent:
                    bYear = "18" + bYear
                else:
                    bYear = "19" + bYear
                birth = birth[:-2] + bYear
                finalVals.append(birth)
                finalVals.append(int(bYear))
            else:
                finalVals.append("")
                finalVals.append("") 
                finalVals.append("")
                finalVals.append("") 
    elif birth == "" and death == "":
        if dYear:
            pass
        elif buried4Year:
            dYear = buried4Year
        elif cent:
            if len(cent) == 4:
                dYear = cent
            elif len(cent) == 2:
                dYear = "19" + cent
        elif buried2Year:
            if war == "Spanish American War" and buried2Year < "98":
                dYear = "19" + buried2Year
            if war in wars:
                if buried2Year > "14":
                    dYear = "19" + buried2Year
                else:
                    dYear = "20" + buried2Year
        if bYear != "" and dYear != "":
            finalVals.append("")
            finalVals.append(int(bYear))
            finalVals.append("")
            finalVals.append(int(dYear))
        elif bYear != "" and dYear == "":
            finalVals.append("")
            finalVals.append(int(bYear))
            finalVals.append("")
            finalVals.append("")
        elif bYear == "" and dYear != "":
            finalVals.append("")
            finalVals.append("") 
            finalVals.append("")
            finalVals.append(int(dYear))
        else:
            finalVals.append("")
            finalVals.append("") 
            finalVals.append("")
            finalVals.append("") 
    return warFlag


'''
Processes and categorizes the war based on the extracted value. It standardizes different 
war names and handles abbreviations and misspellings. It also accounts for brute force 
extraction through the 'world' parameter taken from get_kv_map().

@param value (str) - Extracted raw war name 
@param world (str) - Specific indicator if war name is not in war record field

@return war (str) - The standardized war name or an empty string if no matching category 
                    was found

@author Mike
'''
def warRule(value, world):
    value = value.replace("\n", " ").replace("7", "1").replace("J", "1")
    if value == "" and world != "":
        value = world
    ww1years = ["1914", "1915", "1916", "1917", "1918"]
    war = value.strip()
    identifiedWars = []  
    war = war.replace(".", "").replace("N", "W").replace(" ", "").replace("-", " ")\
             .replace("#", "").replace(",", "").replace("L", "1").replace("I", "1")
    ww1Pattern = re.compile(
        r'WW1|'  # Matches "WW1"
        r'WWI|'  # Matches "WWI"
        r'\b1\b|'  # Matches standalone "1"
        r'W\.?W\.?\s?(1|I|i|l|L|T|ONE)|'  # Matches "WW 1", "WW I", with optional periods and space
        r'WORLD\s*WAR\s*(1|I|i|l|L|T|ONE)',  # Matches "World War 1", "World War I", with optional spaces
        re.IGNORECASE
    ) 
    ww2Pattern = re.compile(
        r'WW2|'  # Matches "WW2"
        r'WWII|'  # Matches "WWII"
        r'\b2\b|'  # Matches standalone "2"
        r'W\.?W\.?\s?(2|II|ii|ll|LL|TT|TWO|11)|'  # Matches "WW 2", "WW II", with optional periods and space
        r'WORLD\s*WAR\s*(2|II|ii|ll|LL|TT|TWO|11)',  # Matches "World War 2", "World War II", with optional spaces
        re.IGNORECASE
    )
    ww1_and_ww2_pattern = re.compile(
        r'(WW|WORLD\s*WAR|WORLD\s*WARS)\s*'  # Starts with "WW" or "World War" with optional spaces
        r'(1|I|i|l|L|T|ONE)'  # First part of the pattern for World War 1
        r'(\s*(and|&)\s*|\sand\s*|\s*&\s*|)'  # Optional connector: "and", "&", with or without spaces
        r'(2|II|ii|ll|LL|TT|TWO|11|WW2|WWII)|'  # Second part of the pattern for World War 2, optional to allow single mentions
        r'WW1andWW11',  # Explicitly matches "WW1andWW11"
        re.IGNORECASE
    )
    korean_war_pattern = re.compile(r'Korea', re.IGNORECASE)
    vietnam_war_pattern = re.compile(r'Vietnam', re.IGNORECASE)
    simple_world_war_pattern = re.compile(r'\bWorld\s*War\b', re.IGNORECASE) 
    if ww1_and_ww2_pattern.search(war):
        identifiedWars.extend(["World War 1", "World War 2"])
    else:
        if ww2Pattern.search(war):
            identifiedWars.append("World War 2")
        elif simple_world_war_pattern.search(war):
            identifiedWars.append("World War 1")
        elif not identifiedWars and ww1Pattern.search(war):
            identifiedWars.append("World War 1")
    if korean_war_pattern.search(war):
        identifiedWars.append("Korean War")
    if vietnam_war_pattern.search(war):
        identifiedWars.append("Vietnam War")
    war = war.replace(".", "").replace("Calv", "").replace("Vols", "")
    if "Civil" in war or "Citil" in war or "Gettysburg" in war \
        or "Fredericksburg" in war or "bull" in war:
        identifiedWars.append("Civil War")
    if "Spanish" in war or "Amer" in war or "American" in war or "SpAm" in war:
        identifiedWars.append("Spanish American War")
    if "Mexican" in war:
        identifiedWars.append("Mexican Border War")
    if "Rebellion" in war:
        identifiedWars.append("War of the Rebellion")
    if "Revolution" in war or "Rev" in war:
        identifiedWars.append("Revolutionary War")
    if "1812" in war:
        identifiedWars.append("War of 1812")
    words = war.split()
    for x in words:
        if x in ww1years:
            war = "World War 1"
            break
    if identifiedWars:
        war = ' and '.join(identifiedWars)
    else:
        war = ""
    if world != "" and war == "":
        war = world
    if war == "World War":
        war = "World War 1"
    if "Army" in war or "US" in war:
        war = ""
    return war


'''
Determines and appends the military branch based on the extracted text. It handles various 
abbreviations and naming conventions.

@param finalVals (list) - A list where the processed military branch will be appended
@param value (str) - The raw branch name
@param war (str) - A net for if war name is caught by the branch key, or misinput on the
                   registration card
                   
@author Mike

'''
def branchRule(finalVals, value, war):
    value = value.replace("\n", " ")
    armys = ["co", "army", "inf", "infantry", "infan", "usa", "med", "cav", "div", \
             "sig", "art", "corps", "corp", "artillery", "army", "q m c "]
    navys = ["hospital", "navy", "naval", "u s n ", "avy", "u s n r ", "u s s ", "usnr", "uss", "usn"]
    guards = ["113", "102d", "114", "44", "181", "250", "112", "national"]
    branch = value
    branch = branch.replace("/", " ").replace(".", " ").replace("th", "").replace("-", "")\
             .replace(", ", " ")
    words = branch.split()
    if war in value and war != "":
        branch = ""
        value = ""
    if "air force" in branch.lower():
        branch = "Air Force"
    elif "marine" in branch.lower():
        branch = "Marine Corps"
    elif branch.lower() in armys:
        branch = "Army"
    elif branch.lower() in navys:
        branch = "Navy"
    elif "u s m c " in branch.lower():
        branch = "Marine Corps" 
    elif branch.lower() in guards:
        branch = "National Guard"
    elif "u s c g " in branch.lower():
        branch = "Coast Guard"
    else:
        for word in words:
            word = word.lower()
            if word in armys:
                branch = "Army"
                break
            elif "air" in word or "aero" in word:
                branch = "Air Force"
                break
            elif word in navys:
                branch = "Navy"
                break
            elif "coast" in word or "USCG" in word:
                branch = "Coast Guard"
                break
            elif word in guards:
                branch = "National Guard"
                break
            else:
                flag1 = False
                flag2 = False
                for x in armys:
                    if x in word:
                        branch = "Army"
                        flag1 = True
                        break
                if flag1:
                    break
                for x in navys:
                    if x in word:
                        branch = "Navy"
                        flag2 = True
                        break
                if flag2:
                    break
                branch = ""
    if value == "N/A":
        value = ""
    finalVals.append(value)   
    finalVals.append(branch)   


'''
Creates a record by extracting and processing information from a document file. It 
handles various data fields, including name, birth, death, military service, and others. 
It calls the specialized functions to handle parsing, cleaning, and standardizing the
extracted text from the veteran cards to be put into the excel sheet.  

@param fileName (str) - The path to the document file
@param id (int) - The id of the file, which is assigned to the record in Excel
@param cemetery (str) - The name of the cemetery where the record is from

@return finalVals (list) - A list of processed values for completed data fields of the 
                           current file
@return flag (bool) - A flag indicated the size type of the document, used for redaction
@return warFlag (bool) - A flag indicating if "<" comparison was made during date parsing,
                         used for highlighting the row for a second look
@return nameCoords (list) - Bounding box coordinates for "NAME" key
@return serialCoords (list) - Bounding box coordinates for "SERIAL" key
@return kvs (defaultdict(list)) - Dictionary containing key-value pairs
    
@author Mike
'''
def createRecord(fileName, id, cemetery):
    world = ""
    war = ""
    buried = ""
    cent = ""
    app = ""
    dob = ""
    badWar = ["n/a", "yes", "not listed", "age", "unknown", "peacetime", "pt"]
    nameCoords = None
    serialCoords = None
    warFlag = False
    finalVals = []
    pageReader = PyPDF2.PdfReader(open(fileName, 'rb'))
    page = pageReader.pages[0]
    pdfWriter = PyPDF2.PdfWriter()
    pdfWriter.add_page(page)
    with open("temp.pdf", 'wb') as output_pdf:
        pdfWriter.write(output_pdf)
    documentResult = analyzeDocument("temp.pdf", id)
    kvs, nameCoords, serialCoords, world = extract_key_value_pairs(documentResult)
    print_kvs(kvs)  
    keys = ["", "NAME", "WAR RECORD", "BORN", "19", "BURIED", "DATE OF DEATH", "WAR RECORD", "BRANCH OF SERVICE" , "IN"]
    flag3 = False
    for x in keys:
        try:
            value = search_value(kvs, x)[0]
        except TypeError:
            value = "" 
        if x == "NAME":
            try:
                nameRule(finalVals, value)
            except Exception:
                finalVals.append("")
                finalVals.append("")
                finalVals.append("")
                finalVals.append("")
        elif x == "BORN":
            dob = value
        elif x == "DATE OF DEATH":
            try:
                app = search_value(kvs, "Application")[0]
            except Exception:
                pass
            try:
                warFlag = dateRule(finalVals, value, dob, buried, cent, war, app)
            except Exception:
                while len(finalVals) < 8:
                    finalVals.append("")
        elif x == "WAR RECORD" and flag3:
            if value.lower() in badWar:
                value = ""
            finalVals.append(value)
            finalVals.append(war)
            flag3 = False
        elif x == "WAR RECORD":
            war = warRule(value, world)
            war = war.strip().strip('and').strip()
            flag3 = True
        elif x == "BRANCH OF SERVICE":
            try:
                branchRule(finalVals, value, war)
            except Exception:
                finalVals.append("")
                finalVals.append("")
        elif x == "IN":
            if fuzz.partial_ratio(value.lower(), cemetery.lower()) > 80:
                finalVals.append(cemetery)
            else:
                value = value.replace("The ", "").replace("Cemetery", "").replace(".", "")
                finalVals.append(value)
        elif x == "19":
            if value:
                try:
                    tempCent = value.replace(",", "").replace(".", "").replace(":", "").replace(";", "").replace("/", "")\
                        .replace(" ", "").replace("\n", "").replace("_", "").replace("in", "").replace("...", "")
                    while tempCent and not tempCent[-1].isnumeric():
                        tempCent = tempCent[:-1]
                    if tempCent.count("19") == 2:
                        tempCent = tempCent.split("19")
                        if all(item == "" for item in tempCent):
                            cent = "1919"
                        else:
                            for x in tempCent:
                                if x != "":
                                    cent = "19" + x
                    else:
                        if tempCent[2:] == "19":
                            cent = tempCent[2:] + tempCent[:2]
                        else:
                            cent = tempCent
                except IndexError:
                    pass
            else:
                pass
        elif x == "BURIED":
            buried = value
    return finalVals, warFlag, nameCoords, serialCoords, kvs


'''
Creates a temporary record for processing 'A' and 'B' cards. It's similar to 
'createRecord' but tailored for handling these specific card types. It extracts 
and processes information from a document file and calls other functions for 
specific fields.

@param fileName (str) - The path to the document file
@param val (str) - A value indicating the type of card ('A' or 'B')
@param int (int) - The ID to be assigned to the record
@param cemetery (str) - The name of the cemetery to associate with the record

@return finalVals (list) - A list of processed values for different data fields 
                           of the record
@return flag (bool) - A flag indicating if a specific condition (e.g., special 
                      handling or a specific record type) was encountered during processing
@return warFlag (bool) - A flag indicating if war-related data was processed
@return kvs (defaultdict(list)) - Dictionary containing key-value pairs

@author Mike
'''
def tempRecord(fileName, val, id, cemetery):
    print("Performing Temp", id, val.upper())
    world = ""
    war = ""
    buried = ""
    cent = ""
    app = ""
    dob = ""
    badWar = ["n/a", "yes", "not listed", "age", "unknown", "peacetime", "pt"]
    nameCoords = None
    serialCoords = None
    warFlag = False
    finalVals = []
    pageReader = PyPDF2.PdfReader(open(fileName, 'rb'))
    page = pageReader.pages[0]
    pdfWriter = PyPDF2.PdfWriter()
    pdfWriter.add_page(page)
    with open("temp.pdf", 'wb') as output_pdf:
        pdfWriter.write(output_pdf)
    documentResult  = analyzeDocument("temp.pdf", id)
    kvs, nameCoords, serialCoords, world = extract_key_value_pairs(documentResult)
    print_kvs(kvs)
    keys = ["", "NAME", "WAR RECORD", "BORN", "19", "DATE OF DEATH", "WAR RECORD", "BRANCH OF SERVICE" , "IN"]
    flag3 = False
    for x in keys:
        try:
            value = search_value(kvs, x)[0]
        except TypeError:
            value = "" 
        if x == "NAME":
            try:
                nameRule(finalVals, value)
            except Exception:
                finalVals.append("")
                finalVals.append("")
                finalVals.append("")
                finalVals.append("")
        elif x == "BORN":
            dob = value
        elif x == "DATE OF DEATH":
            try:
                app = search_value(kvs, "Application")[0]
            except Exception:
                pass
            try:
                warFlag = dateRule(finalVals, value, dob, buried, cent, war, app)
            except Exception:
                while len(finalVals) < 8:
                    finalVals.append("")
        elif x == "WAR RECORD" and flag3:
            if value.lower() in badWar:
                value = ""
            finalVals.append(value)
            finalVals.append(war)
            flag3 = False
        elif x == "WAR RECORD":
            war = warRule(value, world)
            war = war.strip().strip('and').strip()
            flag3 = True
        elif x == "BRANCH OF SERVICE":
            try:
                branchRule(finalVals, value, war)
            except Exception:
                finalVals.append("")
                finalVals.append("")
        elif x == "IN":
            if fuzz.partial_ratio(value.lower(), cemetery.lower()) > 80:
                finalVals.append(cemetery)
            else:
                value = value.replace("The ", "").replace("Cemetery", "").replace(".", "")
                finalVals.append(value)
        elif x == "19":
            if value:
                try:
                    tempCent = value.replace(",", "").replace(".", "").replace(":", "").replace(";", "").replace("/", "")\
                        .replace(" ", "").replace("\n", "").replace("_", "").replace("in", "").replace("...", "")
                    while tempCent and not tempCent[-1].isnumeric():
                        tempCent = tempCent[:-1]
                    if tempCent.count("19") == 2:
                        tempCent = tempCent.split("19")
                        if all(item == "" for item in tempCent):
                            cent = "1919"
                        else:
                            for x in tempCent:
                                if x != "":
                                    cent = "19" + x
                    else:
                        if tempCent[2:] == "19":
                            cent = tempCent[2:] + tempCent[:2]
                        else:
                            cent = tempCent
                except IndexError:
                    pass
            else:
                pass
        elif x == "BURIED":
            buried = value
    return finalVals, warFlag, nameCoords, serialCoords, kvs


'''
Merges two sets of record values, typically from 'A' and 'B' cards, into a 
single record. It compares and chooses the most appropriate value for each 
field and handles row highlighting in the output based on specific conditions.

@param vals1 (list) - The first set of record values
@param vals2 (list) - The second set of record values
@param rowIndex (int) - The row index in the output where the merged record will be placed
@param id (int) - The ID assigned to the merged record
@param warFlag (bool) - A flag indicating if war-related data was processed

@author Mike
'''
def mergeRecords(vals1, vals2, rowIndex, id, warFlag, letter):
    counter = 1
    def length(item):
        return len(str(item))
    mergedArray = [max(a, b, key=length) for a, b in zip(vals1, vals2)] 
    worksheet.cell(row=rowIndex, column=counter, value=id)
    counter += 1
    for x in mergedArray:
        worksheet.cell(row=rowIndex, column=counter, value=x)
        counter += 1
    highlightColors = {
        "mergedNoIssues": PatternFill(start_color="E3963E", end_color="E3963E", fill_type="solid"),
        "mergedAdjusted": PatternFill(start_color="00FFB9", end_color="00FFB9", fill_type="solid"),
        "mergedNoDOD": PatternFill(start_color="00C6FF", end_color="00C6FF", fill_type="solid"),
        "mergedCemeteryMismatch": PatternFill(start_color="FC80AC", end_color="FC80AC", fill_type="solid"),
        "mergedLastNameMismatch": PatternFill(start_color="D2042D", end_color="D2042D", fill_type="solid"),
        "mergedNameBug": PatternFill(start_color="0B8C36", end_color="0B8C36", fill_type="solid"),
        "badDate": PatternFill(start_color="99CC00", end_color="99CC00", fill_type="solid")}
    requiredColors = []
    requiredColors.append(highlightColors["mergedNoIssues"])
    if warFlag:
        requiredColors.append(highlightColors["mergedAdjusted"])
    if (worksheet[f'{"I"}{rowIndex}'].value == ""):
        requiredColors.append(highlightColors["mergedNoDOD"])
    if worksheet[f'{"G"}{rowIndex}'].value:
        if str(worksheet[f'{"G"}{rowIndex}'].value)[0] != "1":
            requiredColors.append(highlightColors["badDate"])
    if worksheet[f'{"I"}{rowIndex}'].value:
        if str(worksheet[f'{"I"}{rowIndex}'].value)[0] != "1":
            if str(worksheet[f'{"I"}{rowIndex}'].value)[0] == "2":
                if str(worksheet[f'{"I"}{rowIndex}'].value)[1] == "0":
                    if str(worksheet[f'{"I"}{rowIndex}'].value)[2:] > "23":
                        requiredColors.append(highlightColors["badDate"]) 
                else:
                    requiredColors.append(highlightColors["badDate"]) 
            else:
                requiredColors.append(highlightColors["badDate"]) 
    if (worksheet[f'{"N"}{rowIndex}'].value) != cemetery:
        requiredColors.append(highlightColors["mergedCemeteryMismatch"])
    if worksheet[f'{"B"}{rowIndex}'].value[0] != letter:
        requiredColors.append(highlightColors["mergedLastNameMismatch"])
    if (worksheet[f'{"B"}{rowIndex}'].value) == (worksheet[f'{"C"}{rowIndex}'].value):
        requiredColors.append(highlightColors["mergedNameBug"])
    if requiredColors:
        numColors = len(requiredColors)
        cols_per_color = max(1, (14 - 2) // numColors)
        for colIndex in range(2, 16 + 1):
            if colIndex == 15:
                continue
            colorIndex = (colIndex - 2) // cols_per_color
            colorIndex = min(colorIndex, numColors - 1) 
            cell = worksheet.cell(row=rowIndex, column=colIndex)
            cell.fill = requiredColors[colorIndex]
        cell = worksheet.cell(row=rowIndex, column=16)
        cell.fill = requiredColors[-1]
        

'''
Finds the next empty row in the worksheet to begin processing at that index.

@param worksheet - The worksheet object being operated on

@return worksheet.max_row + 1 (int) - The row index of the next empty row in the worksheet

@author Mike
'''  
def find_next_empty_row(worksheet):
    for row in worksheet.iter_rows(min_row=3000, min_col=1, max_col=1):
        if row[0].value is None:
            return row[0].row
    return worksheet.max_row + 1 


'''
The main function that controls files sent to be processed as well as 
the data that is put into the excel sheet. It sets up the current working 
directory, the current sheet in the spreadsheet, and loops through all the 
images in a set letter in a set cemetery. This function is the primary 
controller for processing, saving, and handling of records.

Main functions that controls files sent to be processed as well as 
the data that is put into the excel sheet. Sets up the current working 
directory as well as the current sheet in the excel spreadsheet. Loops
through all the images in a set Letter in a set Cemetery, this is for 
damage control. 

The loop finds the last record recorded in the excel sheet
and then starts indexing from there. Calls createRecord is normal card, 
calls tempRecord and mergeRecord for A and B cards.

calls redactImage for every card processed. Places the record ID and a 
hyperlink to the redacted image in the excel sheet during processing.
Saves worksheet after every image is done processing to reduce loss of 
data upon errors.

Controls highlighting of entire row of record upon different conditions 
that are caught by the program. 

@author Mike
'''
def main(singleFlag, singleLetter):
    global cemSet
    global miscSet
    global jewishSet
    networkFolder = r"\\ucclerk\pgmdoc\Veterans"
    os.chdir(networkFolder)
    cemeterys = []
    for x in os.listdir(r"Cemetery"):
        cemeterys.append(x)
    miscs = []
    for x in os.listdir(r"Cemetery\Misc"):
        miscs.append(x)
    jewishs = []
    for x in os.listdir(r"Cemetery\Jewish"):
        jewishs.append(x)
    cemSet = set(cemeterys)
    miscSet = set(miscs)
    jewishSet = set(jewishs)
    workbook = openpyxl.load_workbook('Veterans.xlsx')
    global cemetery
    cemetery = "Extra" # Change this to continue running through cemeteries
    cemPath = os.path.join(networkFolder, fr"Cemetery\{cemetery}")
    letter = singleLetter # Change this to continue running through the current cemetery
    namePath = letter 
    namePath = os.path.join(cemPath, namePath)
    pathA = ""
    rowIndex = 2
    global worksheet
    worksheet = workbook[cemetery]
    warFlag = False
    pdfFiles = sorted(os.listdir(namePath))
    initialID = 1
    breakFlag = False
    for y in range(len(pdfFiles)):
        try:
            warFlag = False
            tempFlag = False
            tempFlag2 = False
            filePath = os.path.join(namePath, pdfFiles[y])
            rowIndex = find_next_empty_row(worksheet)
            try:
                id = worksheet[f'{"A"}{rowIndex-1}'].value + 1
            except TypeError:
                id = initialID
            if "output" in pdfFiles[y] or "redacted" in pdfFiles[y]:
                continue
            else:
                string = pdfFiles[y][:-4]
                string = string.split(letter) 
                string = string[-1].lstrip('0')
                if "a" not in string and "b" not in string:
                    if id != int(string.replace("a", "").replace("b", "")):
                        continue
                    vals, warFlag, nameCoords, serialCoords, kvs = createRecord(filePath, id, cemetery)
                    redactedFile = redact(filePath, cemetery, letter, nameCoords, serialCoords)
                    worksheet.cell(row=rowIndex, column=15).value = "PDF Image"
                    worksheet.cell(row=rowIndex, column=15).font = Font(underline="single", color="0563C1")
                    worksheet.cell(row=rowIndex, column=15).hyperlink = redactedFile
                    counter = 1
                    worksheet.cell(row=rowIndex, column=counter, value=id)
                    counter += 1
                    for x in vals:
                        worksheet.cell(row=rowIndex, column=counter, value=x)
                        counter += 1
                    highlightColors = {
                        "warFlag": PatternFill(start_color="899499", end_color="899499", fill_type="solid"),
                        "cemeteryMismatch": PatternFill(start_color="CF9FFF", end_color="CF9FFF", fill_type="solid"),
                        "noDOD": PatternFill(start_color="A7C7E7", end_color="A7C7E7", fill_type="solid"),
                        "lastNameMismatch": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
                        "shortFirstName": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
                        "badWarName": PatternFill(start_color="F5CBA7", end_color="F5CBA7", fill_type="solid"),
                        "badDate": PatternFill(start_color="99CC00", end_color="99CC00", fill_type="solid")}
                    requiredColors = []
                    if warFlag:
                        requiredColors.append(highlightColors["warFlag"])
                    if (worksheet[f'{"N"}{rowIndex}'].value) != cemetery:
                        requiredColors.append(highlightColors["cemeteryMismatch"])
                    if (worksheet[f'{"I"}{rowIndex}'].value) == "":
                        requiredColors.append(highlightColors["noDOD"])
                    if worksheet[f'{"G"}{rowIndex}'].value:
                        if str(worksheet[f'{"G"}{rowIndex}'].value)[0] != "1":
                            requiredColors.append(highlightColors["badDate"])
                    if worksheet[f'{"I"}{rowIndex}'].value:
                        if str(worksheet[f'{"I"}{rowIndex}'].value)[0] != "1":
                            if str(worksheet[f'{"I"}{rowIndex}'].value)[0] == "2":
                                if str(worksheet[f'{"I"}{rowIndex}'].value)[1] == "0":
                                    if str(worksheet[f'{"I"}{rowIndex}'].value)[2:] > "23":
                                        requiredColors.append(highlightColors["badDate"]) 
                                else:
                                    requiredColors.append(highlightColors["badDate"]) 
                            else:
                                requiredColors.append(highlightColors["badDate"]) 
                    if (worksheet[f'{"J"}{rowIndex}'].value) != "" and (worksheet[f'{"K"}{rowIndex}'].value) == ""\
                    or (worksheet[f'{"L"}{rowIndex}'].value) != "" and (worksheet[f'{"M"}{rowIndex}'].value) == "":
                        if worksheet[f'{"J"}{rowIndex}'].value != "Regular Service" or\
                           worksheet[f'{"J"}{rowIndex}'].value != "Peacetime":
                               requiredColors.append(highlightColors["badWarName"])
                    try:
                        if (worksheet[f'B{rowIndex}'].value)[0] != letter:
                            if (worksheet[f'C{rowIndex}'].value)[0] == letter:
                                tempLname = worksheet[f'B{rowIndex}'].value
                                tempFname = worksheet[f'C{rowIndex}'].value
                                worksheet[f'B{rowIndex}'].value = tempFname
                                worksheet[f'C{rowIndex}'].value = tempLname
                            else:
                                tempFlag = True
                        if len((worksheet[f'C{rowIndex}'].value)) < 3:
                            if len((worksheet[f'D{rowIndex}'].value)) >= 3:
                                tempLname = worksheet[f'C{rowIndex}'].value
                                tempSuffix = worksheet[f'D{rowIndex}'].value
                                worksheet[f'C{rowIndex}'].value = tempSuffix
                                worksheet[f'D{rowIndex}'].value = tempLname + "."
                                tempFlag2 = True
                        if tempFlag and tempFlag2:
                            tempLname = worksheet[f'B{rowIndex}'].value
                            tempFname = worksheet[f'C{rowIndex}'].value
                            worksheet[f'B{rowIndex}'].value = tempFname
                            worksheet[f'C{rowIndex}'].value = tempLname
                        if (worksheet[f'B{rowIndex}'].value)[0] != letter or len((worksheet[f'C{rowIndex}'].value)) < 3:
                            requiredColors.append(highlightColors["lastNameMismatch"])
                        if worksheet[f'C{rowIndex}'].value[-1].isupper():
                            worksheet[f'D{rowIndex}'].value = worksheet[f'C{rowIndex}'].value[-1] + "."
                            worksheet[f'C{rowIndex}'].value = worksheet[f'C{rowIndex}'].value[:-1]
                    except IndexError:
                        requiredColors.append(highlightColors["lastNameMismatch"])
                    two_uppercase_pattern = re.compile(r'.*[A-Z].*[A-Z].*')
                    if two_uppercase_pattern.match(worksheet[f'C{rowIndex}'].value):
                        requiredColors.append(highlightColors["lastNameMismatch"])
                    if requiredColors:
                        numColors = len(requiredColors)
                        cols_per_color = max(1, (14 - 2) // numColors)  
                        for colIndex in range(2, 16 + 1):
                            if colIndex == 15:
                                continue
                            colorIndex = (colIndex - 2) // cols_per_color
                            colorIndex = min(colorIndex, numColors - 1) 
                            cell = worksheet.cell(row=rowIndex, column=colIndex)
                            cell.fill = requiredColors[colorIndex]
                        cell = worksheet.cell(row=rowIndex, column=16)
                        cell.fill = requiredColors[-1]
                    id += 1
                    rowIndex += 1
                else:
                    if id != int(string.replace("a", "").replace("b", "")):
                        continue
                    else:
                        if "a" in string:
                            if (filePath.replace("a.pdf", "") in pdfFiles):
                                continue
                            pathA = filePath
                            vals1, warFlag, nameCoords, serialCoords, kvs = tempRecord(filePath, "a", id, cemetery)
                            redactedFile = redact(filePath, cemetery, letter, nameCoords, serialCoords)
                        if "b" in string:
                            if (filePath.replace("b.pdf", "") in pdfFiles):
                                continue
                            vals2, warFlagB, nameCoords, serialCoords, kvs = tempRecord(filePath, "b", id, cemetery)
                            if not warFlag or not warFlagB:
                                warFlag = False
                            else:
                                warFlag = True
                            redactedFile = redact(filePath, cemetery, letter, nameCoords, serialCoords)
                            mergeRecords(vals1, vals2, rowIndex, id, warFlag, letter)
                            mergeImages(pathA, filePath, cemetery, letter)
                            link_text = "PDF Image"
                            worksheet.cell(row=rowIndex, column=15).value = link_text
                            worksheet.cell(row=rowIndex, column=15).font = Font(underline="single", color="0563C1")
                            worksheet.cell(row=rowIndex, column=15).hyperlink = redactedFile.replace("b redacted.pdf", " redacted.pdf")
                            id += 1
                            rowIndex += 1
                            if singleFlag:
                                breakFlag = True
        except Exception as e:
            errorTraceback = traceback.format_exc()
            print(errorTraceback)
            print(f"An error occurred: {e}")
            errorMessage = f"SKIPPED DUE TO ERROR : {errorTraceback}"
            worksheet.cell(row=rowIndex, column=1, value=id)
            worksheet.cell(row=rowIndex, column=2, value=errorMessage)
            highlightColor = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for colIndex in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=rowIndex, column=colIndex)
                cell.fill = highlightColor
            error_file_path = fr'Errors/{cemetery}{letter}{str(id).zfill(5)} Error.txt' 
            with open(error_file_path, 'a') as error_file:
                error_file.write(f'{kvs} \n\n {errorTraceback}')
            id += 1
            rowIndex += 1
        extension1 = str(id-1).zfill(5)
        extension2 = ""
        if "a.pdf" in filePath:
            extension1 = str(id).zfill(5)
            extension2 = "a"
        elif "b.pdf" in filePath:
            extension1 = str(id-1).zfill(5)
            extension2 = "b"
        logFilePath = fr'Logs/{cemetery}{letter}{extension1}{extension2} Extracted.txt' 
        with open(logFilePath, 'w', encoding='utf-8') as logFile:
            logFile.write("----Key-value pairs found in document----\n")
            for key, values in kvs.items():
                for value in values:
                    logFile.write(f'{key}: {value}\n')
            logFile.write("\n")
        workbook.save('Veterans.xlsx')
        if breakFlag:
            break

if __name__ == "__main__":
    global letter
    letter = "A"
    main(False, letter)