from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, \
     QHBoxLayout, QPushButton, QLineEdit, QLabel, QGroupBox, QFormLayout, \
     QComboBox, QMessageBox
from PyQt6.QtGui import QPixmap, QFont, QIcon
from PyQt6.QtCore import Qt, QThread, pyqtSignal
import os
import openpyxl
from openpyxl.styles import PatternFill, Font
import traceback
import sys
import microsoftOCR
import qdarktheme


'''
Dark Mode QMessageBox Button Styling
'''
darkB = ("""
    QPushButton {
        background-color: #0078D4; 
        color: #E4E7EB;           
        border: 1px solid #8A8A8A; 
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                    stop:0 #00A2FF, stop:1 #002D59);
    }
    QPushButton:hover {
        background-color: #669df2; 
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                    stop:0 #80CFFF, stop:1 #004080);
    }
    QPushButton:pressed {
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                    stop:0 #004080, stop:1 #001B3D);
    }
    QPushButton:disabled {
        background-color: #1A1A1C; 
        border: 1px solid #3B3B3B;
        color: #3B3B3B;   
    }
""")

darkNB = ("""
    QPushButton {
        color: #E4E7EB;          
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                    stop:0 #00A2FF, stop:1 #002D59); 
    }
    QPushButton:hover {
        background-color: #669df2; 
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                    stop:0 #80CFFF, stop:1 #004080);
    }
""")

'''
Light Mode QMessageBox Button Styling
'''
lightB = ("""
    QPushButton {
        background-color: #70c5ff;  
        color: #111111;            
        border: 1px solid #111111; 
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #9EDFFF, stop:1 #1A8FE3);
    }
    QPushButton:hover {
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #70C5FF, stop:1 #0078D4);
    }
    QPushButton:pressed {
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                    stop:0 #0078D4, stop:1 #004080);
    }
    QPushButton:disabled {
        background-color: #DBDBDB; 
        color: #686868; 
        border: 1px solid #686868;
    }
""")

'''
Dark Mode Styling
'''
dark = qdarktheme.load_stylesheet(
            theme="dark",
            custom_colors=
            {
                "[dark]": 
                {
                    "primary": "#0078D4",
                    "border": "#8A8A8A",
                    "primary>button.hoverBackground": "#2B456E",
                    "background>list": "#3F4042",
                    "background>popup": "#303136",
                }
            },
        ) + """
            QMessageBox {
                background-color: #303135;
            }
            QMessageBox QLabel {
                color: #E4E7EB;
            }
        """

'''
Light Mode Styling
'''
light = qdarktheme.load_stylesheet(
            theme="light",
            custom_colors=
            {
                "[light]": 
                {
                    "foreground": "#111111",
                    "border": "#111111",
                    "background": "#f0f0f0",
                    "primary>button.hoverBackground": "#adcaf7",
                    "background>list": "#FFFFFF",
                    "background>popup": "#cfcfd1",
                }
            },
        ) + """
            QMessageBox {
                background-color: #d4d4d4;
            }
            QMessageBox QLabel {
                color: #111111;
            }
        """

class Worker(QThread):
    idSignal = pyqtSignal(str)  # Signal to emit IDs
    kvsSignal = pyqtSignal(str) # Signal to emit KVS
    errorSignal = pyqtSignal(str, bool) # Signal to emit Error Message
    imageSignal = pyqtSignal(str) # Signal to emit Image Path
    popupSignal = pyqtSignal(str, str) # Signal to update app status 
    finishedSignal = pyqtSignal(str, str)
    
    
    '''
    The main processing logic that runs when the "Run" button is pressed and all the required
    fields are filled out. 

    @param self - the main window
    @param singleFlag - flag that indicates if main function is being called 
                        by the duplicates utility program
    @param singleCem - cemetery name variable for if main function is being 
                    called by the duplicates utility program
    @param singleLetter - folder letter variable for if main function is being 
                        called by the duplicates utility program

    @author Mike
    '''
    def __init__(self, singleFlag, singleCem, singleLetter):
        super().__init__()
        self.singleFlag = singleFlag
        self.singleCem = singleCem
        self.singleLetter = singleLetter.upper()
        self.paused = False
        self.stopped = False

    
    '''
    The main processing logic that runs when the "Run" button is pressed and all the required
    fields are filled out. 

    @param self - the main window

    @author Mike
    '''
    def run(self):
        try:
            global cemSet
            global miscSet
            global jewishSet
            networkFolder = r"\\ucclerk\pgmdoc\Veterans"
            os.chdir(networkFolder)
            cemeterys = []
            for x in os.listdir(r"Cemetery"):
                cemeterys.append(x)
            miscs = []
            for x in os.listdir(r"Cemetery\Misc"):
                miscs.append(x)
            jewishs = []
            for x in os.listdir(r"Cemetery\Jewish"):
                jewishs.append(x)
            cemSet = set(cemeterys)
            miscSet = set(miscs)
            jewishSet = set(jewishs)
            workbook = openpyxl.load_workbook('Veterans.xlsx')
            cemetery = self.singleCem 
            cemPath = os.path.join(networkFolder, fr"Cemetery\{cemetery}")
            letter = self.singleLetter 
            namePath = letter 
            namePath = os.path.join(cemPath, namePath)
            pathA = ""
            warFlag = False
            breakFlag = False
            rowIndex = 2
            global worksheet
        except Exception:
            error = traceback.format_exc()
            self.errorSignal.emit(error[-775:], True)
            return
        try:
            worksheet = workbook[cemetery]
        except Exception:
            self.popupSignal.emit('Critical', f"Cemetery '{cemetery}' not found in Veterans.xslx.")
            return
        try:
            pdfFiles = sorted(os.listdir(namePath))
        except Exception:
            self.popupSignal.emit('Critical', f"Letter '{letter}' not found in '{cemetery}' folder.")
            return
        for y in range(len(pdfFiles)):
            try:
                warFlag = False
                filePath = os.path.join(namePath, pdfFiles[y])
                rowIndex = microsoftOCR.find_next_empty_row(worksheet)
                try:
                    id = worksheet[f'{"A"}{rowIndex-1}'].value + 1
                except Exception:
                    id = int(pdfFiles[0][:-4].split(letter)[-1].lstrip('0'))
                while self.paused:
                    self.sleep(1)
                string = pdfFiles[y][:-4]
                string = string.split(letter) 
                string = string[-1].lstrip('0')
                if "a" not in string and "b" not in string:
                    if id != int(string.replace("a", "").replace("b", "")):
                        if self.stopped:
                            break
                        continue
                    if self.stopped:
                            break
                    self.idSignal.emit(str(id))
                    self.kvsSignal.emit("")
                    self.errorSignal.emit("", False)
                    self.imageSignal.emit("")
                    vals, warFlag, nameCoords, serialCoords, printedKVS, kinLast = microsoftOCR.createRecord(filePath, id, cemetery, "")
                    self.kvsSignal.emit(printedKVS)
                    redactedFile = microsoftOCR.redact(filePath, cemetery, letter, nameCoords, serialCoords)
                    self.imageSignal.emit(r"\\ucclerk\pgmdoc\Veterans\temp.png")
                    worksheet.cell(row= rowIndex, column= 15).value = "PDF Image"
                    worksheet.cell(row= rowIndex, column= 15).font = Font(underline= "single", color= "0563C1")
                    worksheet.cell(row= rowIndex, column= 15).hyperlink = redactedFile
                    counter = 1
                    worksheet.cell(row= rowIndex, column= counter, value= id)
                    counter += 1
                    for x in vals:
                        worksheet.cell(row= rowIndex, column= counter, value= x)
                        counter += 1
                    microsoftOCR.highlightSingle(worksheet, cemetery, letter, warFlag, rowIndex, kinLast)
                    id += 1
                    rowIndex += 1
                else:
                    if id != int(string.replace("a", "").replace("b", "")):
                        if self.stopped:
                            break
                        continue
                    if self.stopped:
                        break
                    if "a" in string:
                        if (filePath.replace("a.pdf", "") in pdfFiles):
                            continue
                        pathA = filePath
                        self.idSignal.emit(f"{id} A")
                        self.kvsSignal.emit("")
                        self.errorSignal.emit("", False)
                        self.imageSignal.emit("")
                        vals1, warFlag, nameCoords, serialCoords, printedKVS, kinLast = microsoftOCR.createRecord(filePath, id, cemetery, "A")
                        self.kvsSignal.emit(printedKVS)
                        redactedFile = microsoftOCR.redact(filePath, cemetery, letter, nameCoords, serialCoords)
                        self.imageSignal.emit(r"\\ucclerk\pgmdoc\Veterans\temp.png")
                    if "b" in string:
                        if (filePath.replace("b.pdf", "") in pdfFiles):
                            continue
                        self.idSignal.emit(f"{id} B")
                        self.kvsSignal.emit("")
                        self.errorSignal.emit("", False)
                        self.imageSignal.emit("")
                        vals2, warFlagB, nameCoords, serialCoords, printedKVS, kinLast = microsoftOCR.createRecord(filePath, id, cemetery, "B")
                        self.kvsSignal.emit(printedKVS)
                        if not warFlag or not warFlagB:
                            warFlag = False
                        else:
                            warFlag = True
                        redactedFile = microsoftOCR.redact(filePath, cemetery, letter, nameCoords, serialCoords)
                        self.imageSignal.emit(r"\\ucclerk\pgmdoc\Veterans\temp.png")
                        microsoftOCR.mergeRecords(worksheet, vals1, vals2, rowIndex, id, warFlag, cemetery, letter)
                        microsoftOCR.mergeImages(pathA, filePath, cemetery, letter)
                        linkText = "PDF Image"
                        worksheet.cell(row=rowIndex, column=15).value = linkText
                        worksheet.cell(row=rowIndex, column=15).font = Font(underline="single", color="0563C1")
                        worksheet.cell(row=rowIndex, column=15).hyperlink = redactedFile.replace("b redacted.pdf", " redacted.pdf")
                        id += 1
                        rowIndex += 1
                        if self.singleFlag:
                            breakFlag = True
            except Exception as e:
                errorTraceback = traceback.format_exc()
                print(errorTraceback)
                print(f"An error occurred: {e}")
                errorMessage = f"SKIPPED DUE TO ERROR : {errorTraceback}"
                worksheet.cell(row= rowIndex, column= 1, value= id)
                worksheet.cell(row= rowIndex, column= 2, value= errorMessage)
                highlightColor = PatternFill(start_color= "FF0000", end_color= "FF0000", fill_type= "solid")
                for colIndex in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row= rowIndex, column= colIndex)
                    cell.fill = highlightColor
                errorFilePath = fr'Errors/{cemetery}{letter}{str(id).zfill(5)} Error.txt' 
                with open(errorFilePath, 'a') as errorFile:
                    errorFile.write(f'{printedKVS} \n\n {errorTraceback}')
                    self.errorSignal.emit(errorTraceback, False)
                id += 1
                rowIndex += 1
            extension1 = str(id-1).zfill(5)
            extension2 = ""
            if "a.pdf" in filePath:
                extension1 = str(id).zfill(5)
                extension2 = "a"
            elif "b.pdf" in filePath:
                extension1 = str(id-1).zfill(5)
                extension2 = "b"
            logFilePath = fr'Logs/{cemetery}{letter}{extension1}{extension2} Extracted.txt' 
            with open(logFilePath, 'w', encoding='utf-8') as logFile:
                logFile.write(printedKVS)
            workbook.save('Veterans.xlsx')
            if breakFlag:
                break
        self.finishedSignal.emit(cemetery, letter)
        return
   
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Veteran Extraction")
        self.setGeometry(500, 35, 950, 150)
        self.worker = None  
        self.mainLayout()


    '''
    Saves display mode to display_mode.txt

    @param self - the main window
    @param mode - the name of the display mode that is being used, and therefore saved

    @author Mike
    '''
    def saveDisplayMode(self, mode):
        parentPath = os.path.dirname(os.getcwd())
        with open(f"{parentPath}/_internal/veteranData/display_mode.txt", "w") as file:
            file.write(mode)


    '''
    Loads display mode name from display_mode.txt

    @param self - the main window
    
    @return string - the name of the display mode that was saved

    @author Mike
    '''
    def loadDisplayMode(self):
        parentPath = os.path.dirname(os.getcwd())
        try:
            with open(f"{parentPath}/_internal/veteranData/display_mode.txt", "r") as file:
                return file.read().strip()
        except FileNotFoundError:
            return "Light"
        
        
    '''
    Updates the display mode anytime the selection is changed within the app through
    the display mode selection box. Saves the mode selection to a local .txt file.
    Manually adjusts the styling of the three buttons on the app.

    @param window - the main window
    @param runButton - the QPushButton widget for the run button
    @param pauseButton - the QPushButton widget for the pause button
    @param stopButton - the QPushButton widget for the stop button
    @param displayMode - the name of the display mode selected

    @author Mike
    '''
    def changeDisplayStyle(self, window, runButton, pauseButton, stopButton, displayMode):
        if displayMode == "Dark":
            window.setStyleSheet(dark)
        else:
            window.setStyleSheet(light)
        self.saveDisplayMode(displayMode)
        self.updateButtonStyle(runButton, pauseButton, stopButton, displayMode)
    
    
    '''
    Updates the button styling anytime the selection is changed within the app through
    the display mode selection box. 

    @param window - the main window 
    @param runButton - the QPushButton widget for the run button
    @param pauseButton - the QPushButton widget for the pause button
    @param stopButton - the QPushButton widget for the stop button
    @param displayMode - the name of the display mode selected

    @author Mike
    '''
    def updateButtonStyle(self, runButton, pauseButton, stopButton, displayMode):
        if displayMode == "Light":
            runButton.setStyleSheet(lightB)
            pauseButton.setStyleSheet(lightB)
            stopButton.setStyleSheet(lightB)
        else:
            runButton.setStyleSheet(darkB)
            pauseButton.setStyleSheet(darkB)
            stopButton.setStyleSheet(darkB)
        
        
    '''
    Creates the main window UI. Places all the widgets in their respective places
    and adjusts their size and other properties.

    @param self - the main window

    @author Mike
    '''        
    def mainLayout(self):
        centralWidget = QWidget(self)
        centralWidget.setObjectName("mainCentralWidget")
        self.setCentralWidget(centralWidget)
        layout1 = QVBoxLayout(centralWidget)
        
        # Create the top container
        topContainer = QWidget()
        topLayout = QHBoxLayout()
        logoLabel = QLabel() 
        parentPath = os.path.dirname(os.getcwd())
        logoImage = QPixmap(f"{parentPath}/_internal/veteranData/ucLogo.png") 
        logoLabel.setPixmap(logoImage)
        programName = QLabel("Union County Clerk's Office\n\n       Veteran Extraction")
        font = QFont()
        font.setBold(True)
        font.setPointSize(26)
        programName.setFont(font)
        topLayout.addWidget(logoLabel)
        topLayout.addWidget(programName)
        topContainer.setLayout(topLayout)

        # Create a group box for the middle top left
        middleTopLeftGroupBox = QGroupBox("Parameters")
        middleTopLeftGroupBox.setFixedSize(375, 375)
        middleTopLeftLayout = QFormLayout()
        self.cemeteryBox = QLineEdit()
        self.cemeteryBox.setFixedWidth(140)
        self.letterBox = QLineEdit()
        self.letterBox.setFixedWidth(30)
        font = QFont("Monterchi Sans Book", 8)
        noteLabel1 = QLabel("Note: This is the cemetery name based on its folder.")
        noteLabel2 = QLabel("Note: This is the last name letter to start processing.")
        noteLabel1.setFont(font)
        noteLabel2.setFont(font)
        self.displayModeBox = QComboBox()
        self.displayModeBox.setFixedWidth(75)
        modes = ["Light", "Dark"]
        self.displayModeBox.addItems(modes)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftLayout.addRow("Display Mode: ", self.displayModeBox)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftLayout.addRow("Cemetery :   ", self.cemeteryBox)
        middleTopLeftLayout.addRow(None, noteLabel1)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftLayout.addRow("Letter :", self.letterBox)
        middleTopLeftLayout.addRow(None, noteLabel2)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftGroupBox.setLayout(middleTopLeftLayout)
        
        # Create a group box for the middle top right
        middleTopRightGroupBox = QGroupBox("Results")
        middleTopRightGroupBox.setFixedSize(535, 375)
        middleTopRightLayout = QFormLayout()
        self.idLabel = QLabel("")
        self.kvsLabel = QLabel("")
        self.errorLabel = QLabel("")
        self.imageLabel = QLabel("")
        middleTopRightLayout.addRow(" ", None)
        middleTopRightLayout.addRow("Current ID :", self.idLabel)
        middleTopRightLayout.addRow(" ", None)
        middleTopRightLayout.addRow("Key-Value Pairs :", self.kvsLabel)
        middleTopRightLayout.addRow(" ", None)
        middleTopRightLayout.addRow("Error :", self.errorLabel)
        middleTopRightLayout.addRow(" ", None)
        middleTopRightGroupBox.setLayout(middleTopRightLayout)

        # Create container for middle top left and top right group boxes
        middleTopContainer = QWidget()
        middleTopLayout = QHBoxLayout()
        middleTopLayout.addWidget(middleTopLeftGroupBox)
        middleTopLayout.addWidget(middleTopRightGroupBox)
        middleTopContainer.setLayout(middleTopLayout)
        
        # Create a group box for the middle bottom
        middleBottomGroupBox = QGroupBox("Image")
        middleBottomGroupBox.setFixedSize(500, 287)
        middleBottomLayout = QFormLayout()
        self.imageLabel = QLabel("")
        middleBottomLayout.addRow(None, self.imageLabel)
        middleBottomGroupBox.setLayout(middleBottomLayout)
        
        # Create a container for the middle bottom group box
        middleBottomContainer = QWidget()
        middleBottomLayout = QHBoxLayout()
        middleBottomLayout.addWidget(middleBottomGroupBox)
        middleBottomContainer.setLayout(middleBottomLayout)
        
        # Create the bottom container with a button
        bottomContainer = QWidget()
        bottomLayout = QHBoxLayout()
        self.runButton = QPushButton("Run Code")
        self.runButton.clicked.connect(self.startProcessing)
        self.runButton.setFixedWidth(150)
        self.pauseButton = QPushButton("Pause Code")
        self.pauseButton.clicked.connect(self.togglePauseResume)
        self.pauseButton.setFixedWidth(150)
        self.stopButton = QPushButton("Stop Code")
        self.stopButton.setFixedWidth(225)
        self.stopButton.clicked.connect(self.stopProcessing)
        self.status = QLabel("              Status :  Idle")
        self.status.setFixedWidth(150)
        self.pauseButton.setDisabled(True)
        self.stopButton.setDisabled(True)
        bottomLayout.addWidget(self.runButton)
        bottomLayout.addWidget(self.status)
        bottomLayout.addWidget(self.pauseButton)
        self.displayModeBox.currentTextChanged.connect(lambda: self.changeDisplayStyle(self, self.runButton, self.pauseButton, self.stopButton, self.displayModeBox.currentText()))
        self.displayModeBox.setCurrentIndex(modes.index(self.loadDisplayMode()))
        self.updateButtonStyle(self.runButton, self.pauseButton, self.stopButton, self.loadDisplayMode())
        bottomContainer.setLayout(bottomLayout)
        # Create a main layout to arrange the containers
        mainLayout = QVBoxLayout()
        mainLayout.addWidget(topContainer)
        mainLayout.addWidget(middleTopContainer)
        mainLayout.addWidget(middleBottomContainer)
        mainLayout.addWidget(bottomContainer)
        mainLayout.addWidget(self.stopButton, 0, alignment= Qt.AlignmentFlag.AlignCenter)
        layout1.addLayout(mainLayout)
    
    
    '''
    Calls a QMessageBox from the main run function from the Worker class. Passes in data needed for popup message.
    Resets the window status. 

    @param self - the main window
    @param type - the icon needed for the popup
    @param text - the text needed for the popup

    @author Mike
    '''        
    def updateStatus(self, type, text):
        if type == "Critical":
            msgBox = QMessageBox(QMessageBox.Icon.Critical, 'Error', text, QMessageBox.StandardButton.Ok, window)
            if self.loadDisplayMode() == "Light":
                msgBox.setStyleSheet(lightB)
            else:
                msgBox.setStyleSheet(darkB)
            msgBox.exec()
        elif type == "Warning":
            msgBox = QMessageBox(QMessageBox.Icon.Warning, 'Warning', text, QMessageBox.StandardButton.Ok, window)
            if self.loadDisplayMode() == "Light":
                msgBox.setStyleSheet(lightB)
            else:
                msgBox.setStyleSheet(darkB)
            msgBox.exec()
        elif type == "Info":
            msgBox = QMessageBox(QMessageBox.Icon.Information, 'Information', text, QMessageBox.StandardButton.Ok, window)
            if self.loadDisplayMode() == "Light":
                msgBox.setStyleSheet(lightB)
            else:
                msgBox.setStyleSheet(darkB)
            msgBox.exec()
        self.runButton.setDisabled(False)
        self.status.setText("              Status :  Idle")
    
    
    '''
    Function that updates the image display with the current image being processed

    @param self - the main window
    @param filePath - the path for the current image being processed

    @author Mike
    '''  
    def updateImage(self, filePath):
        pixmap = QPixmap(filePath)
        self.imageLabel.setPixmap(pixmap.scaled(420, 490, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)) 
    
    
    '''
    Function that updates the current ID number display for the current image being processed

    @param self - the main window
    @param id - the current ID number being processed

    @author Mike
    '''      
    def updateID(self, id):
        self.idLabel.setText(id)
    
    
    '''
    Function that updates the raw extracted text display for the current image being processed

    @param self - the main window
    @param printedKVS - the raw extracted text from the current image

    @author Mike
    '''  
    def updateKVS(self, printedKVS):
        self.kvsLabel.setText(f"{printedKVS}")  
    
    
    '''
    Function that updates error display for the current image being processed

    @param self - the main window
    @param printedKVS - the error message that occured when processing the current image

    @author Mike
    '''  
    def updateError(self, errorMsg, resetFlag):
        if resetFlag:
            self.runButton.setDisabled(False)
            self.pauseButton.setDisabled(True)
            self.stopButton.setDisabled(True)
            self.status.setText("              Status :  Idle")
        self.errorLabel.setText(f"{errorMsg}")  

    
    
    '''
    Function that updates resets the window when processing loop has finished

    @param self - the main window
    @param cemetery - the name of the cemetery that is being processed
    @param letter - the letter folder that is being processed

    @author Mike
    '''  
    def updateFinished(self, cemetery, letter):
        self.updateStatus('Info', f"{cemetery} letter {letter} has finished processing. Please enter the next letter.") 
        self.runButton.setDisabled(False)
        self.pauseButton.setDisabled(True)
        self.stopButton.setDisabled(True)
        self.status.setText("              Status :  Idle")

        
    '''
    Function that sets up the Worker and the signals that the Worker will emit to 
    call the display update funcitons.

    @param self - the main window

    @author Mike
    '''        
    def startProcessing(self):
        if self.cemeteryBox.text() == "":
            QMessageBox.warning(window, 'Missing Info', "Cemetery field not filled out.")
            return
        if self.letterBox.text() == "":
            QMessageBox.warning(window, 'Missing Info', "Letter field not filled out.")
            return
        self.worker = Worker(False, self.cemeteryBox.text(), self.letterBox.text())
        self.worker.idSignal.connect(lambda id: self.updateID(id))
        self.worker.kvsSignal.connect(lambda printedKVS: self.updateKVS(printedKVS))
        self.worker.errorSignal.connect(lambda errorMsg, resetFlag: self.updateError(errorMsg, resetFlag))
        self.worker.imageSignal.connect(lambda imagePath: self.updateImage(imagePath))
        self.worker.popupSignal.connect(lambda type, text: self.updateStatus(type, text))
        self.worker.finishedSignal.connect(lambda cemetery, letter: self.updateFinished(cemetery, letter))
        self.worker.start()
        self.status.setText("          Status :  Running...")
        self.runButton.setDisabled(True)
        self.pauseButton.setDisabled(False)
        self.stopButton.setDisabled(False)
    
    
    '''
    Function that changes the text of the Pause/Resume button and updates the 
    app status accordingly. 

    @param self - the main window, used to apply its properties to QMessageBox

    @author Mike
    ''' 
    def togglePauseResume(self):
        if self.worker.paused:
            self.stopButton.setDisabled(False)
            self.worker.paused = False
            self.pauseButton.setText("Pause")
            self.status.setText("          Status :  Running...")
        else:
            self.stopButton.setDisabled(True)
            self.worker.paused = True
            self.pauseButton.setText("Resume")
            self.status.setText("           Status :  Paused" )
    
    
    '''
    Function that stops the Worker and updates the status and buttons on the app.

    @param self - the main window, used to apply its properties to QMessageBox

    @author Mike
    ''' 
    def stopProcessing(self):
        self.worker.stopped = True
        self.status.setText("          Status :  Stopping...")
        self.pauseButton.setDisabled(True)
        
        
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.setWindowTitle("Veteran Extraction")
    parentPath = os.path.dirname(os.getcwd())
    window.setWindowIcon(QIcon(f"{parentPath}/_internal/veteranData/veteranLogo.png"))
    msgBox = QMessageBox(QMessageBox.Icon.Information, 'Instructions', "If code is running, please press \"Stop Code\" before closing the application.", QMessageBox.StandardButton.Ok, window)
    window.show()
    if window.loadDisplayMode() == "Light":
        app.setStyleSheet(light)
        msgBox.setStyleSheet(lightB)
    else:
        app.setStyleSheet(dark)
        msgBox.setStyleSheet(darkB)
    msgBox.exec()
    sys.exit(app.exec())