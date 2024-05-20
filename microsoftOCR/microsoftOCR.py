import re
import openpyxl 
from azure.core.credentials import AzureKeyCredential
from azure.ai.formrecognizer import DocumentAnalysisClient
from openpyxl.styles import Font
from collections import defaultdict
import traceback
import os
import PyPDF2
import fitz
import cv2
import time
from reportlab.pdfgen import canvas
from PIL import Image
from nameparser import HumanName
from nameparser.config import CONSTANTS
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz
import sys
sys.path.append(r'C:\workspace\veterans\microsoftOCR')
import nameRule
import dateRule
import warRule
import branchRule


'''
Redacts sensitive information (e.g., social security or ID numbers) from a registration 
card by adding a black box over the specified area. The method of redaction varies based 
on the card format. The first page of the PDF is saved as a PNG, and the black box is 
added to the PNG. A new PDF is then created using the redacted PNG for the first page and 
the original second page.

@param file (str) - Path to the file to be processed
@param cemetery (str) - Name of the cemetery, used for categorizing and saving the 
                        file in the correct directory
@param letter (str) - Last name letter, used for categorizing and saving the 
                      file in the correct directory
@param nameCoords (list) - Bounding box coordinates for "NAME" key
@param serialCoords (list) - Bounding box coordinates for "SERIAL" key

@return new_pdf_file (str) - Path to the newly created redacted PDF file

@author Mike
'''
def redact(filePath, cemetery, letter, nameCoords, serialCoords):
    imageFile  = "temp.png"
    img = cv2.imread(imageFile)
    if serialCoords:
        pt1 = (serialCoords[0][0]-75, serialCoords[0][1]-350)
        pt3 = (3525, serialCoords[1][1]+50)
    else:
        pt1 = (nameCoords[0][0]+2050, nameCoords[0][1]-350)
        pt3 = (3525, nameCoords[1][1]+50)
    cv2.rectangle(img, pt1, pt3, (0, 0, 0), thickness=cv2.FILLED)
    cv2.imwrite(imageFile, img)
    image = Image.open(imageFile)
    pdfFile = "temp.pdf"
    c = canvas.Canvas(pdfFile, pagesize=(image.width, image.height))
    c.drawImage(imageFile, 0, 0, width=image.width, height=image.height)
    c.save()
    image.close()

    pdfDocument = fitz.open(filePath)
    fullLocation = r"\\ucclerk\pgmdoc\Veterans\Cemetery - Redacted"
    redactedLocation = f'{cemetery} - Redacted'
    fullLocation = os.path.join(fullLocation, redactedLocation)
    fullLocation = os.path.join(fullLocation, letter)
    redacted_pdf_document  = fitz.open(pdfFile)
    new_pdf_document  = fitz.open()
    new_pdf_document.insert_pdf(redacted_pdf_document, from_page=0, to_page=0)
    new_pdf_document.insert_pdf(pdfDocument, from_page=1, to_page=1)
    tempName = filePath[-9:]
    if "a" in tempName or "b" in tempName:
        tempName = filePath[-10:]
    new_pdf_file = f'{fullLocation}\\{cemetery}{letter}{tempName.replace(".pdf", "")} redacted.pdf'
    new_pdf_document.save(new_pdf_file)
    new_pdf_document.close()
    redacted_pdf_document.close()
    pdfDocument.close()
    return new_pdf_file    


'''
Merges 'A' and 'B' redacted files into a single combined file, showing both versions 
of the card, both redacted.

@param pathA (str) - File path for file A (first document to be merged)
@param pathB (str) - File path for file B (second document to be merged)
@param cemetery (str) - Name of the cemetery, used for categorizing and saving the 
                        merged file in the correct directory
@param letter (str) - Last name letter, used for categorizing and saving the 
                      merged file in the correct directory

@author Mike
'''
def mergeImages(pathA, pathB, cemetery, letter):
    fullLocation = r"\\ucclerk\pgmdoc\Veterans\Cemetery - Redacted"
    redactedLocation = f'{cemetery} - Redacted'
    fullLocation = os.path.join(fullLocation, redactedLocation)
    fullLocation = os.path.join(fullLocation, letter)
    fileName = pathA.split(letter)
    merger = PyPDF2.PdfMerger()
    merger.append(pathA)
    merger.append(pathB)
    # with open(f'\\ucclerk\pgmdoc\Veterans\\{cemetery}{letter}{fileName[2].replace("a.pdf", ".pdf")}', 'wb') as out_pdf:
    #     merger.write(out_pdf)
    merger2 = PyPDF2.PdfMerger()
    merger2.append(f'{fullLocation}\\{cemetery}{letter}{fileName[-1].replace(".pdf", "")} redacted.pdf')
    merger2.append(f'{fullLocation}\\{cemetery}{letter}{fileName[-1].replace("a.pdf", "b")} redacted.pdf')
    with open(f'{fullLocation}\\{cemetery}{letter}{fileName[-1].replace("a.pdf", "")} redacted.pdf', 'wb') as out_pdf:
        merger2.write(out_pdf)
        
        
'''
Extracts field matches from the specified file using Microsoft Azure Textract, which 
performs an AI analysis to identify and extract text from the form. Uses a custom built
model that was trained to understand the specific form layout and variable placement per
field needed from extraction.

@param filePath (str) - File path for the file being processed 
@param id (int) - ID of the file, used for referencing the specific file that is 
                  currently being processed, printed in the terminal

@return result (AnalyzeResult) - The result text extraction of the fields that were found 
                                 by the custom model.

@author Mike
'''
def analyzeDocument(filePath, id, suffix):
    endpoint = os.environ["VISION_ENDPOINT"]
    key = os.environ["VISION_KEY"]
    document_analysis_client = DocumentAnalysisClient(endpoint=endpoint, credential=AzureKeyCredential(key))
    with open(filePath, "rb") as file:
        imgTest = file.read()
        bytesTest = bytearray(imgTest)
        print('\nImage loaded', f"{id} {suffix}")
    poller = document_analysis_client.begin_analyze_document("Test2.4n", document=bytesTest)
    result = poller.result()
    return result


'''
Creates a dictionary storing the values for each field extracted from the document. 
Associates the values with the specific key names that are used throughout the code.

@param result (AnalyzeResult) - The result text extraction of the fields that were found 
                                by the custom model.

@return kvs (defaultdict(list)) - The result text extraction of the fields that were found 
                                  by the custom model.
@return nameCoords (list) - Bounding box coordinates for "NAME" key
@return serialCoords (list) - Bounding box coordinates for "SERIAL" key
@return war (string) - Secondary check for if war record info is placed outside of war 
                       record field.

@author Mike
'''
def extract_key_value_pairs(result):
    kvs = defaultdict(list)
    war = ""
    nameCoords = []
    serialCoords = []
    for idx, invoice in enumerate(result.documents):
        name = invoice.fields.get("Name")
        if name.content != None:
            kvs["NAME"].append(name.content)
        kin = invoice.fields.get("Kin")
        if kin.content != None:
            kvs["KIN"].append(kin.content)
        nameCoord = invoice.fields.get("nameCoord")
        if nameCoord.content != None:
            for region in nameCoord.bounding_regions:
                polygon = region.polygon
                nameCoords.append((int(polygon[0].x), int(polygon[0].y)))
                nameCoords.append((int(polygon[2].x), int(polygon[2].y)))
        serial = invoice.fields.get("Serial No.")
        if serial:
            for region in serial.bounding_regions:
                polygon = region.polygon
                serialCoords.append((int(polygon[0].x), int(polygon[0].y)))
                serialCoords.append((int(polygon[2].x), int(polygon[2].y)))
        born = invoice.fields.get("Born")
        if born.content != None:
            kvs["BORN"].append(born.content)
        death = invoice.fields.get("Date of Death")
        if death.content != None:
            kvs["DATE OF DEATH"].append(death.content)
        buried = invoice.fields.get("Buried")
        if buried.content != None:
            kvs["BURIED"].append(buried.content)
        century = invoice.fields.get("19")
        if century.content != None:
            kvs["19"].append(century.content)
        cemetery = invoice.fields.get("IN")
        if cemetery.content != None:
            kvs["IN"].append(cemetery.content)
        warRecord = invoice.fields.get("War Record")
        if warRecord.content != None:
            kvs["WAR RECORD"].append(warRecord.content)
        branch = invoice.fields.get("Branch of Service")
        if branch.content != None:
            kvs["BRANCH OF SERVICE"].append(branch.content)
        app = invoice.fields.get("App")
        if app.content != None:
            kvs["APPLICATION"].append(app.content)
        world = invoice.fields.get("War")
        if world.content != None:
            if "War" in world.content:
                kvs[""].append(world.content)
                war = world.content
        # care = invoice.fields.get("Care Assigned")
        # if care.content != None:
        #     extracted.append(True)
    return kvs, nameCoords, serialCoords, war


'''
Prints the key-value pairs in a readable format. Useful for debugging 
or presenting the extracted data.

@param kvs (defaultdict(list)) - Dictionary containing the key-value pairs to be printed

@author Mike
'''
def print_kvs(kvs):
    result = "----Key-value pairs found in document----\n"
    for key, value in kvs.items():
        result += f"{key} : {value}\n"
    print(result)
    return result
    
    
'''
Searches for and retrieves the value associated with a specified key from the dictionary 
of key-value pairs. Compares strings for equality.

@param kvs (defaultdict(list)) - Dictionary containing key-value pairs
@param search_key (str) - The key for which the value needs to be retrieved for

@return value (str or None) - The value associated with the search_key, if found. 
                              Otherwise, None

@author Mike
''' 
def search_value(kvs, searchKey):
    for key, value in kvs.items():
        if key.strip().upper() == searchKey.upper():
            return value


'''
Searches for and retrieves the value associated with a specified key from the dictionary 
of key-value pairs. Uses REGEX instead of comparing strings.

@param kvs (defaultdict(list)) - Dictionary containing key-value pairs
@param search_key (str) - The key for which the value needs to be retrieved for

@return value (str or None) - The value associated with the search_key, if found. 
                            Otherwise, None

@author Mike
''' 
def search_value_regex(kvs, searchKey):
    searchPattern = r'\b' + re.escape(searchKey) + r'\b'
    for key, value in kvs.items():
        if re.search(searchPattern, key, re.IGNORECASE):
            return value


'''
Creates a record by extracting and processing information from a document file. It 
handles various data fields, including name, birth, death, military service, and others. 
It calls the specialized functions to handle parsing, cleaning, and standardizing the
extracted text from the veteran cards to be put into the excel sheet.  

@param fileName (str) - The path to the document file
@param id (int) - The id of the file, which is assigned to the record in Excel
@param cemetery (str) - The name of the cemetery where the record is from

@return finalVals (list) - A list of processed values for completed data fields of the 
                           current file
@return flag (bool) - A flag indicated the size type of the document, used for redaction
@return warFlag (bool) - A flag indicating if "<" comparison was made during date parsing,
                         used for highlighting the row for a second look
@return nameCoords (list) - Bounding box coordinates for "NAME" key
@return serialCoords (list) - Bounding box coordinates for "SERIAL" key
@return kvs (defaultdict(list)) - Dictionary containing key-value pairs
    
@author Mike
'''
def createRecord(fileName, id, cemetery, suffix):
    world = ""
    war = ""
    buried = ""
    cent = ""
    app = ""
    dob = ""
    kinLast = ""
    badWar = ["n/a", "yes", "not listed", "age", "unknown", "peacetime", "pt", "honorable", \
        "not shown", "no date shown", "Peace time", "none", "Peace Time"]
    nameCoords = None
    serialCoords = None
    warFlag = False
    finalVals = []
    attempt = False
    # pageReader = PyPDF2.PdfReader(open(fileName, 'rb'))
    # page = pageReader.pages[0]
    # pdfWriter = PyPDF2.PdfWriter()
    # pdfWriter.add_page(page)
    # with open("temp.pdf", 'wb') as output_pdf:
    #     pdfWriter.write(output_pdf)
    while attempt == False:
        try:
            doc = fitz.open(fileName)
            page = doc.load_page(0)
            pix = page.get_pixmap(matrix=fitz.Matrix(600/72, 600/72))
            pix.save('temp.png')
            attempt = True
        except Exception as e:
            print(f"Failed to process document: {e}")
            time.sleep(5)  
    documentResult = analyzeDocument("temp.png", id, suffix)
    kvs, nameCoords, serialCoords, world = extract_key_value_pairs(documentResult)
    printedKVS = print_kvs(kvs)  
    keys = ["", "NAME", "KIN", "WAR RECORD", "BORN", "19", "BURIED", "DATE OF DEATH", "WAR RECORD", "BRANCH OF SERVICE" , "IN"]
    flag3 = False
    for x in keys:
        try:
            value = search_value(kvs, x)[0]
        except TypeError:
            value = "" 
        if x == "NAME":
            try:
                nameRule.nameRule(finalVals, value)
            except Exception:
                finalVals.append("")
                finalVals.append("")
                finalVals.append("")
                finalVals.append("")
        elif x == "KIN":
            value = value.replace("NAME", "").replace("Name", "").replace("name", "")\
                .replace("\n", " ").replace(".", " ")
            if value:
                try:
                    CONSTANTS.force_mixed_case_capitalization = True
                    name = HumanName(value)
                    flag = True
                    if name.last.isupper():
                        name.last = name.last[0] + name.last[1:].lower()
                    tempLast = name.last[0]
                    for letter in name.last[1:]:
                        if letter.isupper() and flag:
                            tempLast += " "
                            flag = False
                        elif letter == " ":
                            flag = False
                        tempLast += letter
                    name.last = tempLast
                    name.capitalize(force=True)
                    lastName =  name.last.replace("' ", "'")
                    lastName = lastName[0].upper() + lastName[1:]
                    kinLast = re.sub(r"[^a-zA-Z' ]", '', lastName)
                    kinLast = kinLast.replace("Mom" , "").replace("Daughter" , "")\
                        .replace("Dad" , "").replace("Sister" , "").replace("Brother" , "")\
                        .replace("Mother" , "").replace("Father" , "").replace("Son" , "")\
                        .replace("Wife" , "").replace("Husband" , "")
                    print("KIN LAST: " + kinLast)
                except Exception:
                    pass
        elif x == "BORN":
            dob = value
        elif x == "DATE OF DEATH":
            try:
                app = search_value(kvs, "Application")[0]
            except Exception:
                pass
            try:
                warFlag = dateRule.dateRule(finalVals, value, dob, buried, cent, war, app)
            except Exception:
                while len(finalVals) < 8:
                    finalVals.append("")
        elif x == "WAR RECORD" and flag3:
            for x in badWar:
                if x in value.lower():
                    value = ""
            pattern = re.compile(re.escape("recor"), re.IGNORECASE)
            value = pattern.sub(lambda x: "", value)
            pattern = re.compile(re.escape("record"), re.IGNORECASE)
            value = pattern.sub(lambda x: "", value)
            finalVals.append(value)
            finalVals.append(war)
            flag3 = False
        elif x == "WAR RECORD":
            war = warRule.warRule(value, world)
            war = war.strip().strip('and').strip()
            flag3 = True
        elif x == "BRANCH OF SERVICE":
            try:
                branchRule.branchRule(finalVals, value, war)
            except Exception:
                finalVals.append("")
                finalVals.append("")
        elif x == "IN":
            if fuzz.partial_ratio(value.lower(), cemetery.lower()) > 80:
                finalVals.append(cemetery)
            else:
                value = value.replace("The ", "").replace("Cemetery", "").replace(".", "")
                finalVals.append(value)
        elif x == "19":
            if value:
                try:
                    tempCent = value.replace(",", "").replace(".", "").replace(":", "").replace(";", "").replace("/", "")\
                        .replace(" ", "").replace("\n", "").replace("_", "").replace("in", "").replace("...", "")
                    while tempCent and not tempCent[-1].isnumeric():
                        tempCent = tempCent[:-1]
                    if tempCent.count("19") == 2:
                        tempCent = tempCent.split("19")
                        if all(item == "" for item in tempCent):
                            cent = "1919"
                        else:
                            for x in tempCent:
                                if x != "":
                                    cent = "19" + x
                    else:
                        if tempCent[2:] == "19":
                            cent = tempCent[2:] + tempCent[:2]
                        else:
                            cent = tempCent
                except IndexError:
                    pass
            else:
                pass
        elif x == "BURIED":
            buried = value
    return finalVals, warFlag, nameCoords, serialCoords, printedKVS, kinLast


'''
Merges two sets of record values, typically from 'A' and 'B' cards, into a 
single record. It compares and chooses the most appropriate value for each 
field and handles row highlighting in the output based on specific conditions.

@param vals1 (list) - The first set of record values
@param vals2 (list) - The second set of record values
@param rowIndex (int) - The row index in the output where the merged record will be placed
@param id (int) - The ID assigned to the merged record
@param warFlag (bool) - A flag indicating if war-related data was processed

@author Mike
'''
def mergeRecords(worksheet, vals1, vals2, rowIndex, id, warFlag, cemetery, letter):
    counter = 1
    def length(item):
        return len(str(item))
    mergedArray = [max(a, b, key=length) for a, b in zip(vals1, vals2)] 
    worksheet.cell(row=rowIndex, column=counter, value=id)
    counter += 1
    for x in mergedArray:
        worksheet.cell(row=rowIndex, column=counter, value=x)
        counter += 1
    highlightColors = {
        "mergedNoIssues": PatternFill(start_color="E3963E", end_color="E3963E", fill_type="solid"),
        "mergedAdjusted": PatternFill(start_color="00FFB9", end_color="00FFB9", fill_type="solid"),
        "mergedNoDOD": PatternFill(start_color="00C6FF", end_color="00C6FF", fill_type="solid"),
        "mergedCemeteryMismatch": PatternFill(start_color="FC80AC", end_color="FC80AC", fill_type="solid"),
        "mergedLastNameMismatch": PatternFill(start_color="D2042D", end_color="D2042D", fill_type="solid"),
        "mergedNameBug": PatternFill(start_color="0B8C36", end_color="0B8C36", fill_type="solid"),
        "badDate": PatternFill(start_color="99CC00", end_color="99CC00", fill_type="solid")}
    requiredColors = []
    requiredColors.append(highlightColors["mergedNoIssues"])
    if warFlag:
        requiredColors.append(highlightColors["mergedAdjusted"])
    if (worksheet[f'{"I"}{rowIndex}'].value == ""):
        requiredColors.append(highlightColors["mergedNoDOD"])
    if worksheet[f'{"G"}{rowIndex}'].value:
        if str(worksheet[f'{"G"}{rowIndex}'].value)[0] != "1":
            requiredColors.append(highlightColors["badDate"])
    if worksheet[f'{"I"}{rowIndex}'].value:
        if str(worksheet[f'{"I"}{rowIndex}'].value)[0] != "1":
            if str(worksheet[f'{"I"}{rowIndex}'].value)[0] == "2":
                if str(worksheet[f'{"I"}{rowIndex}'].value)[1] == "0":
                    if str(worksheet[f'{"I"}{rowIndex}'].value)[2:] > "23":
                        requiredColors.append(highlightColors["badDate"]) 
                else:
                    requiredColors.append(highlightColors["badDate"]) 
            else:
                requiredColors.append(highlightColors["badDate"]) 
    if (worksheet[f'{"N"}{rowIndex}'].value) != cemetery:
        requiredColors.append(highlightColors["mergedCemeteryMismatch"])
    try:
        if worksheet[f'{"B"}{rowIndex}'].value[0] != letter:
            requiredColors.append(highlightColors["mergedLastNameMismatch"])
    except Exception:
        pass
    if (worksheet[f'{"B"}{rowIndex}'].value) == (worksheet[f'{"C"}{rowIndex}'].value):
        requiredColors.append(highlightColors["mergedNameBug"])
    if requiredColors:
        numColors = len(requiredColors)
        cols_per_color = max(1, (14 - 2) // numColors)
        for colIndex in range(2, 16 + 1):
            if colIndex == 15:
                continue
            colorIndex = (colIndex - 2) // cols_per_color
            colorIndex = min(colorIndex, numColors - 1) 
            cell = worksheet.cell(row=rowIndex, column=colIndex)
            cell.fill = requiredColors[colorIndex]
        cell = worksheet.cell(row=rowIndex, column=16)
        cell.fill = requiredColors[-1]
        

'''
Finds the next empty row in the worksheet to begin processing at that index.

@param worksheet - The worksheet object being operated on

@return worksheet.max_row + 1 (int) - The row index of the next empty row in the worksheet

@author Mike
'''  
def find_next_empty_row(worksheet):
    for row in worksheet.iter_rows(min_row=1, min_col=1, max_col=1):
        if row[0].value is None:
            return row[0].row
    return worksheet.max_row + 1 


def highlightSingle(worksheet, cemetery, letter, warFlag, rowIndex, kinLast):
    tempFlag = False
    tempFlag2 = False
    highlightColors = {
        "warFlag": PatternFill(start_color="899499", end_color="899499", fill_type="solid"),
        "cemeteryMismatch": PatternFill(start_color="CF9FFF", end_color="CF9FFF", fill_type="solid"),
        "noDOD": PatternFill(start_color="A7C7E7", end_color="A7C7E7", fill_type="solid"),
        "lastNameMismatch": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "shortFirstName": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "badWarName": PatternFill(start_color="F5CBA7", end_color="F5CBA7", fill_type="solid"),
        "badDate": PatternFill(start_color="99CC00", end_color="99CC00", fill_type="solid")}
    requiredColors = []
    if warFlag:
        requiredColors.append(highlightColors["warFlag"])
    if (worksheet[f'{"N"}{rowIndex}'].value) != cemetery:
        requiredColors.append(highlightColors["cemeteryMismatch"])
    if (worksheet[f'{"I"}{rowIndex}'].value) == "":
        requiredColors.append(highlightColors["noDOD"])
    if worksheet[f'{"G"}{rowIndex}'].value:
        if str(worksheet[f'{"G"}{rowIndex}'].value)[0] != "1":
            requiredColors.append(highlightColors["badDate"])
    if worksheet[f'{"I"}{rowIndex}'].value:
        if str(worksheet[f'{"I"}{rowIndex}'].value)[0] != "1":
            if str(worksheet[f'{"I"}{rowIndex}'].value)[0] == "2":
                if str(worksheet[f'{"I"}{rowIndex}'].value)[1] == "0":
                    if str(worksheet[f'{"I"}{rowIndex}'].value)[2:] > "23":
                        requiredColors.append(highlightColors["badDate"]) 
                else:
                    requiredColors.append(highlightColors["badDate"]) 
            else:
                requiredColors.append(highlightColors["badDate"]) 
    if (worksheet[f'{"J"}{rowIndex}'].value) != "" and (worksheet[f'{"K"}{rowIndex}'].value) == ""\
    or (worksheet[f'{"L"}{rowIndex}'].value) != "" and (worksheet[f'{"M"}{rowIndex}'].value) == "":
        if worksheet[f'{"J"}{rowIndex}'].value != "Regular Service" or\
            worksheet[f'{"J"}{rowIndex}'].value != "Peacetime":
                requiredColors.append(highlightColors["badWarName"])
    try:
        if (worksheet[f'B{rowIndex}'].value)[0] != letter:
            if (worksheet[f'C{rowIndex}'].value)[0] == letter:
                tempLname = worksheet[f'B{rowIndex}'].value
                tempFname = worksheet[f'C{rowIndex}'].value
                worksheet[f'B{rowIndex}'].value = tempFname
                worksheet[f'C{rowIndex}'].value = tempLname
            else:
                tempFlag = True
        if len((worksheet[f'C{rowIndex}'].value)) < 3:
            if len((worksheet[f'D{rowIndex}'].value)) >= 3:
                tempLname = worksheet[f'C{rowIndex}'].value
                tempSuffix = worksheet[f'D{rowIndex}'].value
                worksheet[f'C{rowIndex}'].value = tempSuffix
                worksheet[f'D{rowIndex}'].value = tempLname + "."
                tempFlag2 = True
        if tempFlag and tempFlag2:
            tempLname = worksheet[f'B{rowIndex}'].value
            tempFname = worksheet[f'C{rowIndex}'].value
            worksheet[f'B{rowIndex}'].value = tempFname
            worksheet[f'C{rowIndex}'].value = tempLname
        if (worksheet[f'B{rowIndex}'].value)[0] != letter or len((worksheet[f'C{rowIndex}'].value)) < 3:
            requiredColors.append(highlightColors["lastNameMismatch"])
        if worksheet[f'C{rowIndex}'].value[-1].isupper():
            worksheet[f'D{rowIndex}'].value = worksheet[f'C{rowIndex}'].value[-1] + "."
            worksheet[f'C{rowIndex}'].value = worksheet[f'C{rowIndex}'].value[:-1]
    except IndexError:
        requiredColors.append(highlightColors["lastNameMismatch"])
    two_uppercase_pattern = re.compile(r'.*[A-Z].*[A-Z].*')
    if two_uppercase_pattern.match(worksheet[f'C{rowIndex}'].value):
        requiredColors.append(highlightColors["lastNameMismatch"])
    if kinLast:
        if kinLast != worksheet[f'B{rowIndex}'].value and \
        (kinLast == worksheet[f'C{rowIndex}'].value or
            kinLast == worksheet[f'D{rowIndex}'].value):
                requiredColors.append(highlightColors["lastNameMismatch"])
    if requiredColors:
        numColors = len(requiredColors)
        cols_per_color = max(1, (14 - 2) // numColors)  
        for colIndex in range(2, 16 + 1):
            if colIndex == 15:
                continue
            colorIndex = (colIndex - 2) // cols_per_color
            colorIndex = min(colorIndex, numColors - 1) 
            cell = worksheet.cell(row=rowIndex, column=colIndex)
            cell.fill = requiredColors[colorIndex]
        cell = worksheet.cell(row=rowIndex, column=16)
        cell.fill = requiredColors[-1]

'''
The main function that controls files sent to be processed as well as 
the data that is put into the excel sheet. It sets up the current working 
directory, the current sheet in the spreadsheet, and loops through all the 
images in a set letter in a set cemetery. This function is the primary 
controller for processing, saving, and handling of records.

The loop finds the last record recorded in the excel sheet
and then starts indexing from there. Calls createRecord is normal card, 
calls tempRecord and mergeRecord for A and B cards.

Calls redactImage for every card processed. Places the record ID and a 
hyperlink to the redacted image in the excel sheet during processing.
Saves worksheet after every image is done processing to reduce loss of 
data upon errors.

Controls highlighting of entire row of record upon different conditions 
that are caught by the program. 

@param singleFlag - flag that indicates if main function is being called 
                    by the duplicates utility program
@param singleCem - cemetery name variable for if main function is being 
                   called by the duplicates utility program
@param singleLetter - folder letter variable for if main function is being 
                      called by the duplicates utility program

@author Mike
'''
def main(singleFlag, singleCem, singleLetter):
    global cemSet, miscSet, jewishSet
    networkFolder = r"\\ucclerk\pgmdoc\Veterans"
    os.chdir(networkFolder)
    cemeterys = []
    for x in os.listdir(r"Cemetery"):
        cemeterys.append(x)
    miscs = []
    for x in os.listdir(r"Cemetery\Misc"):
        miscs.append(x)
    jewishs = []
    for x in os.listdir(r"Cemetery\Jewish"):
        jewishs.append(x)
    cemSet, miscSet, jewishSet = set(cemeterys), set(miscs), set(jewishs)
    workbook = openpyxl.load_workbook('Veterans.xlsx')
    cemetery = singleCem 
    cemPath = os.path.join(networkFolder, fr"Cemetery\{cemetery}")
    letter = singleLetter 
    namePath = letter 
    namePath = os.path.join(cemPath, namePath)
    pathA = ""
    rowIndex = 2
    worksheet = workbook[cemetery]
    warFlag = False
    pdfFiles = sorted(os.listdir(namePath))
    breakFlag = False
    for y in range(len(pdfFiles)):
        try:
            warFlag = False
            filePath = os.path.join(namePath, pdfFiles[y])
            rowIndex = find_next_empty_row(worksheet)
            try:
                id = worksheet[f'{"A"}{rowIndex-1}'].value + 1
            except Exception:
                id = int(pdfFiles[0][:-4].split(letter)[-1].lstrip('0'))
            string = pdfFiles[y][:-4].split(letter)[-1].lstrip('0')
            if "a" not in string and "b" not in string:
                if id != int(string.replace("a", "").replace("b", "")):
                    continue
                vals, warFlag, nameCoords, serialCoords, printedKVS, kinLast = createRecord(filePath, id, cemetery, "")
                redactedFile = redact(filePath, cemetery, letter, nameCoords, serialCoords)
                worksheet.cell(row=rowIndex, column=15).value = "PDF Image"
                worksheet.cell(row=rowIndex, column=15).font = Font(underline="single", color="0563C1")
                worksheet.cell(row=rowIndex, column=15).hyperlink = redactedFile
                counter = 1
                worksheet.cell(row=rowIndex, column=counter, value=id)
                counter += 1
                for x in vals:
                    worksheet.cell(row=rowIndex, column=counter, value=x)
                    counter += 1
                highlightSingle(worksheet, cemetery, letter, warFlag, rowIndex, kinLast)
                id += 1
                rowIndex += 1
            else:
                if id != int(string.replace("a", "").replace("b", "")):
                    continue
                if "a" in string:
                    if (filePath.replace("a.pdf", "") in pdfFiles):
                        continue
                    pathA = filePath
                    vals1, warFlag, nameCoords, serialCoords, printedKVS, kinLast = createRecord(filePath, id, cemetery, "A")
                    redactedFile = redact(filePath, cemetery, letter, nameCoords, serialCoords)
                if "b" in string:
                    if (filePath.replace("b.pdf", "") in pdfFiles):
                        continue
                    vals2, warFlagB, nameCoords, serialCoords, printedKVS, kinLast = createRecord(filePath, id, cemetery, "B")
                    if not warFlag or not warFlagB:
                        warFlag = False
                    else:
                        warFlag = True
                    redactedFile = redact(filePath, cemetery, letter, nameCoords, serialCoords)
                    mergeRecords(worksheet, vals1, vals2, rowIndex, id, warFlag, cemetery, letter)
                    mergeImages(pathA, filePath, cemetery, letter)
                    link_text = "PDF Image"
                    worksheet.cell(row=rowIndex, column=15).value = link_text
                    worksheet.cell(row=rowIndex, column=15).font = Font(underline="single", color="0563C1")
                    worksheet.cell(row=rowIndex, column=15).hyperlink = redactedFile.replace("b redacted.pdf", " redacted.pdf")
                    id += 1
                    rowIndex += 1
                    if singleFlag:
                        breakFlag = True
        except Exception as e:
            errorTraceback = traceback.format_exc()
            print(errorTraceback)
            print(f"An error occurred: {e}")
            errorMessage = f"SKIPPED DUE TO ERROR : {errorTraceback}"
            worksheet.cell(row=rowIndex, column=1, value=id)
            worksheet.cell(row=rowIndex, column=2, value=errorMessage)
            highlightColor = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for colIndex in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=rowIndex, column=colIndex)
                cell.fill = highlightColor
            error_file_path = fr'Errors/{cemetery}{letter}{str(id).zfill(5)} Error.txt' 
            with open(error_file_path, 'a') as error_file:
                error_file.write(f'{printedKVS} \n\n {errorTraceback}')
            id += 1
            rowIndex += 1
        extension1 = str(id-1).zfill(5)
        extension2 = ""
        if "a.pdf" in filePath:
            extension1 = str(id).zfill(5)
            extension2 = "a"
        elif "b.pdf" in filePath:
            extension1 = str(id-1).zfill(5)
            extension2 = "b"
        logFilePath = fr'Logs/{cemetery}{letter}{extension1}{extension2} Extracted.txt' 
        with open(logFilePath, 'w', encoding='utf-8') as logFile:
            logFile.write(printedKVS)
            logFile.write("\n")
        workbook.save('Veterans.xlsx')
        if breakFlag:
            break

if __name__ == "__main__":
    global cemetery
    cemetery = "Graceland" # Change this to continue running through cemeteries
    global letter
    letter = "P" # Change this to continue running through the current cemetery
    main(False, cemetery, letter)