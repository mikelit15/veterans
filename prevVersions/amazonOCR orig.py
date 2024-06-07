import boto3
import re
import openpyxl 
from openpyxl.styles import Font
from collections import defaultdict
import os
import PyPDF2
import fitz
import cv2
from reportlab.pdfgen import canvas
from PIL import Image
from nameparser import HumanName
import dateparser
from nameparser.config import CONSTANTS
from openpyxl.styles import PatternFill

def redact(file, flag):
    PDF_Document  = fitz.open(file)
    firstPage  = PDF_Document.load_page(0)
    firstPageImage  = firstPage.get_pixmap(matrix=fitz.Matrix(600/72, 600/72))
    imageFile  = "temp.png"
    firstPageImage.save(imageFile)
    img = cv2.imread(imageFile)
    pt1 = (2670, 500)
    pt3 = (3526, 950)
    if flag:
        pt1 = (2670, 260)
        pt3 = (3500, 550)
    cv2.rectangle(img, pt1, pt3, (0, 0, 0), thickness=cv2.FILLED)
    cv2.imwrite(imageFile, img)
    image = Image.open(imageFile)
    pdfFile = "temp.pdf"
    c = canvas.Canvas(pdfFile, pagesize=(image.width, image.height))
    c.drawImage(imageFile, 0, 0, width=image.width, height=image.height)
    c.save()
    
    redactedPDF_Document  = fitz.open(pdfFile)
    newPDF_Document  = fitz.open()
    newPDF_Document.insert_pdf(redactedPDF_Document, from_page=0, to_page=0)
    newPDF_Document.insert_pdf(PDF_Document, from_page=1, to_page=1)
    newPDF_File  = f'{file.replace(".pdf", "")} redacted.pdf'
    newPDF_Document.save(newPDF_File)
    newPDF_Document.close()
    redactedPDF_Document.close()
    PDF_Document.close()    

def getKV_Map(fileName):
    with open(fileName, 'rb') as file:
        imgTest = file.read()
        bytesTest = bytearray(imgTest)
        print('Image loaded', fileName)
    client = boto3.client('textract', region_name='us-east-1')
    response = client.analyze_document(Document={'Bytes': bytesTest}, FeatureTypes=['FORMS'])
    blocks = response['Blocks']
    keyMap = {}
    civil = ""
    world = ""
    badVal = ""
    valueMap = {}
    blockMap = {}
    for block in blocks:
        try:
            if block['Text'] == "Civil War":
                civil = "Civil War"
            if block['Text'] == 'World War':
                world = "World War"
            # if block['Text'] in cemSet:
            #     if block['Text'] not in badVal:
            #         badVal.append(block['Text'])
        except KeyError:
            None
        blockID = block['Id']
        blockMap[blockID] = block
        if block['BlockType'] == "KEY_VALUE_SET":
            if 'KEY' in block['EntityTypes']:
                keyMap[blockID] = block
            else:
                valueMap[blockID] = block

    return keyMap, valueMap, blockMap, civil, badVal, world

def getKV_Relationship(keyMap, valueMap, blockMap):
    kvs = defaultdict(list)
    for blockID, keyBlock in keyMap.items():
        valueBlock = findValueBlock(keyBlock, valueMap)
        key = getText(keyBlock, blockMap)
        key = re.sub(r'[^a-zA-Z0-9\s]', '', key)
        val = getText(valueBlock, blockMap)
        kvs[key].append(val)
    return kvs

def findValueBlock(keyBlock, valueMap):
    for relationship in keyBlock['Relationships']:
        if relationship['Type'] == 'VALUE':
            for valueID in relationship['Ids']:
                valueBlock = valueMap[valueID]
    return valueBlock

def getText(result, blocksMap):
    text = ''
    if 'Relationships' in result:
        for relationship in result['Relationships']:
            if relationship['Type'] == 'CHILD':
                for childID in relationship['Ids']:
                    word = blocksMap[childID]
                    if word['BlockType'] == 'WORD':
                        text += word['Text'] + ' '
                    if word['BlockType'] == 'SELECTION_ELEMENT':
                        if word['SelectionStatus'] == 'SELECTED':
                            text += 'X '
    return text

def printKVS(kvs):
    for key, value in kvs.items():
        print(key, ":", value)

def searchValue(kvs, searchKey):
    for key, value in kvs.items():
        key = str.rstrip(key)
        if key.upper() == searchKey.upper():
            return value

def searchValueX(kvs, searchKey):
    for key, value in kvs.items():
        if re.search(searchKey, key, re.IGNORECASE):
            return value
        
def nameRule(worksheet, rowIndex, letter, value, counter):
    CONSTANTS.force_mixed_case_capitalization = True
    name = HumanName(value)
    name.capitalize()
    firstName = name.first
    middleName = name.middle
    lastName = name.last
    suffix = name.suffix
    temp = value.replace("Jr.", "").replace("Sr.", "")
    if ("," in value and "." in firstName):
        if len(middleName) > 0:
            middleName = name.first
            firstName = name.middle
        else:
            lastName = name.last
            firstName = name.first
    elif ("." in firstName):
        if len(middleName) > 0:
            lastName = name.first
            firstName = name.middle
            middleName = name.last
        else:
            lastName = name.first
            firstName = name.last
    elif ("." in lastName):
        if len(middleName) == 0:
            lastName = name.last
            firstName = name.first
        else:
            lastName = name.first
            middleName = name.last
            firstName = name.middle
    elif len(suffix) > 0:
        suffix = suffix.replace(", ", "")
        middleName = suffix.replace("Sr.", "").replace("Jr.", "")
        suffix = suffix.replace(middleName, "")
    if ("Jr." in value or "Sr." in value) and "." not in temp and "," not in temp:
        if len(middleName) > 0:
            firstName = name.middle
            middleName = name.last
            lastName = name.first
        else:
            firstName = name.last
            lastName = name.first
    elif "." not in temp and "," not in temp:
        if len(middleName) > 0:
            firstName = name.middle
            middleName = name.last
            lastName = name.first
        else:
            lastName = name.first
            firstName = name.last
    if len(middleName) > 2:
        middleName = middleName.replace(".", "")
    dots = 0
    for x in firstName:
        if (x == "."):
            dots+=1
    if (dots == 2):
        firstName = firstName.upper()
    else:
        firstName = firstName.replace(".", "")
    worksheet.cell(row=rowIndex, column=counter, value=lastName.replace(",", "").replace(".", ""))
    counter += 1
    worksheet.cell(row=rowIndex, column=counter, value=firstName.replace(",", ""))
    counter += 1
    worksheet.cell(row=rowIndex, column=counter, value=middleName.replace(",", "."))
    counter += 1
    worksheet.cell(row=rowIndex, column=counter, value=suffix.replace(",", "."))
    highlightColor = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    try:
        if lastName[0] != letter:
            for cell in worksheet[rowIndex]:
                cell.fill = highlightColor
    except IndexError:
        None
    return counter

def dobRule(value, counter):
    date = value
    counter -= 1
    return counter, date     

def dodRule(worksheet, rowIndex, value, counter, dob, century, buried, war):
    wars19 = ["Vietnam War", "Korean War", "World War 2"]
    birthFlag = False
    deathFlag = False
    birth = dob
    birth = birth.replace(":", ".")
    if birth[-1:] == " ":
        birth = birth.replace(" ", "")
    try:
        if dateparser.parse(birth, settings={'STRICT_PARSING': True}) == None:
            birth = birth
        else:
            birthFlag = True
            birth = dateparser.parse(birth)
    except dateparser.ParserError:
        birth = "" 
    death = value
    death = death.replace(":", ".")
    try:
        if dateparser.parse(death, settings={'STRICT_PARSING': True}) == None:
            death = death
        else:
            deathFlag = True
            death = dateparser.parse(death)
    except dateparser.ParserError:
        death = "" 
    if birth == "" and death == "":
        counter += 3
        buried = buried.replace(" ", "")
        if len(buried) == 4:
            worksheet.cell(row=rowIndex, column=counter, value=int(buried))
        elif len(buried) == 2:
            worksheet.cell(row=rowIndex, column=counter, value=int("19" + buried))
    elif birth != "" and death == "":
        if isinstance(birth, str):
            counter += 1
            birth = re.sub(r'[^0-9]', '', birth)
            if(birth != ""):
                worksheet.cell(row=rowIndex, column=counter, value=int(birth))
            counter += 2
        else:
            if(birth.strftime("%Y")[:2] == "20"):
                dobYear = "19" + birth.strftime("%x")[-2:]
                worksheet.cell(row=rowIndex, column=counter, value=birth.strftime("%m/%d/" + dobYear))
                counter += 1
                if(dobYear != ""):
                    worksheet.cell(row=rowIndex, column=counter, value=int(dobYear))
                counter += 2
            else:
                worksheet.cell(row=rowIndex, column=counter, value=birth.strftime("%m/%d/%Y"))
                counter += 1
                if(birth.year != ""):
                    worksheet.cell(row=rowIndex, column=counter, value=int(birth.year))
                counter += 2
    elif birth == "" and death != "":
        if isinstance(death, str):
            counter += 3
            death = re.sub(r'[^0-9]', '', death)
            if(death != ""):
                worksheet.cell(row=rowIndex, column=counter, value=int(death))
        else:
            if(death.strftime("%Y")[:2] == "20"):
                dodYear = "19" + death.strftime("%x")[-2:]
                counter += 2
                worksheet.cell(row=rowIndex, column=counter, value=death.strftime("%m/%d/" + dodYear))
                counter += 1
                if(dodYear != ""):
                    worksheet.cell(row=rowIndex, column=counter, value=int(dodYear))
            else:    
                counter += 2
                worksheet.cell(row=rowIndex, column=counter, value=death.strftime("%m/%d/%Y"))
                counter += 1
                if(death.year != ""):
                    worksheet.cell(row=rowIndex, column=counter, value=int(death.year))
    else:
        try:
            birthLen = len(birth.strftime("%x"))
            deathLen = len(death.strftime("%x"))
            if birthLen >= 4 and deathLen >= 4:
                dobYear = birth.strftime("%x")[-2:]
                dodYear = death.strftime("%x")[-2:]
            if(dobYear > dodYear):
                yearb = "18" + dobYear
                yeard = "19" + dodYear
            else:
                yearb = "19" + dobYear
                yeard = "19" + dodYear
            worksheet.cell(row=rowIndex, column=counter, value=birth.strftime("%m/%d/" + yearb))
            counter += 1
            if(yearb != ""):
                worksheet.cell(row=rowIndex, column=counter, value=int(yearb))
            counter += 1
            worksheet.cell(row=rowIndex, column=counter, value=death.strftime("%m/%d/" + yeard))
            counter += 1
            if(yeard != ""):
                worksheet.cell(row=rowIndex, column=counter, value=int(yeard))
        except AttributeError:
            if not birthFlag and not deathFlag:
                counter += 1
                birth = re.sub(r'[^0-9]', '', birth)
                # if len(birth) > 4:
                #     worksheet.cell(row=rowIndex, column=counter, value=int(birth[4:]))
                # else:
                if(birth != ""):
                    worksheet.cell(row=rowIndex, column=counter, value=int(birth))
                counter += 2
                death = re.sub(r'[^0-9]', '', death)
                # if len(death) > 4:
                #     worksheet.cell(row=rowIndex, column=counter, value=int(death[-4:]))
                # else:
                if(death != ""):
                    worksheet.cell(row=rowIndex, column=counter, value=int(death))
            elif birthFlag and not deathFlag:
                dodYear = death[-2:]
                dobYear = birth.strftime("%x")[-2:]
                dobCent = birth.strftime("%Y")[:2]
                if dobYear > dodYear:
                    yearb = "18" + dobYear
                    yeard = "19" + dodYear
                else:
                    if dobCent == "18":
                        yearb = "18" + dobYear
                        yeard = "18" + dodYear
                    else:    
                        yearb = "19" + dobYear
                        yeard = "19" + dodYear
                worksheet.cell(row=rowIndex, column=counter, value=birth.strftime("%m/%d/" + yearb))
                counter += 1
                if(yearb != ""):
                    worksheet.cell(row=rowIndex, column=counter, value=int(yearb))
                counter += 1
                if(yeard != ""):
                    worksheet.cell(row=rowIndex, column=counter, value=int(yeard))
            elif not birthFlag and deathFlag:
                if len(birth) > 4:
                    index = birth.find("18")
                    if index:
                        dobYear = birth[index+2:index+4]
                        dobYear = re.sub(r'[^0-9]', '', dobYear)
                    elif not index:
                        index = birth.find("19")
                        dobYear = birth[index+2:index+4]
                        dobYear = re.sub(r'[^0-9]', '', dobYear)
                    else:
                        dobYear = birth[:4]
                else:
                    dobYear = birth[-2:]
                dodYear = death.strftime("%x")[-2:]
                dodCent = death.strftime("%Y")[:2]
                if dobYear > dodYear:
                    yearb = "18" + dobYear
                    yeard = "19" + dodYear
                else:
                    if dodCent == "18":
                        yearb = "18" + dobYear
                        yeard = "18" + dodYear
                    else:    
                        if dobYear == "":
                            yearb = ""
                            yeard = "19" + dodYear
                        else:
                            yearb = "19" + dobYear
                            yeard = "19" + dodYear
                counter += 1
                birth = re.sub(r'[^0-9]', '', birth)
                if(yearb != ""):
                    worksheet.cell(row=rowIndex, column=counter, value=int(yearb))
                counter += 1
                worksheet.cell(row=rowIndex, column=counter, value=death.strftime("%m/%d/" + yeard))
                counter += 1
                if(yeard != ""):
                    worksheet.cell(row=rowIndex, column=counter, value=int(yeard))
    if century:
        counter += 1
        worksheet.cell(row=rowIndex, column=counter, value="Yes")
    else:
        counter += 1
        worksheet.cell(row=rowIndex, column=counter, value="No Field")
    return counter

def warRule(value, civil, world):
    war = value
    if war:
        if war[0] == " ":
            war = war[1:]
    ww1years = ["1914", "1915", "1916", "1917", "1918"]
    if "N." in war:
        war = war.replace("N", "W")
    tempText = war.replace("-", "").replace(".", "").replace("#", "").replace(" ", "").replace(",", "")
    # WW1and2_Pattern = re.compile(r'WWI&?II|WW1&2|WW\s*1\s*&\s*2|World\s*War\s*(1|I)&(2|II)|World\s*War\s*(1|I)\s*&\s*(2|II)')
    WW1_Pattern = re.compile(r'WWI|WW1|WW\s*I|World\s*War\s*(1|I)|World\s*War|WW\s*1|1')
    WW2_Pattern = re.compile(r'WWII|WW2|WW\s*2|WW\s*#?II|World\s*War\s*(II|2)|2')
    if (WW2_Pattern.findall(tempText)):
        war = "World War 2"
    elif (WW1_Pattern.findall(tempText)):
        war = "World War 1"
    # if (ww1_and_2_pattern.findall(tempText)):
    #     war = "World War 1 and 2"
    else:
        war = war.replace(".", "").replace("Calv", "").replace("Vols", "")
        if "Korea" in war:
            war = "Korean War"
        elif "Vietnam" in war:
            war = "Vietnam War"
        elif "Civil" in war:
            war = "Civil War"
        elif "Spanish" in war or "Amer" in war or "American" in war:
            war = "Spanish American War"
        elif "Peacetime" in war or "Not Listed" in war:
            war = ""
        words = war.split()
        for x in words:
            if x in ww1years:
                war = "World War 1"
                break
    if civil != "" and war != "Civil War":
        war = "Civil War"
    if world != "" and war != "World War 1":
        war = "World War 1"
    if "Army" in war or "US" in war:
        war = ""
    return war

def warRule2(worksheet, rowIndex, counter, war):
    worksheet.cell(row=rowIndex, column=counter, value=war)
    return counter 

def branchRule(worksheet, rowIndex, value, counter, war):
    # stateInfo = {
    # ('Alabama', 'AL'), ('Alaska', 'AK'), ('Arizona', 'AZ'), ('Arkansas', 'AR'), ('California', 'CA'),
    # ('Colorado', 'CO'), ('Connecticut', 'CT'), ('Delaware', 'DE'), ('Florida', 'FL'), ('Georgia', 'GA'),
    # ('Hawaii', 'HI'), ('Idaho', 'ID'), ('Illinois', 'IL'), ('Indiana', 'IN'), ('Iowa', 'IA'), ('Kansas', 'KS'),
    # ('Kentucky', 'KY'), ('Louisiana', 'LA'), ('Maine', 'ME'), ('Maryland', 'MD'), ('Massachusetts', 'MA'),
    # ('Michigan', 'MI'), ('Minnesota', 'MN'), ('Mississippi', 'MS'), ('Missouri', 'MO'), ('Montana', 'MT'),
    # ('Nebraska', 'NE'), ('Nevada', 'NV'), ('New Hampshire', 'NH'), ('New Jersey', 'NJ'), ('New Mexico', 'NM'),
    # ('New York', 'NY'), ('North Carolina', 'NC'), ('North Dakota', 'ND'), ('Ohio', 'OH'), ('Oklahoma', 'OK'),
    # ('Oregon', 'OR'), ('Pennsylvania', 'PA'), ('Rhode Island', 'RI'), ('South Carolina', 'SC'),
    # ('South Dakota', 'SD'), ('Tennessee', 'TN'), ('Texas', 'TX'), ('Utah', 'UT'), ('Vermont', 'VT'),
    # ('Virginia', 'VA'), ('Washington', 'WA'), ('West Virginia', 'WV'), ('Wisconsin', 'WI'), ('Wyoming', 'WY'),
    # ('United States', 'US')}
    # branch = value
    # if "Army" in branch:
    #     branch = "Army"
    # elif "Navy" in branch:
    #     branch = "Navy"
    # elif "Air Force" in branch:
    #     branch = "Air Force"
    # elif "USA" in branch:
    #     branch = ""
    # else:
    #     company = ""
    #     state = ""
    #     branch = branch.replace('"', "").replace("/", " ")
    #     companyLetterMatch = re.search(r'(Co\.|Co|C)\s*([A-Z])', branch)
    #     try:
    #         if len(companyLetterMatch.group(1)) != 1:
    #             company = companyLetterMatch.group(2)
    #         else:
    #             company = companyLetterMatch.group(1)
    #         numberMatch = re.search(r'(\d+)', branch)
    #         reg = numberMatch.group(1)
    #         word = branch.split()
    #         for y in word:
    #             y = y.replace(".", "").replace("Inf", "")
    #             for z in stateInfo:
    #                 if y == z[1]:
    #                     state = z[1]
    #                 elif y == z[0]:
    #                     state = z[1]
    #     except AttributeError:
    #         numberMatch = re.search(r'(\d+)', branch)
    #         reg = numberMatch.group(1)
    #         word = branch.split()
    #         for y in word:
    #             y = y.replace(".", "").replace("Inf", "")
    #             for z in stateInfo:
    #                 if y == z[1]:
    #                     state = z[1]
    #                 elif y == z[0]:
    #                     state = z[1]
    #     if state == "":
    #         branch = branch.replace(" ", "")
    #         branch = branch.replace("N.Y.", " NY ").replace("N.J.", " NJ ").replace("U.S.", " US ")
    #         word = branch.split()
    #         for y in word:
    #             for z in stateInfo:
    #                 if y == z[1]:
    #                     state = z[1]
    #                 elif y == z[0]:
    #                     state = z[1]
    #     if state == "":
    #         branch = "US Reg " + reg
    #     elif company == "":
    #         branch = state + " Reg " + reg
    #     else:
    #         branch = state + " Reg " + reg + " - Company " + company
    # worksheet.cell(row=rowIndex, column=counter, value=branch)
    # return counter  
    armys = ["co", "army", "inf", "infantry", "Infan", "usa", "med", \
            "cav", "div", "sig", "art", "corps", "corp"]
    navys = ["hospital", "navy", "naval"]
    guards = ["113", "102d", "114", "44", "181", "250", "112"]
    branch = value
    branch = branch.replace("/", " ").replace(".", " ").replace("th", "").replace("-", "")
    words = branch.split()
    if war == "Civil War":
        branch = "Army"
    for word in words:
        word = word.lower()
        if "Air Force" in branch:
            branch = "Air Force"
            break
        if "Marine" in branch:
            branch = "Marine Corps"
            break
        if word in armys:
            branch = "Army"
        if word in navys:
            branch = "Navy"
        if word in guards:
            branch = "National Guard"
            break
        if "Not Listed" in branch:
            branch = ""
    WW1_Pattern = re.compile(r'WWI|WW1|WW\s*I|World\s*War\s*(1|I)|World\s*War|WW\s*1')
    WW2_Pattern = re.compile(r'WWII|WW2|WW\s*2|WW\s*#?II|World\s*War\s*(II|2)|2')
    if (WW2_Pattern.findall(branch) or (WW1_Pattern.findall(branch))):
        branch = ""
    worksheet.cell(row=rowIndex, column=counter, value=value)   
    counter += 1
    worksheet.cell(row=rowIndex, column=counter, value=branch)   
    return counter

def cemeteryRule(worksheet, rowIndex, value, counter, cemetery, badVal):
    if badVal != "":
        worksheet.cell(row=rowIndex, column=counter, value=badVal)
    else:
        strippedText = ""
        if cemetery in value:
            strippedText = cemetery
        else:
            strippedText = value
        worksheet.cell(row=rowIndex, column=counter, value=strippedText)
    return counter      

def createRecord(rowIndex, fileName, id):
    century = None
    buried = ""
    civil = ""
    world = ""
    war = ""
    badVal = []
    pageReader = PyPDF2.PdfReader(open(fileName, 'rb'))
    page = pageReader.pages[0]
    pdfWriter = PyPDF2.PdfWriter()
    pdfWriter.add_page(page)
    with open("temp.pdf", 'wb') as outputPDF:
        pdfWriter.write(outputPDF)
    keyMap, valueMap, blockMap, civil, badVal, world = getKV_Map("temp.pdf")
    kvs = getKV_Relationship(keyMap, valueMap, blockMap)
    printKVS(kvs)
    keys = ["", "NAME", "WAR RECORD", "BORN", "19", "DATE OF DEATH", "WAR RECORD", "BRANCH OF SERVICE" , "IN"]
    counter = 1
    flag = False
    flag3 = False
    worksheet.cell(row=rowIndex, column=counter, value=id)
    counter += 1
    if searchValueX(kvs, "Care Assigned") == None:
        flag = False
    else:
        flag = True
    dob = ""
    for x in keys:
        try:
            value = searchValue(kvs, x)[0]
        except TypeError:
            value = "" 
        if x == "NAME":
            counter = nameRule(worksheet, rowIndex, letter, value, counter)
        elif x == "":
            value = value.replace(" ", "")
            buried = value[-2:]
            counter -= 1
        elif x == "BORN":
            counter, dob = dobRule(value, counter)
        elif x == "DATE OF DEATH":
            counter = dodRule(worksheet, rowIndex, value, counter, dob, century, buried, war)
        elif x == "WAR RECORD" and flag3:
            counter = warRule2(worksheet, rowIndex, counter, value)
            counter += 1
            counter = warRule2(worksheet, rowIndex, counter, war)
            flag3 = False
        elif x == "WAR RECORD":
            if "&" in value:
                wars = value.split("&")
                war1 = warRule(wars[0], civil, world)
                war2 = warRule(wars[1], civil, world)
                war = war1 + " and " + war2
            else:
                war = warRule(value, civil, world)
            flag3 = True
            counter -= 1
        elif x == "BRANCH OF SERVICE":
            counter = branchRule(worksheet, rowIndex, value, counter, war)
        elif x == "19":
            if value != "":
                century = True
                buried = "19" + value
            counter -= 1
        elif x == "IN":
            counter = cemeteryRule(worksheet, rowIndex, value, counter, cemetery, badVal)
        counter += 1
    redact(fileName, flag)
    linkText = "PDF Image"
    worksheet.cell(row=rowIndex, column=counter).value = linkText
    worksheet.cell(row=rowIndex, column=counter).font = Font(underline="single", color="0563C1")
    worksheet.cell(row=rowIndex, column=counter).hyperlink = f'{fileName.replace(".pdf", "")} redacted.pdf'

def createRecord2(rowIndex, fileName, id):
    century = None
    buried = ""
    civil = ""
    world = ""
    war = ""
    badVal = []
    pageReader = PyPDF2.PdfReader(open(fileName, 'rb'))
    page = pageReader.pages[0]
    pdfWriter = PyPDF2.PdfWriter()
    pdfWriter.add_page(page)
    with open("temp.pdf", 'wb') as outputPDF:
        pdfWriter.write(outputPDF)
    keyMap, valueMap, blockMap, civil, badVal, world = getKV_Map("temp.pdf")
    kvs = getKV_Relationship(keyMap, valueMap, blockMap)
    printKVS(kvs)
    keys = ["", "NAME", "WAR RECORD", "BORN", "19", "DATE OF DEATH", "WAR RECORD", "BRANCH OF SERVICE" , "IN"]
    counter = 1
    flag = False
    flag3 = False
    worksheet.cell(row=rowIndex, column=counter, value=id)
    counter += 1
    if searchValueX(kvs, "Care Assigned") == None:
        flag = False
    else:
        flag = True
    dob = ""
    for x in keys:
        try:
            value = searchValue(kvs, x)[0]
        except TypeError:
            value = "" 
        if x == "NAME":
            counter = nameRule(worksheet, rowIndex, letter, value, counter)
        elif x == "":
            value = value.replace(" ", "")
            buried = value[-2:]
            counter -= 1
        elif x == "BORN":
            counter, dob = dobRule(value, counter)
        elif x == "DATE OF DEATH":
            counter = dodRule(worksheet, rowIndex, value, counter, dob, century, buried, war)
        elif x == "WAR RECORD" and flag3:
            counter = warRule2(worksheet, rowIndex, counter, value)
            counter += 1
            counter = warRule2(worksheet, rowIndex, counter, war)
            flag3 = False
        elif x == "WAR RECORD":
            if "&" in value:
                wars = value.split("&")
                war1 = warRule(wars[0], civil, world)
                war2 = warRule(wars[1], civil, world)
                war = war1 + " and " + war2
            else:
                war = warRule(value, civil, world)
            flag3 = True
            counter -= 1
        elif x == "BRANCH OF SERVICE":
            counter = branchRule(worksheet, rowIndex, value, counter, war)
        elif x == "19":
            if value != "":
                century = True
                buried = "19" + value
            counter -= 1
        elif x == "IN":
            counter = cemeteryRule(worksheet, rowIndex, value, counter, cemetery, badVal)
        counter += 1
    redact(fileName, flag)
    linkText = "PDF Image"
    worksheet.cell(row=rowIndex, column=counter).value = linkText
    worksheet.cell(row=rowIndex, column=counter).font = Font(underline="single", color="0563C1")
    worksheet.cell(row=rowIndex, column=counter).hyperlink = f'{fileName.replace(".pdf", "")} redacted.pdf'

def clean(cemetery, letter, namePath):
    PDF_Files = [file for file in os.listdir(namePath) if file.lower().endswith('.pdf')]
    letterBefore = list(namePath)
    letterBefore[-1] = chr(ord(letter) - 1)
    letterBefore = ''.join(letterBefore)
    PDF_FilesBefore = [file for file in os.listdir(letterBefore) if file.lower().endswith('.pdf')]
    counter = PDF_FilesBefore[len(PDF_FilesBefore) - 1]
    match = re.search(r'\d+', counter)
    number = match.group()
    counter = int(number) + 1
    for x in PDF_Files:
        if "redacted" in x:
            newName = f"{cemetery}{letter}{counter:05d} redacted.pdf"
            os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
            counter -= 1
        else:    
            newName = f"{cemetery}{letter}{counter:05d}.pdf"
            os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
        counter += 1

def main():
    cemeterys = ["Graceland", "Hazelwood", "Evergreen", "Fairview", "Hillside", "Hollywood", 
                "Hollywood Memorial", "Rosehill", "St. Gertrude's", "St. Mary's", "St. Teresa's",
                "Mt. Olivet", "Misc.", "Rosedale", "Rahway", "Mt. Calvary", "Beth David",
                "B'Nai Abraham", "B'Nai Israel", "B'Nai Jeshuran", "Gemel Chesed", "Ohed Sholom"]
    global cemSet
    cemSet = set(cemeterys)
    networkFolder = r"\\ucclerk\pgmdoc\Veterans"
    os.chdir(networkFolder)
    workbook = openpyxl.load_workbook('Veterans2.xlsx')
    global cemetery
    cemetery = "Evergreen"
    cemPath = cemetery
    cemPath = os.path.join(networkFolder, cemPath)
    global letter
    letter = "M"
    namePath = letter
    namePath = os.path.join(cemPath, namePath)
    rowIndex = 1
    global worksheet
    worksheet = workbook[cemetery]
    for cell in worksheet["A"]:
        if cell.value is None:
            rowIndex = cell.row
            break
        else:
            rowIndex = cell.row + 1
    id = rowIndex - 1
    clean(cemetery, letter, namePath)
    PDF_Files = [file for file in os.listdir(namePath) if file.lower()]
    for y in range(len(PDF_Files)):
        filePath = os.path.join(namePath, PDF_Files[y])
        if "output" in PDF_Files[y] or "redacted" in PDF_Files[y]:
            continue
        else:
            string = PDF_Files[y][:-4]
            string = string.split(letter) 
            string = string[-1].lstrip('0')
            if id != int(string):
                continue
            createRecord(rowIndex, filePath, id)
            id += 1
            rowIndex += 1
        workbook.save('Veterans2.xlsx')

if __name__ == "__main__":
    main()