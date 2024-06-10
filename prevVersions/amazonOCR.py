import boto3
import re
import openpyxl 
from openpyxl.styles import Font
from collections import defaultdict
import os
import PyPDF2
import fitz
import cv2
from reportlab.pdfgen import canvas
from PIL import Image
from nameparser import HumanName
import dateparser
from nameparser.config import CONSTANTS
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz


'''
Redacts sensitive information (e.g., social security or ID numbers) from a registration 
card by adding a black box over the specified area. The method of redaction varies based 
on the card format. The first page of the PDF is saved as a PNG, and the black box is 
added to the PNG. A new PDF is then created using the redacted PNG for the first page and 
the original second page.

@param file (str) - Path to the file to be processed
@param flag (bool) - Flag indicating the card format, effects the position 
                     of the redaction
@param cemetery (str) - Name of the cemetery, used for categorizing and saving the 
                        file in the correct directory
@param letter (str) - Last name letter, used for categorizing and saving the 
                      file in the correct directory

@return newPDF_File (str) - Path to the newly created redacted PDF file

@author Mike
'''
def redact(file, flag, cemetery, letter):
    pdfDocument  = fitz.open(file)
    firstPage  = pdfDocument.load_page(0)
    firstPageImage  = firstPage.get_pixmap(matrix=fitz.Matrix(600/72, 600/72))
    imageFile  = "temp.png"
    firstPageImage.save(imageFile)
    img = cv2.imread(imageFile)
    pt1 = (2670, 500)
    pt3 = (3526, 950)
    if flag:
        pt1 = (2670, 260)
        pt3 = (3500, 570)
    cv2.rectangle(img, pt1, pt3, (0, 0, 0), thickness=cv2.FILLED)
    cv2.imwrite(imageFile, img)
    image = Image.open(imageFile)
    pdfFile = "temp.pdf"
    c = canvas.Canvas(pdfFile, pagesize=(image.width, image.height))
    c.drawImage(imageFile, 0, 0, width=image.width, height=image.height)
    c.save()

    fullLocation = r"\\ucclerk\pgmdoc\Veterans\Cemetery - Redacted"
    redactedLocation = f'{cemetery} - Redacted'
    fullLocation = os.path.join(fullLocation, redactedLocation)
    fullLocation = os.path.join(fullLocation, letter)
    fileName = file.split(letter)
    redactedPDF_Document  = fitz.open(pdfFile)
    newPDF_Document  = fitz.open()
    newPDF_Document.insert_pdf(redactedPDF_Document, from_page=0, to_page=0)
    newPDF_Document.insert_pdf(pdfDocument, from_page=1, to_page=1)
    newPDF_File  = f'{fullLocation}\\{cemetery}{letter}{fileName[2].replace(".pdf", "")} redacted.pdf'
    newPDF_Document.save(newPDF_File)
    newPDF_Document.close()
    redactedPDF_Document.close()
    pdfDocument.close()
    return newPDF_File    


'''
Merges 'A' and 'B' redacted files into a single combined file, showing both versions 
of the card, both redacted.

@param pathA (str) - File path for file A (first document to be merged)
@param pathB (str) - File path for file B (second document to be merged)
@param cemetery (str) - Name of the cemetery, used for categorizing and saving the 
                        merged file in the correct directory
@param letter (str) - Last name letter, used for categorizing and saving the 
                      merged file in the correct directory

@author Mike
'''
def mergeImages(pathA, pathB, cemetery, letter):
    fullLocation = r"\\ucclerk\pgmdoc\Veterans\Cemetery - Redacted"
    redactedLocation = f'{cemetery} - Redacted'
    fullLocation = os.path.join(fullLocation, redactedLocation)
    fullLocation = os.path.join(fullLocation, letter)
    fileName = pathA.split(letter)
    merger = PyPDF2.PdfMerger()
    merger.append(pathA)
    merger.append(pathB)
    string = pathA[:-5]
    string = string.split(letter) 
    string = string[-1].lstrip('0')
    # with open(f'\\ucclerk\pgmdoc\Veterans\\{cemetery}{letter}{fileName[2].replace("a.pdf", ".pdf")}', 'wb') as outputPDF:
    #     merger.write(outputPDF)
    merger2 = PyPDF2.PdfMerger()
    merger2.append(f'{fullLocation}\\{cemetery}{letter}{fileName[2].replace(".pdf", "")} redacted.pdf')
    merger2.append(f'{fullLocation}\\{cemetery}{letter}{fileName[2].replace("a.pdf", "b")} redacted.pdf')
    string = pathA[:-5]
    string = string.split(letter) 
    string = string[-1].lstrip('0')
    with open(f'{fullLocation}\\{cemetery}{letter}{fileName[2].replace("a.pdf", "")} redacted.pdf', 'wb') as outputPDF:
        merger2.write(outputPDF)


'''
Extracts key-value pairs from the specified file using AWS Textract, which performs 
an AI analysis to identify and extract text from the form.

@param fileName (str) - Path to the file to be processed
@param id (int) - ID of the file, used for referencing the specific file that is 
                  currently being processed, printed in the terminal
@param cemetery (str) - Name of the cemetery, used for brute force extraction, for example
                        if the key was not found or combined with another key. Also used 
                        with fuzzy API to catch cemetery if it has minor spelling mistakes

@return keyMap (dict) - Dictionary of keys extracted from the document
@return valueMap (dict) - Dictionary of values extracted from the document
@return blockMap (dict) - Dictionary of blocks extracted from the document
@return civil (str) - Extracted "Civil War" if text found outside of the war key
@return cem (str) - Brute force extraction of cemetery, if key does not work
@return world (str) - Extracted "World War 1" if text found outside of the war key
@return bYear (str) - Extracted 2-digit birth year, for key "19"

@author Mike
'''
def getKV_Map(fileName, id, cemetery):
    with open(fileName, 'rb') as file:
        imgTest = file.read()
        bytesTest = bytearray(imgTest)
        print('Image loaded', id)
    client = boto3.client('textract', region_name='us-east-1')
    response = client.analyze_document(Document={'Bytes': bytesTest}, FeatureTypes=['FORMS'])
    blocks = response['Blocks']
    keyMap = {}
    civil = ""
    world = ""
    cem = ""
    bYear = ""
    valueMap = {}
    blockMap = {}
    flag = False
    flag2 = False
    stopFlag = False
    for block in blocks:
        try:
            if block['Text'] == "Civil War":
                civil = "Civil War"
            if block['Text'] == 'World War':
                world = "World War"
            if block['Text'] == cemetery:
                cem = cemetery
                stopFlag = True
            if fuzz.partial_ratio(block['Text'].lower(), cemetery.lower()) > 80:
                cem = cemetery
                stopFlag = True
            if flag:
                if len(block['Text']) == 4:
                    bYear = block['Text'][:2]
                    if block['Text'][-2:].lower() == "in":
                        flag2 = True
                elif block['Text'].lower() == "in":
                    flag2 = True
                else:
                    bYear = block['Text']
                flag = False
            elif block['BlockType'] == "WORD" and block['Text'] == "19":
                flag = True
            elif block['BlockType'] == "WORD" and block['Text'][:2] == "19" \
                and len(block['Text']) > 2:
                if len(block['Text']) == 4:
                    if block['Text'][-2:].isnumeric():
                        bYear = block['Text'][-2:]
                elif len(block['Text']) == 3:
                    pass
                elif len(block['Text']) == 5:
                    temp = re.sub(r'[^0-9]', '', block['Text'])
                    if temp[-2:].isnumeric():
                        bYear = temp[-2:]
            elif flag2 and not stopFlag:
                cem = block['Text']
                flag2 = False
            elif block['BlockType'] == "WORD" and block['Text'].lower() == "in":
                flag2 = True
        except KeyError:
            pass
        blockID = block['Id']
        blockMap[blockID] = block
        if block['BlockType'] == "KEY_VALUE_SET":
            if 'KEY' in block['EntityTypes']:
                keyMap[blockID] = block
            else:
                valueMap[blockID] = block
    return keyMap, valueMap, blockMap, civil, cem, world, bYear


'''
Establishes relationships between keys and their corresponding values by analyzing 
block relationships and types.

@param keyMap (dict) - Dictionary containing blocks identified as keys
@param valueMap (dict) - Dictionary containing blocks identified as values
@param blockMap (dict) - Dictionary containing all blocks

@return kvs (defaultdict(list)) - Dictionary of keys and their corresponding list of values

@author Mike
'''
def getKV_Relationship(keyMap, valueMap, blockMap):
    kvs = defaultdict(list)
    for blockID, keyBlock in keyMap.items():
        valueBlock = findValueBlock(keyBlock, valueMap)
        key = get_text(keyBlock, blockMap)
        key = re.sub(r'[^a-zA-Z0-9\s]', '', key)
        val = get_text(valueBlock, blockMap)
        kvs[key].append(val)
    return kvs


'''
Identifies the value block associated with a given key block by analyzing their 
relationships.

@param keyBlock (dict) - Dictionary containing blocks identified as keys
@param valueMap (dict) - Dictionary containing blocks identified as values

@return valueBlock (dict) - The value block corresponding to the given key block

@author Mike
'''
def findValueBlock(keyBlock, valueMap):
    for relationship in keyBlock['Relationships']:
        if relationship['Type'] == 'VALUE':
            for valueID in relationship['Ids']:
                valueBlock = valueMap[valueID]
    return valueBlock


'''
Extracts and concatenates text from a block or a collection of blocks.

@param result (dict) - Block from which text needs to be extracted
@param blockMap (dict) - Dictionary containing all blocks

@return text (str) - Text extracted from the specified block or collection of blocks

@author Mike
'''
def getText(result, blockMap):
    text = ''
    if 'Relationships' in result:
        for relationship in result['Relationships']:
            if relationship['Type'] == 'CHILD':
                for childID in relationship['Ids']:
                    word = blockMap[childID]
                    if word['BlockType'] == 'WORD':
                        text += word['Text'] + ' '
                    if word['BlockType'] == 'SELECTION_ELEMENT':
                        if word['SelectionStatus'] == 'SELECTED':
                            text += 'X '
    return text


'''
Prints the key-value pairs in a readable format. Useful for debugging 
or presenting the extracted data.

@param kvs (defaultdict(list)) - Dictionary containing the key-value pairs to be printed

@author Mike
'''
def printKVS(kvs):
    for key, value in kvs.items():
        print(key, ":", value)  
    print("\n")


'''
Searches for and retrieves the value associated with a specified key from the dictionary 
of key-value pairs. Compares strings for equality.

@param kvs (defaultdict(list)) - Dictionary containing key-value pairs
@param searchKey (str) - The key for which the value needs to be retrieved for

@return value (str or None) - The value associated with the searchKey, if found. 
                              Otherwise, None

@author Mike
''' 
def searchValue(kvs, searchKey):
    for key, value in kvs.items():
        key = str.rstrip(key)
        if key.upper() == searchKey.upper():
            return value


'''
Searches for and retrieves the value associated with a specified key from the dictionary 
of key-value pairs. Uses REGEX instead of comparing strings.

@param kvs (defaultdict(list)) - Dictionary containing key-value pairs
@param searchKey (str) - The key for which the value needs to be retrieved for

@return value (str or None) - The value associated with the searchKey, if found. 
                            Otherwise, None

@author Mike
''' 
def searchValueX(kvs, searchKey):
    searchPattern = r'\b' + re.escape(searchKey) + r'\b' 
    for key, value in kvs.items():
        if re.search(searchPattern, key, re.IGNORECASE):
            return value


'''
Parses and formats a full name into its constituent parts: first name, middle name, 
last name, and suffix. The function handles various naming conventions and appends 
the formatted name components to a list.

@param finalVals (list) - List where the processed name components will be appended
@param value (str) - The extracted raw name to be processed

@author Mike
''' 
def nameRule(finalVals, value):
    CONSTANTS.force_mixed_case_capitalization = True
    name = HumanName(value)
    name.capitalize()
    firstName = name.first
    middleName = name.middle
    lastName = name.last
    suffix = name.suffix
    suffi = ["Jr.", "Sr.", "I", "II", "III", "IV", "V"]
    temp = value.replace("Jr.", "").replace("Sr.", "").replace("I", "").replace("II", "")\
        .replace("III", "").replace("IV", "").replace("V", "")
    if ("," in value and "." in firstName):
        if len(middleName) > 0:
            middleName = name.first
            firstName = name.middle
        else:
            lastName = name.last
            firstName = name.first
    elif ("." in firstName):
        if len(middleName) > 0:
            lastName = name.first
            firstName = name.middle
            middleName = name.last
        else:
            lastName = name.first
            firstName = name.last
    elif ("." in lastName):
        if len(middleName) == 0:
            lastName = name.last
            firstName = name.first
        else:
            lastName = name.first
            middleName = name.last
            firstName = name.middle
    elif len(suffix) > 0 and not middleName:
        suffix = suffix.replace(", ", "")
        middleName = suffix.replace("Sr.", "").replace("Jr.", "").replace("I", "").replace("II", "")\
        .replace("III", "").replace("IV", "").replace("V", "")
        suffix = suffix.replace(middleName, "")
    elif len(suffix) > 0 and middleName:
        suffix = suffix.replace(", ", "")
        suffix = suffix.replace(middleName, "")
    if value in suffi and "." not in temp and "," not in temp:
        if len(middleName) > 0:
            firstName = name.middle
            middleName = name.last
            lastName = name.first
        else:
            firstName = name.last
            lastName = name.first
    elif "." not in temp and "," not in temp:
        if len(middleName) > 0:
            firstName = name.middle
            middleName = name.last
            lastName = name.first
        else:
            lastName = name.first
            firstName = name.last
    if len(middleName) > 2:
        middleName = middleName.replace(".", "")
    dots = 0
    for x in firstName:
        if (x == "."):
            dots+=1
    if (dots == 2):
        firstName = firstName.upper()
    else:
        firstName = firstName.replace(".", "")
    if suffix:
        if not "." in suffix:
            suffix += "."
    finalVals.append(lastName.replace(",", "").replace(".", ""))
    finalVals.append(firstName.replace(",", ""))
    finalVals.append(middleName.replace(",", "."))
    finalVals.append(suffix.replace(",", "."))


'''
Processes and formats birth date information. Handles partial and full dates, converting 
them into a consistent format.

@param birth (str) - Extracted raw full birth date
@param bYear (str) - Birth year, particularly if only the year is provided
@param birthYYFlag (bool) - Flag indicating if the birth year is provided as a two-digit 
                            number

@return birth (str) - Cleaned and formatted full birth date
@return bYear (str) - Birth year, seperated from full date if applicable
@return birthYYFlag (bool) - Updated flag indicating if the birth year is provided as 
                             a two-digit number

@author Mike
''' 
def parseBirth(birth, bYear, birthYYFlag):
    pattern = r'([A-Za-z]+),\s*(\d{4})'
    match = re.match(pattern, birth.strip())
    if birth[-1:] == " ":
        birth = birth[:-1]
    if match:
        bYear = match.group(2)
        birth = ""
    elif "/" in birth:
        year = birth.split('/')[-1]
        year = year.replace(" ", "")
        if len(year) == 4:
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) == 2:
            birthYYFlag = True
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")[-2:]
            birth = birth.strftime("%m/%d/%Y")[:-2]
        elif len(year) == 6 or len(year) == 5:
            bYear = year[-4:]
            birth = birth.replace(",", " ")
            birth = dateparser.parse(birth)
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) > 6:
            temp = year[:4]
            if temp.isnumeric():
                deathNoSpaces = birth.replace(' ', '')
                startIndex = deathNoSpaces.find(year[4:])
                numSpacesBefore = birth[:startIndex].count(' ')
                adjustedStartIndex = startIndex + numSpacesBefore
                birth = birth[:adjustedStartIndex].strip()
                birth = dateparser.parse(birth)
                birth = birth.strftime("%m/%d/%Y")
                bYear = temp
            else:
                birth = ""
        else:
            birth = ""
            bYear = ""
    elif "," in birth:
        year = birth.split(',')[-1]
        year = year.replace(" ", "")
        if len(year) == 4:
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) == 2:
            birthYYFlag = True
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")[-2:]
            birth = birth.strftime("%m/%d/%Y")[:-2]
        elif len(year) == 6 or len(year) == 5:
            bYear = year[-4:]
            birth = birth.replace(",", " ")
            birth = dateparser.parse(birth)
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) > 6:
            temp = year[:4]
            if temp.isnumeric():
                deathNoSpaces = birth.replace(' ', '')
                startIndex = deathNoSpaces.find(year[4:])
                numSpacesBefore = birth[:startIndex].count(' ')
                adjustedStartIndex = startIndex + numSpacesBefore
                birth = birth[:adjustedStartIndex].strip()
                birth = dateparser.parse(birth)
                birth = birth.strftime("%m/%d/%Y")
                bYear = temp
            else:
                birth = ""
        else:
            birth = ""
    elif " " in birth:
        year = birth.split(' ')[-1]
        year = year.replace(" ", "")
        if len(year) == 4:
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) == 2:
            birthYYFlag = True
            birth = dateparser.parse(birth)
            bYear = birth.strftime("%Y")[-2:]
            birth = birth.strftime("%m/%d/%Y")[:-2]
        elif len(year) == 6 or len(year) == 5:
            bYear = year[-4:]
            birth = birth.replace(",", " ")
            birth = dateparser.parse(birth)
            birth = birth.strftime("%m/%d/%Y")
        elif len(year) > 6:
            temp = year[:4]
            if temp.isnumeric():
                deathNoSpaces = birth.replace(' ', '')
                startIndex = deathNoSpaces.find(year[4:])
                numSpacesBefore = birth[:startIndex].count(' ')
                adjustedStartIndex = startIndex + numSpacesBefore
                birth = birth[:adjustedStartIndex].strip()
                birth = dateparser.parse(birth)
                birth = birth.strftime("%m/%d/%Y")
                bYear = temp
            else:
                birth = ""
        else:
            birth = ""
    return birth, bYear, birthYYFlag


'''
Processes and formats death date information. Handles partial and full dates, converting 
them into a consistent format.

@param death (str) - Extracted raw full death date 
@param dYear (str) - Death year, particularly if only the year is provided
@param deathYYFlag (bool) - Flag indicating if the death year is provided as a two-digit 
                            number
 
@return death (str) - Cleaned and formatted full death date
@return dYear (str) - Death year, seperated from full date if applicable
@return deathYYFlag (bool) - Updated flag indicating if the death year is provided as a 
                             two-digit number

@author Mike
''' 
def parseDeath(death, dYear, deathYYFlag):
    pattern = r'([A-Za-z]+),\s*(\d{4})'
    match = re.match(pattern, death.strip())
    if death[-1:] == " ":
        death = death[:-1]
    if match:
        dYear = match.group(2)
        death = ""
    elif "/" in death:
        year3 = death.split('/')[-1]
        year3 = year3.replace(" ", "")
        if len(year3) == 4:
            death = dateparser.parse(death)
            dYear = death.strftime("%Y")
            death = death.strftime("%m/%d/%Y")
        elif len(year3) == 2:
            deathYYFlag = True
            death = dateparser.parse(death)
            dYear = death.strftime("%Y")[-2:]
            death = death.strftime("%m/%d/%Y")[:-2]
        elif len(year3) == 6 or len(year3) == 5:
            dYear = year3[-4:]
            death = death.replace(",", " ")
            death = dateparser.parse(death)
            death = death.strftime("%m/%d/%Y")
        elif len(year3) > 6:
            temp = year3[:4]
            if temp.isnumeric():
                deathNoSpaces = death.replace(' ', '')
                startIndex = deathNoSpaces.find(year3[4:])
                numSpacesBefore = death[:startIndex].count(' ')
                adjustedStartIndex = startIndex + numSpacesBefore
                death = death[:adjustedStartIndex].strip()
                death = dateparser.parse(death)
                death = death.strftime("%m/%d/%Y")
                dYear = temp
            else:
                death = ""
        else:
            death = ""
    elif "," in death:
        year3 = death.split(',')[-1]
        year3 = year3.replace(" ", "")
        if len(year3) == 4:
            death = dateparser.parse(death)
            if death:
                dYear = death.strftime("%Y")
                death = death.strftime("%m/%d/%Y")
            else:
                dYear = year3
                death = ""
        elif len(year3) == 2:
            deathYYFlag = True     
            death = dateparser.parse(death)
            dYear = death.strftime("%Y")[-2:]
            death = death.strftime("%m/%d/%Y")[:-2]
        elif len(year3) == 6 or len(year3) == 5:
            dYear = year3[-4:]
            death = death.replace(",", " ")
            death = dateparser.parse(death)
            death = death.strftime("%m/%d/%Y")
        elif len(year3) > 6:
            temp = year3[:4]
            if temp.isnumeric():
                deathNoSpaces = death.replace(' ', '')
                startIndex = deathNoSpaces.find(year3[4:])
                numSpacesBefore = death[:startIndex].count(' ')
                adjustedStartIndex = startIndex + numSpacesBefore
                death = death[:adjustedStartIndex].strip()
                death = dateparser.parse(death)
                death = death.strftime("%m/%d/%Y")
                dYear = temp
            else:
                death = ""
        else:
            death = ""
    elif " " in death:
        year3 = death.split(' ')[-1]
        year3 = year3.replace(" ", "")
        if len(year3) == 4:
            death = dateparser.parse(death)
            dYear = death.strftime("%Y")
            death = death.strftime("%m/%d/%Y")
        elif len(year3) == 2:
            deathYYFlag = True     
            death = dateparser.parse(death)
            dYear = death.strftime("%Y")[-2:]
            death = death.strftime("%m/%d/%Y")[:-2]
        elif len(year3) == 6 or len(year3) == 5:
            dYear = year3[-4:]
            death = death.replace(",", " ")
            death = dateparser.parse(death)
            death = death.strftime("%m/%d/%Y")
        elif len(year3) > 6:
            temp = year3[:4]
            if temp.isnumeric():
                deathNoSpaces = death.replace(' ', '')
                startIndex = deathNoSpaces.find(year3[4:])
                numSpacesBefore = death[:startIndex].count(' ')
                adjustedStartIndex = startIndex + numSpacesBefore
                death = death[:adjustedStartIndex].strip()
                death = dateparser.parse(death)
                death = death.strftime("%m/%d/%Y")
                dYear = temp
            else:
                death = ""
        else:
            death = ""
    return death, dYear, deathYYFlag


'''
Analyzes and formats date-related information based on various rules and conditions. 
It reconciles birth and death dates with other date-related information. Handles conditions
where missing birth and/or death dates and/or years. Attempts to correct 2-digit years
using cent date, else uses buried date, else uses war info, else uses "<" comparison and 
this comparison calls for the activation of warFlag signaling to highlight the row due to 
possible wrong century.

@param finalVals (list) - List where the processed date information will be appended
@param value (string) - Current value being processed (typically related to death date)
@param dob (string) - Date of birth
@param buried (string) - Date of burial
@param buriedYear (string) - Year of burial
@param cent (string) -  Century of the event (used for 2-digit years)
@param war (string) - War during which the person served, if applicable

@return warFlag (bool) - Flag indicating if a 2-digit year comparison was made to 
                         determine the century

@author Mike
''' 
def dateRule(finalVals, value, dob, buried, buriedYear, cent, war):
    warFlag = False
    warsFlag = False
    wars = ["World War 1", "World War 2", "Korean War", "Vietnam War", "Mexican Border War"]
    if war in wars:
        warsFlag = True
    if not buriedYear:
        buriedYear = buriedRule(buried, buriedYear, cent, warsFlag)
    birthYYFlag = False
    deathYYFlag = False
    bYear = ""
    dYear = ""
    birth = dob.replace(":", ".").replace("I", "1").replace(".", " ").replace("&", "").replace("x", "")
    if "At" in birth:
        birth = birth.split("At")
        birth = birth[0] 
    elif "AT" in birth:
        birth = birth.split("AT")
        birth = birth[0] 
    if birth[-1:] == " ":
        birth = birth[:-1]
    if birth[-1:] == ".":
        birth = birth[:-1]
    if birth[-1:] == "/":
        birth = birth[:-1]
    if "Age" in birth:
        match = re.search(r'\b\d{4}\b', birth)
        if match:
            bYear = match.group()
            birth = ""
        else:
            match = re.search(r'\b\d{2}\b', birth)
            if match:
                birth = ""
    death = value.replace(":", ".").replace("I", "1").replace(".", " ").replace("&", "").replace("x", "")
    if death[-1:] == " ":
        death = death[:-1]
    if death[-1:] == ".":
        death = death[:-1]
    if death[-1:] == "/":
        death = death[:-1]
    if "-" in birth:
        birth = birth.replace("-", "/")
    if "-" in death:
        death = death.replace("-", "/")
    if len(birth.replace(" ", "")) == 4:
        if len(death.replace(" ", "")) == 4:
            bYear = birth
            birth = ""
            dYear = death
            death = ""
        else:
            bYear = birth
            birth = ""
    elif len(death.replace(" ", "")) == 4:
        dYear = death
        death = ""
    if birth != "" and death != "":
        if dateparser.parse(birth, settings={'STRICT_PARSING': True}) != None:
            if dateparser.parse(death, settings={'STRICT_PARSING': True}) != None:
                birth, bYear, birthYYFlag = parseBirth(birth, bYear, birthYYFlag)
                death, dYear, deathYYFlag = parseDeath(death, dYear, deathYYFlag)
                if birthYYFlag and deathYYFlag:
                    if cent.isnumeric():
                        if bYear > dYear:
                            bYear = "18" + bYear
                            dYear = "19" + dYear
                        elif bYear < dYear:
                            bYear = "19" + bYear
                            dYear = "19" + dYear
                        birth = birth[:-2] + bYear
                        death = death[:-2] + dYear
                        finalVals.append(birth)
                        finalVals.append(int(bYear))
                        finalVals.append(death)
                        finalVals.append(int(dYear))
                    elif war:
                        if war in wars and dYear[0] == "0":
                            bYear = "19" + bYear
                            birth = birth[:-2] + bYear
                            dYear = "20" + dYear
                            death = death[:-2] + dYear
                            finalVals.append(birth)
                            finalVals.append(int(bYear))
                            finalVals.append(death)
                            finalVals.append(int(dYear))
                        elif war in wars and bYear > dYear:
                            bYear = "18" + bYear
                            birth = birth[:-2] + bYear
                            dYear = "19" + dYear
                            death = death[:-2] + dYear
                            finalVals.append(birth)
                            finalVals.append(int(bYear))
                            finalVals.append(death)
                            finalVals.append(int(dYear))
                        elif war in wars and bYear < dYear:
                            bYear = "19" + bYear
                            birth = birth[:-2] + bYear
                            dYear = "19" + dYear
                            death = death[:-2] + dYear
                            finalVals.append(birth)
                            finalVals.append(int(bYear))
                            finalVals.append(death)
                            finalVals.append(int(dYear))
                    elif buriedYear:
                        if buriedYear[:2] == "18" and dYear > bYear:
                            bYear = "18" + bYear
                            dYear = "18" + dYear
                        elif buriedYear[:2] == "19" and dYear < bYear:
                            bYear = "18" + bYear
                            dYear = "19" + dYear
                        elif buriedYear[:2] == "19" and dYear > bYear:
                            bYear = "19" + bYear
                            dYear = "19" + dYear
                        elif bYear < dYear:
                            bYear = "19" + bYear
                            dYear = "19" + dYear
                            warFlag = True
                        elif bYear > dYear:
                            bYear = "18" + bYear
                            dYear = "19" + dYear
                            warFlag = True
                        birth = birth[:-2] + bYear
                        death = death[:-2] + dYear
                        finalVals.append(birth)
                        if bYear:
                            finalVals.append(int(bYear))
                        else:
                            finalVals.append("")
                        finalVals.append(death)
                        if dYear:
                            finalVals.append(int(dYear))
                        else:
                            finalVals.append("")
                    elif dYear < bYear:
                        bYear = "18" + bYear
                        birth = birth[:-2] + bYear
                        dYear = "19" + dYear
                        death = death[:-2] + dYear
                        finalVals.append(birth)
                        finalVals.append(int(bYear))
                        finalVals.append(death)
                        finalVals.append(int(dYear))
                        warFlag = True
                    elif dYear > bYear:
                        bYear = "19" + bYear
                        birth = birth[:-2] + bYear
                        dYear = "19" + dYear
                        death = death[:-2] + dYear
                        finalVals.append(birth)
                        finalVals.append(int(bYear))
                        finalVals.append(death)
                        finalVals.append(int(dYear))
                        warFlag = True
                    else:
                        finalVals.append("")
                        finalVals.append("") 
                        finalVals.append("")
                        finalVals.append("")
                    return warFlag
                elif not birthYYFlag and deathYYFlag:
                    if bYear[:2] == "19" and bYear[-2:] < dYear:
                        dYear = "19" + dYear
                    elif bYear[:2] == "19" and bYear[-2:] > dYear:
                        dYear = "20" + dYear
                    elif bYear[:2] == "18" and bYear[-2:] < dYear:
                        dYear = "18" + dYear
                    elif bYear[:2] == "18" and bYear[-2:] > dYear: 
                        dYear = "19" + dYear
                    death = death[:-2] + dYear
                    if birth:
                        finalVals.append("")
                    else:
                        finalVals.append("")
                    if bYear:
                        finalVals.append("")
                    else:
                        finalVals.append("")
                    if len(dYear) == 2:
                        dYear = "19" + dYear
                        finalVals.append(death[:-2] + dYear)
                        finalVals.append(int(dYear))
                    elif len(dYear) == 4:
                        finalVals.append(death)
                        finalVals.append(int(dYear))
                    else:
                        finalVals.append("")
                    return warFlag
                elif birthYYFlag and not deathYYFlag:
                    if dYear[:2] == "19" and dYear[-2:] < bYear:
                        bYear = "18" + bYear
                    elif dYear[:2] == "19" and dYear[-2:] > bYear:
                        bYear = "19" + bYear
                    elif dYear[:2] == "18" and dYear[-2:] < bYear:
                        bYear = "19" + bYear
                    elif dYear[:2] == "18" and dYear[-2:] > bYear: 
                        bYear = "18" + bYear
                    elif dYear[:2] == "20":
                        bYear = "19" + bYear
                    birth = birth[:-2] + bYear
                    finalVals.append(birth)
                    finalVals.append(int(bYear))
                    finalVals.append(death)
                    finalVals.append(int(dYear))
                    return warFlag
                else:
                    finalVals.append(birth)
                    finalVals.append(int(bYear))
                    finalVals.append(death)
                    finalVals.append(int(dYear))
                    return warFlag
            else:
                birth = dateparser.parse(birth)
                bYear = birth.strftime("%Y")
                birth = birth.strftime("%m/%d/%Y")
                death, dYear, deathYYFlag = parseDeath(death, dYear, deathYYFlag)
                if not deathYYFlag:
                    finalVals.append(birth)
                    if bYear:
                        finalVals.append(int(bYear))
                    else:
                        finalVals.append("")
                    finalVals.append(death)
                    if dYear:
                        finalVals.append(int(dYear))
                    else:
                        finalVals.append("")
                    return warFlag
                else:
                    if bYear[:2] == "19" and bYear[-2:] < dYear:
                        dYear = "20" + dYear
                    elif bYear[:2] == "19" and bYear[-2:] > dYear:
                        dYear = "19" + dYear
                    elif bYear[:2] == "18" and bYear[-2:] < dYear:
                        dYear = "19" + dYear
                    elif bYear[:2] == "18" and bYear[-2:] > dYear: 
                        dYear = "18" + dYear
                    death = death[:-2] + dYear
                    finalVals.append(birth)
                    if bYear:
                        finalVals.append(int(bYear))
                    else:
                        finalVals.append("")
                    finalVals.append(death)
                    if dYear:
                        finalVals.append(int(dYear))
                    else:
                        finalVals.append("")
                    return warFlag
        elif dateparser.parse(death, settings={'STRICT_PARSING': True}) != None:
            finalVals.append("")
            finalVals.append("")
            death = dateparser.parse(death)
            finalVals.append(death.strftime("%m/%d/%Y"))
            finalVals.append(int(death.strftime("%Y")))
            return warFlag
        else:
            birth, bYear, birthYYFlag = parseBirth(birth, bYear, birthYYFlag)
            death, dYear, deathYYFlag = parseDeath(death, dYear, deathYYFlag)
            if birthYYFlag and deathYYFlag:
                if cent.isnumeric() or buriedYear.isnumeric():
                    if bYear > dYear:
                        bYear = "18" + bYear
                        dYear = "19" + dYear
                    elif bYear < dYear:
                        bYear = "19" + bYear
                        dYear = "19" + dYear
                    birth = birth[:-2] + bYear
                    death = death[:-2] + dYear
                    finalVals.append(birth)
                    finalVals.append(int(bYear))
                    finalVals.append(death)
                    finalVals.append(int(dYear))
                    return warFlag
                elif dYear < bYear:
                    finalVals.append("")
                    finalVals.append(int("18" + bYear))
                    finalVals.append("")
                    finalVals.append(int("19" + dYear))
                    return warFlag
            elif birthYYFlag and not deathYYFlag:
                if dYear[:2] == "19" and dYear[-2:] < bYear:
                    bYear = "18" + bYear
                elif dYear[:2] == "19" and dYear[-2:] > bYear:
                    bYear = "19" + bYear
                elif dYear[:2] == "18" and dYear[-2:] < bYear:
                    bYear = "19" + bYear
                elif dYear[:2] == "18" and dYear[-2:] > bYear: 
                    bYear = "18" + bYear
                elif dYear[:2] == "20":
                    bYear = "19" + bYear
                birth = birth[:-2] + bYear
                finalVals.append(birth)
                if bYear: 
                    finalVals.append(int(bYear))
                else:
                    finalVals.append("")
                finalVals.append(death)
                finalVals.append(int(dYear))
                return warFlag
            elif not birthYYFlag and deathYYFlag:
                if bYear[:2] == "19" and bYear[-2:] < dYear:
                    dYear = "20" + dYear
                elif bYear[:2] == "19" and bYear[-2:] > dYear:
                    dYear = "19" + dYear
                elif bYear[:2] == "18" and bYear[-2:] < dYear:
                    dYear = "19" + dYear
                elif bYear[:2] == "18" and bYear[-2:] > dYear: 
                    dYear = "18" + dYear
                death = death[:-2] + dYear
                finalVals.append(birth)
                finalVals.append(int(bYear))
                finalVals.append(death)
                finalVals.append(int(dYear))
                return warFlag
    elif birth == "" and death != "":
        death, dYear, deathYYFlag = parseDeath(death, dYear, deathYYFlag)
        if not deathYYFlag:
            if bYear:
                finalVals.append("")
                finalVals.append(int(bYear))
            else:
                finalVals.append("")
                finalVals.append("")
            finalVals.append(death)
            if dYear:
                finalVals.append(int(dYear))
            else:
                finalVals.append("")
        else:
            if bYear:
                if bYear[:2] == "18" and bYear[-2:] < dYear:
                    dYear = "18" + dYear
                elif bYear[:2] == "18" and bYear[-2:] > dYear:
                    dYear = "19" + dYear
                elif bYear[:2] == "19" and bYear[-2:] < dYear:
                    dYear = "19" + dYear
                elif bYear[:2] == "19" and bYear[-2:] > dYear:
                    dYear = "20" + dYear
                finalVals.append("")
                finalVals.append(int(bYear))
                death = death[:-2] + dYear
                finalVals.append(death)
                finalVals.append(int(dYear))
            elif cent.isnumeric() or buriedYear.isnumeric():
                finalVals.append("")
                finalVals.append("") 
                dYear = "19" + dYear
                finalVals.append(death[:-2] + dYear)
                finalVals.append(int(dYear))
            elif war in wars:
                finalVals.append("")
                finalVals.append("") 
                dYear = "19" + dYear
                finalVals.append(death[:-2] + dYear)
                finalVals.append(int(dYear))
            else:
                finalVals.append("")
                finalVals.append("") 
                finalVals.append("")
                finalVals.append("") 
    elif birth != "" and death == "":
        birth, bYear, birthYYFlag = parseBirth(birth, bYear, birthYYFlag)
        if not birthYYFlag:
            finalVals.append(birth)
            finalVals.append(int(bYear))
            if dYear != "":
                finalVals.append("")
                finalVals.append(int(dYear))
                return warFlag
            elif buriedYear != "":
                if buriedYear < bYear[:-2]:
                    dYear = "19" + buriedYear
                else:
                    dYear = "18" + buriedYear
                finalVals.append("")
                finalVals.append(int(dYear))
                return warFlag
            else:
                finalVals.append("")
                finalVals.append("") 
        else:
            if dYear:
                if dYear[:2] == "18" and dYear[-2:] > bYear:
                    bYear = "18" + bYear
                elif dYear[:2] == "19" and dYear[-2:] < bYear:
                    bYear = "18" + bYear
                elif dYear[:2] == "19" and dYear[-2:] > bYear:
                    bYear = "19" + bYear
                elif dYear[:2] == "20" and dYear[-2:] < bYear:
                    bYear = "19" + bYear
                birth = birth[:-2] + bYear
                finalVals.append(birth)
                finalVals.append(int(bYear))
                finalVals.append("")
                finalVals.append(int(dYear))
                return warFlag
            elif cent.isnumeric():
                if bYear > cent:
                    bYear = "18" + bYear
                else:
                    bYear = "19" + bYear
                birth = birth[:-2] + bYear
                finalVals.append(birth)
                finalVals.append(int(bYear))
            else:
                finalVals.append("")
                finalVals.append("") 
                finalVals.append("")
                finalVals.append("") 
    elif birth == "" and death == "":
        if buriedYear.isnumeric():
            if len(buriedYear) == 2:
                dYear = "19" + buriedYear
            elif len(buriedYear) == 4:
                dYear = buriedYear
        elif cent.isnumeric():
            dYear = "19" + cent
        if bYear != "" and dYear != "":
            finalVals.append("")
            finalVals.append(int(bYear))
            finalVals.append("")
            finalVals.append(int(dYear))
        elif bYear != "" and dYear == "":
            finalVals.append("")
            finalVals.append(int(bYear))
            finalVals.append("")
            finalVals.append("")
        elif bYear == "" and dYear != "":
            finalVals.append("")
            finalVals.append("") 
            finalVals.append("")
            finalVals.append(int(dYear))
        else:
            finalVals.append("")
            finalVals.append("") 
            finalVals.append("")
            finalVals.append("") 
    return warFlag


'''
Processes and formats the burial year information. It adjusts the year based on the 
century and war involvement, and reconciles it with other date-related information.

@param value (str) - Extracted raw burial date
@param bYear (str) - Burial year, particularly if only the year is printed on the card
@param cent (str) - If not none, provides century data for death year
@param warsFlag (bool) - Flag indicating involvement in a 19th century war

@return buriedYear (str) - Processed and formatted burial year

@author Mike
'''
def buriedRule(value, bYear, cent, warsFlag):
    buried = value
    buried = buried.replace("I", "1").replace("-", "/")
    buriedYear = bYear
    if cent:
        buriedYear = "19" + cent
    elif buried or buriedYear:
        if dateparser.parse(buried, settings={'STRICT_PARSING': True}) == None:
            if buriedYear:
                if len(buriedYear) == 2:
                    buriedYear = buriedYear
                elif len(buriedYear) == 4:
                    buriedYear = buriedYear[2:]
                else:
                    buriedYear = ""
                if "," in buried:
                    temp = buried + " 19" + buriedYear
                    if dateparser.parse(temp, settings={'STRICT_PARSING': True}) != None:
                        buried = dateparser.parse(temp)
                        buried = buried.strftime("%m/%d/%Y")
                        buriedYear = "19" + buriedYear
                else:
                    temp = buried + ", 19" + buriedYear
                    if dateparser.parse(temp, settings={'STRICT_PARSING': True}) != None:
                        buried = dateparser.parse(temp)
                        buried = buried.strftime("%m/%d/%Y")
                        if len(bYear) == 2:
                            buriedYear = "19" + bYear
                        else:
                            buriedYear = bYear
                    else:
                        buried = ""
                        if len(bYear) == 2:
                            buriedYear = "19" + bYear
                        else:
                            buriedYear = bYear
            else:
                try:
                    beforeComma = buried.split(',', 1)[0].strip()
                    afterComma = buried.split(',', 1)[1].strip()
                    if len(afterComma) > 4:
                        buriedYear = buriedRule(beforeComma, afterComma[:4], cent, warsFlag) 
                except IndexError:
                    print("buried is weird")
                    buried = ""
        else:
            year5 = buried.split('/')[-1]
            if year5[-1:] == " ":
                year5 = year5[:-1]
            if len(year5) == 4:
                buriedYear = year5
            elif len(year5) == 2:
                if warsFlag:
                    buriedYear = "19" + year5
                else:
                    buriedYear = ""
    return buriedYear


'''
Processes and categorizes the war based on the extracted value. It standardizes different 
war names and handles abbreviations and misspellings. It also accounts for brute force 
extraction through 'civil' and 'world' parameters taken from getKV_Map().

@param value (str) - Extracted raw war name 
@param civil (str) - Specific indicator if the person participated in the Civil War
@param world (str) - Specific indicator if the person participated in World War I

@return war (str) - The standardized war name or an empty string if no matching category 
                    was found

@author Mike
'''
def warRule(value, civil, world):
    war = value
    if war:
        if war[0] == " ":
            war = war[1:]
    ww1years = ["1914", "1915", "1916", "1917", "1918"]
    if "N." in war or "N" in war:
        war = war.replace("N", "W")
    if "T" in war:
        if "1" in war:
            war = war.replace("T", "1")
        elif "I" in war:
            war = war.replace("T", "I")
        elif "L" in war:
            war = war.replace("T", "L")
        elif "l" in war:
            war = war.replace("T", "l")
    war = war.replace(" ", "")
    tempText = war.replace("-", "").replace(".", "").replace("#", "").replace(" ", "").replace(",", "")
    WW1and2_Pattern = re.compile(r'(WW1|WWI|W\.?W\.?\s?(1|I|One|ONE)|World\s*War\s*(1|I|One|ONE))\s*(and|&)?\s*'\
    r'(WW2|WWII|W\.?W\.?\s?(2|II|Two|TWO)|World\s*War\s*(2|II|Two|TWO))', re.IGNORECASE)    
    # WW1_Pattern = re.compile(r'WWI|WW1|WW\s?1|World\s*War\s*(1|I|ONE)|WorldWar1', re.IGNORECASE)
    WW1_Pattern = re.compile(r'WW1|WWI|WWl|\b1\b|W\.?W\.?\s?(1|I|l|One|ONE)?|World\s*War\s*(1|I|l|One|ONE)?|WorldWar\s*(1|I|l|One|ONE)?', re.IGNORECASE)    
    # WW2_Pattern = re.compile(r'WWII|WW2|WW11|WWLL|WWll|WW\s?2|World\s*War\s*(II|2|Two|LL|ll)|\b2\b', re.IGNORECASE)
    WW2_Pattern = re.compile(r'WW2|WWII|WWll|WW11|\b2\b|W\.?W\.?\s?(2|II|ll|Two|TWO)|World\s*War\s*(2|II|ll|Two|TWO)|WorldWar\s*(2|II|ll|Two|TWO)', re.IGNORECASE)    
    if WW1and2_Pattern.findall(tempText.upper()):
        war = "World War 1 and World War 2"
    elif (WW2_Pattern.findall(tempText.upper())):
        war = "World War 2"
    elif (WW1_Pattern.findall(tempText.upper())):
        war = "World War 1"
    # if (ww1_and_2_pattern.findall(tempText)):
    #     war = "World War 1 and 2"
    else:
        war = war.replace(".", "").replace("Calv", "").replace("Vols", "")
        if "Korea" in war:
            war = "Korean War"
        elif "Vietnam" in war:
            war = "Vietnam War"
        elif "Civil" in war or "Citil" in war or "Gettysburg" in war:
            war = "Civil War"
        elif "Spanish" in war or "Amer" in war or "American" in war:
            war = "Spanish American War"
        elif "Mexican" in war:
            war = "Mexican Border War"
        elif "Rebellion" in war:
            war = "War of the Rebellion"
        else:
            war = ""
        words = war.split()
        for x in words:
            if x in ww1years:
                war = "World War 1"
                break
    if civil != "" and war != "Civil War":
        war = "Civil War"
    if world != "" and war != "World War 1":
        war = "World War 1"
    if "Army" in war or "US" in war:
        war = ""
    return war


'''
Determines and appends the military branch based on the extracted text. It handles various 
abbreviations and naming conventions.

@param finalVals (list) - A list where the processed military branch will be appended
@param value (str) - The raw branch name
@param war (str) - A net for if war name is caught by the branch key, or misinput on the
                   registration card
                   
@author Mike

'''
def branchRule(finalVals, value, war):
    armys = ["co", "army", "inf", "infantry", "infan", "usa", "med", \
            "cav", "div", "sig", "art", "corps", "corp"]
    navys = ["hospital", "navy", "naval"]
    guards = ["113", "102d", "114", "44", "181", "250", "112"]
    branch = value
    branch = branch.replace("/", " ").replace(".", " ").replace("th", "").replace("-", "")\
        .replace(", ", " ")
    words = branch.split()
    if war in value and war != "":
        branch = ""
        value = ""
    for word in words:
        word = word.lower()
        if "air force" in branch.lower():
            branch = "Air Force"
            break
        elif "marine" in branch.lower():
            branch = "Marine Corps"
            break
        elif "air" in word:
            branch = "Air Force"
            break
        elif word in armys:
            branch = "Army"
            break
        elif word in navys:
            branch = "Navy"
            break
        elif "coast" in word:
            branch = "Coast Guard"
            break
        elif word in guards:
            branch = "National Guard"
            break
        else:
            branch = ""
    finalVals.append(value)   
    finalVals.append(branch)   


'''
Creates a record by extracting and processing information from a document file. It 
handles various data fields, including name, birth, death, military service, and others. 
It calls the specialized functions to handle parsing, cleaning, and standardizing the
extracted text from the veteran cards to be put into the excel sheet.  

@param fileName (str) - The path to the document file
@param id (int) - The id of the file, which is assigned to the record in Excel
@param cemetery (str) - The name of the cemetery where the record is from

@return finalVals (list) - A list of processed values for completed data fields of the 
                           current file
@return flag (bool) - A flag indicated the size type of the document, used for redaction
@return warFlag (bool) - A flag indicating if "<" comparison was made during date parsing,
                         used for highlighting the row for a second look
    
@author Mike
'''
def createRecord(fileName, id, cemetery):
    civil = ""
    world = ""
    war = ""
    bYear = ""
    cent = ""
    buried = ""
    warFlag = False
    cem = []
    finalVals = []
    pageReader = PyPDF2.PdfReader(open(fileName, 'rb'))
    page = pageReader.pages[0]
    pdfWriter = PyPDF2.PdfWriter()
    pdfWriter.add_page(page)
    with open("temp.pdf", 'wb') as outputPDF:
        pdfWriter.write(outputPDF)
    keyMap, valueMap, blockMap, civil, cem, world, bYear = getKV_Map("temp.pdf", id, cemetery)
    kvs = getKV_Relationship(keyMap, valueMap, blockMap)
    printKVS(kvs)
    keys = ["", "NAME", "WAR RECORD", "BORN", "19", "BURIED", "DATE OF DEATH", "WAR RECORD", "BRANCH OF SERVICE" , "IN"]
    flag = False
    flag3 = False
    if searchValueX(kvs, "Care Assigned") == None:
        flag = False
    else:
        flag = True
    dob = ""
    for x in keys:
        try:
            value = searchValueX(kvs, x)[0]
        except TypeError:
            value = "" 
        if x == "NAME":
            nameRule(finalVals, value)
        elif x == "BORN":
            dob = value
        elif x == "DATE OF DEATH":
            warFlag = dateRule(finalVals, value, dob, buried, bYear, cent, war)
        elif x == "WAR RECORD" and flag3:
            finalVals.append(value)
            finalVals.append(war)
            flag3 = False
        elif x == "WAR RECORD":
            if "&" in value:
                wars = value.split("&")
                war1 = warRule(wars[0], civil, world)
                war2 = warRule(wars[1], civil, world)
                war = war1 + " and " + war2
            elif "and" in value:
                wars = value.split("and")
                war1 = warRule(wars[0], civil, world)
                war2 = warRule(wars[1], civil, world)
                war = war1 + " and " + war2
            else:
                war = warRule(value, civil, world)
            flag3 = True
        elif x == "BRANCH OF SERVICE":
            branchRule(finalVals, value, war)
        elif x == "IN":
            try:
                cem = cem.replace(",", "").replace(".", "").replace(":", "")
                cem = cem.lower()
                first = cem[0].upper()
                cem = first + cem[1:]
                finalVals.append(cem)
            except IndexError:
                finalVals.append("")
        elif x == "19":
            if value:
                try:
                    tempCent = value.replace(",", "").replace(".", "").replace(":", "").replace("/", "").replace(" ", "")
                    for y in tempCent:
                        if y.isnumeric():
                            cent += y
                except IndexError:
                    pass
            else:
                pass
        elif x == "BURIED":
            buried = value
            if "in" in buried:
                buried = buried.replace("?", "")
                buried = buried.split("in")
                try:
                    cem = buried[1].replace(" ", "").replace(",", "").replace(".", "").replace(":", "")
                except IndexError:
                    pass
                buried = buried[0]
    return finalVals, flag, warFlag


'''
Creates a temporary record for processing 'A' and 'B' cards. It's similar to 
'createRecord' but tailored for handling these specific card types. It extracts 
and processes information from a document file and calls other functions for 
specific fields.

@param fileName (str) - The path to the document file
@param val (str) - A value indicating the type of card ('A' or 'B')
@param int (int) - The ID to be assigned to the record
@param cemetery (str) - The name of the cemetery to associate with the record

@return finalVals (list) - A list of processed values for different data fields 
                           of the record
@return flag (bool) - A flag indicating if a specific condition (e.g., special 
                      handling or a specific record type) was encountered during processing
@return warFlag (bool) - A flag indicating if war-related data was processed

@author Mike
'''
def tempRecord(fileName, val, id, cemetery):
    print("Performing Temp", id, val.upper())
    buried = ""
    civil = ""
    world = ""
    war = ""
    bYear = ""
    cent = ""
    warFlag = False
    cem = []
    finalVals = []
    pageReader = PyPDF2.PdfReader(open(fileName, 'rb'))
    page = pageReader.pages[0]
    pdfWriter = PyPDF2.PdfWriter()
    pdfWriter.add_page(page)
    with open("temp.pdf", 'wb') as outputPDF:
        pdfWriter.write(outputPDF)
    keyMap, valueMap, blockMap, civil, cem, world, bYear = getKV_Map("temp.pdf", id, cemetery)
    kvs = getKV_Relationship(keyMap, valueMap, blockMap)
    printKVS(kvs)
    keys = ["", "NAME", "WAR RECORD", "BORN", "19", "DATE OF DEATH", "WAR RECORD", "BRANCH OF SERVICE" , "IN"]
    flag = False
    flag3 = False
    if searchValueX(kvs, "Care Assigned") == None:
        flag = False
    else:
        flag = True
    dob = ""
    for x in keys:
        try:
            value = searchValue(kvs, x)[0]
        except TypeError:
            value = "" 
        if x == "NAME":
            nameRule(finalVals, value)
        elif x == "":
            value = value.replace(" ", "")
            buried = value[-2:]
        elif x == "BORN":
            dob = value
        elif x == "DATE OF DEATH":
            warFlag = dateRule(finalVals, value, dob, buried, bYear, cent, war)
        elif x == "WAR RECORD" and flag3:
            finalVals.append(value)
            finalVals.append(war)
            flag3 = False
        elif x == "WAR RECORD":
            if "&" in value:
                wars = value.split("&")
                war1 = warRule(wars[0], civil, world)
                war2 = warRule(wars[1], civil, world)
                war = war1 + " and " + war2
            elif "and" in value:
                wars = value.split("and")
                war1 = warRule(wars[0], civil, world)
                war2 = warRule(wars[1], civil, world)
                war = war1 + " and " + war2
            else:
                war = warRule(value, civil, world)
            flag3 = True
        elif x == "BRANCH OF SERVICE":
            branchRule(finalVals, value, war)
        elif x == "IN":
            try:
                cem = cem.replace(",", "").replace(".", "").replace(":", "")
                cem = cem.lower()
                first = cem[0].upper()
                cem = first + cem[1:]
                finalVals.append(cem)
            except IndexError:
                finalVals.append("")
        elif x == "19":
            if value:
                try:
                    cent = value.replace(",", "").replace(".", "").replace(":", "").replace("/", "").replace(" ", "")
                except IndexError:
                    pass
            else:
                pass
        elif x == "BURIED":
            buried = value
            if "in" in buried:
                buried = buried.replace("?", "")
                buried = buried.split("in")
                try:
                    cem = buried[1].replace(" ", "").replace(",", "").replace(".", "").replace(":", "")
                except IndexError:
                    pass
                buried = buried[0]
    return finalVals, flag, warFlag


'''
Merges two sets of record values, typically from 'A' and 'B' cards, into a 
single record. It compares and chooses the most appropriate value for each 
field and handles row highlighting in the output based on specific conditions.

@param vals1 (list) - The first set of record values
@param vals2 (list) - The second set of record values
@param rowIndex (int) - The row index in the output where the merged record will be placed
@param id (int) - The ID assigned to the merged record
@param warFlag (bool) - A flag indicating if war-related data was processed

@author Mike
'''
def mergeRecords(vals1, vals2, rowIndex, id, warFlag):
    counter = 1
    def length(item):
        return len(str(item))
    mergedArray = [max(a, b, key=length) for a, b in zip(vals1, vals2)] 
    worksheet.cell(row=rowIndex, column=counter, value=id)
    counter += 1
    for x in mergedArray:
        worksheet.cell(row=rowIndex, column=counter, value=x)
        counter += 1
    # Orange if merged record and no issues
    highlightColor = PatternFill(start_color="E3963E", end_color="E3963E", fill_type="solid")
    for colIndex in range(2, 15):
        cell = worksheet.cell(row=rowIndex, column=colIndex)
        cell.fill = highlightColor
        cell = worksheet.cell(row=rowIndex, column=16)
        cell.fill = highlightColor
    # Teal if merged, and record adjusted using comparison
    if warFlag:
        highlightColor = PatternFill(start_color="00FFB9", end_color="00FFB9", fill_type="solid")
        for colIndex in range(2, 15):
            cell = worksheet.cell(row=rowIndex, column=colIndex)
            cell.fill = highlightColor
            cell = worksheet.cell(row=rowIndex, column=16)
            cell.fill = highlightColor
    # Blue if merged, and record has no DOD
    if (worksheet[f'{"I"}{rowIndex}'].value == ""):
        highlightColor = PatternFill(start_color="00C6FF", end_color="00C6FF", fill_type="solid")
        for colIndex in range(2, 15):
            cell = worksheet.cell(row=rowIndex, column=colIndex)
            cell.fill = highlightColor
            cell = worksheet.cell(row=rowIndex, column=16)
            cell.fill = highlightColor
    # Pink if merged, and cemetery does not match
    if (worksheet[f'{"N"}{rowIndex}'].value) != cemetery:
        highlightColor = PatternFill(start_color="FC80AC", end_color="FC80AC", fill_type="solid")
        for colIndex in range(2, 15):
            cell = worksheet.cell(row=rowIndex, column=colIndex)
            cell.fill = highlightColor
            cell = worksheet.cell(row=rowIndex, column=16)
            cell.fill = highlightColor
    # Red is merged record and last name letter does not match
    if worksheet[f'{"B"}{rowIndex}'].value[0] != letter:
        highlightColor = PatternFill(start_color="D2042D", end_color="D2042D", fill_type="solid")
        for colIndex in range(2, 15):
            cell = worksheet.cell(row=rowIndex, column=colIndex)
            cell.fill = highlightColor
            cell = worksheet.cell(row=rowIndex, column=16)
            cell.fill = highlightColor
    # Dark Green if merged, and first and last name are the same, bug
    if (worksheet[f'{"B"}{rowIndex}'].value) == (worksheet[f'{"C"}{rowIndex}'].value):
        highlightColor = PatternFill(start_color="0B8C36", end_color="0B8C36", fill_type="solid")
        for colIndex in range(2, 15):
            cell = worksheet.cell(row=rowIndex, column=colIndex)
            cell.fill = highlightColor
            cell = worksheet.cell(row=rowIndex, column=16)
            cell.fill = highlightColor
       
       
'''
Cleans and renames file names in a given letter and cemetery for inconsistencies 
when scanning. It standardizes the naming convention of files and ensures correct ordering.

@param cemetery (str) - The name of the cemetery
@param letter (str) - The letter or section of the cemetery being processed
@param namePath (str) - The path where the files are located
@param cemPath (str) - The path to the cemetery directory
@param counterA (int) - The starting counter for naming files
@param isFirstFile (bool) - Flag indicating if the current file is the first in the series

@author Mike
'''     
def clean(cemetery, letter, namePath, cemPath, counterA, isFirstFile):
    pdfFiles = sorted(os.listdir(namePath))
    letters = sorted([folder for folder in os.listdir(cemPath) if os.path.isdir(os.path.join(cemPath, folder))])
    try:
        currentLetterIndex = letters.index(letter)
    except ValueError:
        return
    folderBeforeIndex = max(0, currentLetterIndex - 1)
    folderBefore = letters[folderBeforeIndex]
    pdfFilesBefore = sorted([file for file in os.listdir(os.path.join(cemPath, folderBefore)) if file.lower().endswith('.pdf')])
    maxCounter = 0
    for file in pdfFilesBefore:
        match = re.search(r'\d+', file)
        if match:
            number = int(match.group())
            maxCounter = max(maxCounter, number)
    counter = maxCounter + 1
    for x in pdfFiles:
        if isFirstFile:
            counter = counterA 
        if "a.pdf" in x:
            newName = f"{cemetery}{letter}{counter:05d}a.pdf"
            os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
            counter -= 1
        elif "b.pdf"  in x:
            newName = f"{cemetery}{letter}{counter:05d}b.pdf"
            os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
        elif "a redacted" in x:
            newName = f"{cemetery}{letter}{counter:05d}a redacted.pdf"
            os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
            counter -= 1
        elif "b redacted" in x:
            newName = f"{cemetery}{letter}{counter:05d}b redacted.pdf"
            os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
            counter -= 1
        elif "redacted" in x:
            newName = f"{cemetery}{letter}{counter:05d} redacted.pdf"
            os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
            counter -= 1
        else:    
            newName = f"{cemetery}{letter}{counter:05d}.pdf"
            os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
        counter += 1
        isFirstFile = False


'''
Finds the next empty row in the worksheet to begin processing at that index.

@param worksheet - The worksheet object being operated on

@return worksheet.max_row + 1 (int) - The row index of the next empty row in the worksheet

@author Mike
'''  
def findNextEmptyRow(worksheet):
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=1):
        if row[0].value is None:
            return row[0].row
    return worksheet.max_row + 1 


'''
The main function that controls files sent to be processed as well as 
the data that is put into the excel sheet. It sets up the current working 
directory, the current sheet in the spreadsheet, and loops through all the 
images in a set letter in a set cemetery. This function is the primary 
controller for processing, saving, and handling of records.

Main functions that controls files sent to be processed as well as 
the data that is put into the excel sheet. Sets up the current working 
directory as well as the current sheet in the excel spreadsheet. Loops
through all the images in a set Letter in a set Cemetery, this is for 
damage control. 

The loop finds the last record recorded in the excel sheet
and then starts indexing from there. Calls createRecord is normal card, 
calls tempRecord and mergeRecord for A and B cards.

calls redactImage for every card processed. Places the record ID and a 
hyperlink to the redacted image in the excel sheet during processing.
Saves worksheet after every image is done processing to reduce loss of 
data upon errors.

Controls highlighting of entire row of record upon different conditions 
that are caught by the program. 

@author Mike
'''
def main():
    global cemSet
    global miscSet
    global jewishSet
    networkFolder = r"\\ucclerk\pgmdoc\Veterans"
    os.chdir(networkFolder)
    cemeterys = []
    for x in os.listdir():
        cemeterys.append(x)
    miscs = []
    for x in os.listdir("Misc"):
        miscs.append(x)
    jewishs = []
    for x in os.listdir("Jewish"):
        jewishs.append(x)
    cemSet = set(cemeterys)
    miscSet = set(miscs)
    jewishSet = set(jewishs)
    workbook = openpyxl.load_workbook('Veterans.xlsx')
    global cemetery
    cemetery = "Evergreen"
    cemPath = os.path.join(networkFolder, fr"Cemetery\{cemetery}")
    # if cemetery == "Misc":
    #     miscCem = "Arlington"
    #     cemPath = os.path.join(cemPath, miscCem)
    global letter
    letter = "A"
    namePath = letter
    namePath = os.path.join(cemPath, namePath)
    initialCount = 42772
    pathA = ""
    rowIndex = 2
    global worksheet
    worksheet = workbook[cemetery]
    warFlag = False
    
    uppercaseAlphabet = [chr(i) for i in range(ord('A'), ord('Z') + 1)]
    firstFileFlag = True
    for letter in uppercaseAlphabet:
        try:
            namePath = os.path.join(cemPath, letter)
            print(letter)
            clean(cemetery, letter, namePath, cemPath, initialCount, firstFileFlag)
            firstFileFlag = False
        except FileNotFoundError:
            continue
    
    pdfFiles = sorted(os.listdir(namePath))
    initialID = 1
    for y in range(len(pdfFiles)):
        warFlag = False
        filePath = os.path.join(namePath, pdfFiles[y])
        rowIndex = findNextEmptyRow(worksheet)
        try:
            id = worksheet[f'{"A"}{rowIndex-1}'].value + 1
        except TypeError:
            id = initialID
        if "output" in pdfFiles[y] or "redacted" in pdfFiles[y]:
            continue
        else:
            string = pdfFiles[y][:-4]
            string = string.split(letter) 
            string = string[-1].lstrip('0')
            if "a" not in string and "b" not in string:
                if id != int(string.replace("a", "").replace("b", "")):
                    continue
                vals, flag, warFlag = createRecord(filePath, id, cemetery)
                redactedFile = redact(filePath, flag, cemetery, letter)
                linkText = "PDF Image"
                worksheet.cell(row=rowIndex, column=15).value = linkText
                worksheet.cell(row=rowIndex, column=15).font = Font(underline="single", color="0563C1")
                worksheet.cell(row=rowIndex, column=15).hyperlink = redactedFile
                counter = 1
                worksheet.cell(row=rowIndex, column=counter, value=id)
                counter += 1
                for x in vals:
                    worksheet.cell(row=rowIndex, column=counter, value=x)
                    counter += 1
                # Grey if record adjusted using comparison
                if warFlag:
                    highlightColor = PatternFill(start_color="899499", end_color="899499", fill_type="solid")
                    for colIndex in range(2, 15):
                        cell = worksheet.cell(row=rowIndex, column=colIndex)
                        cell.fill = highlightColor
                        cell = worksheet.cell(row=rowIndex, column=16)
                        cell.fill = highlightColor
                # Purple if cemetery does not match
                if (worksheet[f'{"N"}{rowIndex}'].value) != cemetery:
                    highlightColor = PatternFill(start_color="CF9FFF", end_color="CF9FFF", fill_type="solid")
                    for colIndex in range(2, 15):
                        cell = worksheet.cell(row=rowIndex, column=colIndex)
                        cell.fill = highlightColor
                        cell = worksheet.cell(row=rowIndex, column=16)
                        cell.fill = highlightColor
                # Light Blue if record has no DOD
                if (worksheet[f'{"I"}{rowIndex}'].value) == "":
                    highlightColor = PatternFill(start_color="A7C7E7", end_color="A7C7E7", fill_type="solid")
                    for colIndex in range(2, 15):
                        cell = worksheet.cell(row=rowIndex, column=colIndex)
                        cell.fill = highlightColor
                        cell = worksheet.cell(row=rowIndex, column=16)
                        cell.fill = highlightColor
                # Yellow if record last name does not match
                try:
                    if (worksheet[f'B{rowIndex}'].value)[0] != letter:
                        highlightColor = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        for colIndex in range(2, 15):
                            cell = worksheet.cell(row=rowIndex, column=colIndex)
                            cell.fill = highlightColor
                            cell = worksheet.cell(row=rowIndex, column=16)
                        cell.fill = highlightColor
                except IndexError:
                        highlightColor = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        for colIndex in range(2, 15):
                            cell = worksheet.cell(row=rowIndex, column=colIndex)
                            cell.fill = highlightColor
                            cell = worksheet.cell(row=rowIndex, column=16)
                            cell.fill = highlightColor
                id += 1
                rowIndex += 1
            else:
                if id != int(string.replace("a", "").replace("b", "")):
                    continue
                else:
                    if "a" in string:
                        if (filePath.replace("a.pdf", "") in pdfFiles):
                            continue
                        pathA = filePath
                        vals1, flag, warFlag = tempRecord(filePath, "a", id, cemetery)
                        redactedFile = redact(filePath, flag, cemetery, letter)
                    if "b" in string:
                        if (filePath.replace("b.pdf", "") in pdfFiles):
                            continue
                        vals2, flag, warFlagB = tempRecord(filePath, "b", id, cemetery)
                        if not warFlag or not warFlagB:
                            warFlag = False
                        else:
                            warFlag = True
                        redactedFile = redact(filePath, flag, cemetery, letter)
                        mergeRecords(vals1, vals2, rowIndex, id, warFlag)
                        mergeImages(pathA, filePath, cemetery, letter)
                        linkText = "PDF Image"
                        worksheet.cell(row=rowIndex, column=15).value = linkText
                        worksheet.cell(row=rowIndex, column=15).font = Font(underline="single", color="0563C1")
                        worksheet.cell(row=rowIndex, column=15).hyperlink = redactedFile.replace("b redacted.pdf", " redacted.pdf")
                        id += 1
                        rowIndex += 1
        workbook.save('Veterans.xlsx')

if __name__ == "__main__":
    main()