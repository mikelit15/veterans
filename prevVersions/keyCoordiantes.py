import cv2
import easyocr
from PIL import Image
import numpy as np
import openpyxl
import pytesseract
import os

def preProcess(image):
    img = cv2.imread(image)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    thresh = cv2.adaptiveThreshold(~gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 7, -2)
    kernal = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 5))
    thresh = cv2.dilate(thresh, kernal, iterations=2)
    thresh = cv2.erode(thresh, kernal, iterations=1)
    horz = np.copy(thresh)
    (_, horzcol) = np.shape(horz)
    horzSize = int(horzcol/26)
    horzStructure = cv2.getStructuringElement(cv2.MORPH_RECT, (horzSize, 1))
    horz = cv2.erode(horz, horzStructure, iterations=1)
    horz = cv2.dilate(horz, horzStructure, iterations=1)
    outimage = cv2.bitwise_and(~gray, ~horz)
    outimage = ~outimage
    edges = cv2.Canny(outimage, 50, 150, apertureSize=3)
    minLineLength = 5
    maxLineGap = 10
    lines = cv2.HoughLinesP(edges, 1, np.pi/180, 50, minLineLength, maxLineGap)
    for x1, y1, x2, y2 in lines[0]:
        cv2.line(outimage, (x1, y1), (x2, y2), (0, 255, 0), 2)
    cv2.imwrite('output.png', outimage)
    # Load the image
    # image = cv2.imread(imagePath)

    # # Convert the image to grayscale
    # grayImage = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # # Perform text detection using pytesseract
    # customConfig = r'--oem 3 --psm 6'  # OCR Engine Mode: 3 (Default) and Page Segmentation Mode: 6 (Assume a single uniform block of text)
    # results = pytesseract.image_to_data(grayImage, config=customConfig, output_type=pytesseract.Output.DICT)

    # # Iterate over each detected word
    # for i in range(len(results['text'])):
    #     # Extract word and bounding box information
    #     word = results['text'][i]
    #     x, y, w, h = results['left'][i], results['top'][i], results['width'][i], results['height'][i]

    #     # Check if the word is non-empty and small based on the threshold
    #     smallTextThreshold = 370  # Adjust as needed
    #     if word and w < smallTextThreshold and h < smallTextThreshold:
    #         # Invert the color of the small text (assuming a white background)
    #         invertedText = cv2.bitwise_not(image[y:y+h, x:x+w])

    #         # Replace the small text region with the inverted color
    #         image[y:y+h, x:x+w] = invertedText

    # # Display the resulting image
    # resizeFactor = 0.3
    # smallerImage = cv2.resize(image, None, fx=resizeFactor, fy=resizeFactor)
    # cv2.imshow('Image with small text removed', smallerImage)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()

def easyOCR():
    imagePath = "output.png"
    reader = easyocr.Reader(['en'])
    text = reader.readtext(imagePath)
    nameCoordinates = []
    keys = [[1, "name"], [2, "serial"], [3, "home"], [4, "kin"], [5, "address"], [6, "born"], 
            [7, "at"], [8, "death"], [9, "place"], [10, "buried"], [11, "19"], [12, "in"], 
            [13,"city"], [14, "county"], [15, "division"], [16, "section"], [17, "lot"], 
            [18, "block"], [19, "grave no"], [20, "war"], [21, "service"], [22, "rank"],
            [23, "enlisted"], [24, "discharged"], [25, "information"], [26, "relationship"]]
    garb = []
    keyFound = False
    for x in text:
        print(x)
    # Loop through the OCR results
    prevBbox = None  
    currentWord = "" 
    currentBbox = None
    for key in keys:
        keyFound = False
        flag1 = False
        for bbox, word, score in text:
            if key[0] == 7:
                counter = 0
                for x in nameCoordinates:
                    if x[0] == "BORN":
                        nameCoordinates.append(["AT", [[nameCoordinates[counter][1][1][0] + 734, nameCoordinates[counter][1][0][1]], 
                            [nameCoordinates[counter][1][1][0] + 827, nameCoordinates[counter][1][0][1]],
                            [nameCoordinates[counter][1][1][0] + 827, nameCoordinates[counter][1][2][1]], 
                            [nameCoordinates[counter][1][1][0] + 734, nameCoordinates[counter][1][2][1]]]])                
                        break
                    counter += 1
                break
            if word.isupper() or word.lower() == "lot" or word.lower() == "19":
                # if (prevBbox and bbox[0][0] - prevBbox[1][0] <= 150 and prevBbox and bbox[0][0] - prevBbox[1][0] > 0):
                if (prevBbox and 
                    0 <= (bbox[0][0] - prevBbox[1][0]) <= 70 and 
                    (bbox[0][1] - prevBbox[1][1]) <= 10):                    
                    currentWord += " " + word
                    currentBbox[1] = bbox[1]
                    currentBbox[2] = bbox[2]
                else:
                    if currentWord and "UNION" not in currentWord:
                        if key[0] == 5 or key[0] == 12:
                            flag1 = True
                        if key[1].lower() == currentWord.lower():
                            # print([currentWord.upper(), currentBbox])
                            nameCoordinates.append([currentWord.upper(), currentBbox])
                            keyFound = True
                            break
                        elif key[1].lower() in currentWord.lower() and not flag1:
                            # print([currentWord.upper(), currentBbox])
                            nameCoordinates.append([currentWord.upper(), currentBbox])
                            keyFound = True
                            break 
                    currentWord = word
                    currentBbox = bbox.copy()
                prevBbox = bbox
        if not keyFound:
            garb.append(key[1])
            nameCoordinates.append([key[1].upper() + " not found", \
                                     [[2100, 3600], [2100, 3600], [2100, 3600], [2100, 3600]]])

    return garb, nameCoordinates

def tesseractOCR(rowIndex, nameCoordinates):
    imagePath = "output.png"
    image = Image.open(imagePath)
    count = 0
    counter = 1
    # nt = image.crop((419, 748, 1139, 871))
    # print(pytesseract.image_to_string(nt, lang='eng'))
    for x in nameCoordinates: 
        if(count < len(nameCoordinates)-1):
            # if (nameCoordinates[count+1][1][0] > x[1][0]) and x != [[395, 1402], [662, 1402], [662, 1458], [395, 1458]] and count != 18:
            # print(x)
            if (nameCoordinates[count+1][1][1][0] > x[1][1][0]) and "WAR" not in x[0]:
                croppedImage = image.crop((x[1][2][0] + 20, x[1][0][1]-70, 
                                            nameCoordinates[count+1][1][0][0] - 20, x[1][2][1]))
            else:
                croppedImage = image.crop((x[1][2][0] + 20, x[1][0][1]-70, 3600, x[1][2][1]))
            # croppedImage = image.crop((219, 452, 413, 520))
            if x[0] == "NAME" or x[0] == "BORN" or x[0] == "BRANCH OF SERVICE" or x[0] == "WAR RECORD":
                text = pytesseract.image_to_string(croppedImage, lang='eng')
                filteredText = text.replace("♀", "").replace("\n", "").replace("«", "") \
                                .replace("_", " ").replace("—. ", " ").replace("(", " ") \
                                .replace("CEMETERY", "").strip()
                # print(filteredText)
                worksheet.cell(row=rowIndex, column=counter, value=filteredText)
                counter += 1
            count += 1
        else:
            # print(x)
            croppedImage = image.crop((x[1][2][0] + 20, x[1][0][1]-70, 3600, x[1][2][1]))
            text = pytesseract.image_to_string(croppedImage, lang='eng')
            if x[0] == "NAME" or x[0] == "BORN" or x[0] == "BRANCH OF SERVICE" or x[0] == "WAR RECORD":
                text = pytesseract.image_to_string(croppedImage, lang='eng')
                filteredText = text.replace("♀", "").replace("\n", "").replace("«", "") \
                                .replace("_", " ").replace("—. ", " ").replace("(", " ") \
                                .replace("CEMETERY", "").strip()
                worksheet.cell(row=rowIndex, column=counter, value=filteredText)
                # print(filteredText)
workbook = openpyxl.load_workbook('Veterans.xlsx')
worksheet = workbook['Sheet1']
currentDirectory = os.getcwd()
rowIndex = 2
pngFiles = [file for file in os.listdir(currentDirectory) if file.lower().endswith('.png')]
for x in pngFiles:
    if x != "output.png":
        preProcess(x)
        garb, nameCoordinates = easyOCR()
        tesseractOCR(rowIndex, nameCoordinates)
        rowIndex += 1
workbook.save('Veterans.xlsx')
# preProcess("4.png")
# easyOCR()
# imagePath = "output.png"
# image = Image.open(imagePath)
# nt = image.crop((192, 499, 378, 549))
# print(pytesseract.image_to_string(nt, lang='eng'))
# garb, nameCoordinates = easyOCR()
# tesseractOCR(garb, nameCoordinates)
