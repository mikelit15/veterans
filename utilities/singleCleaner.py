import os
import re
import openpyxl
import shutil
import sys
import duplicates
from openpyxl.worksheet.hyperlink import Hyperlink
sys.path.append(r'C:\workspace\veterans\microsoftOCR')
from microsoftOCR import microsoftOCR

'''
Processes each cemetery by iterating through alphabetically named subdirectories and 
renaming PDF files within. It starts renaming files from a specified point, controlled 
by 'start' and 'startFlag'.

@param cemeteryName (str) - The name of the cemetery being processed.
@param basePath (str) - The base directory path where cemeteries are located.
@param initialCount (int) - The initial counter for renaming files.
@param start (int) - The starting file number from which renaming begins.
@param startFlag (bool) - A flag indicating whether the start point has been reached.

@return initialCount (int) - Updated counter after processing.
@return startFlag (bool) - Updated flag indicating if the start point was reached and 
                           surpassed.

@author Mike
'''
def processCemetery(cemeteryName, basePath, initialCount, start, startFlag):
    uppercaseAlphabet = [chr(i) for i in range(ord('A'), ord('Z') + 1)]
    firstFileFlag = True
    for letter in uppercaseAlphabet:
        namePath = os.path.join(basePath, cemeteryName, letter)
        try:
            print(f"Processing {cemeteryName} - {letter}")
            initialCount, startFlag = clean(cemeteryName, letter, namePath, os.path.join(basePath, cemeteryName), \
                initialCount, firstFileFlag, start, startFlag)
            firstFileFlag = False
        except FileNotFoundError:
            continue
    return initialCount, startFlag


'''
Renames PDF files within a specified directory, starting from a given file number. It 
skips renaming until the start condition is met.

@param cemetery (str) - The cemetery name.
@param letter (str) - The letter indicating the subdirectory within the cemetery directory.
@param namePath (str) - The path to the subdirectory being processed.
@param cemPath (str) - The path to the cemetery directory.
@param counterA (int) - The initial renaming counter.
@param isFirstFile (bool) - Flag indicating if the current file is the first to be processed.
@param start (int) - The file number from which to start renaming.
@param startFlag (bool) - Flag indicating if the start file number has been reached.

@return counter (int) - The updated file counter after renaming.
@return startFlag (bool) - Updated flag indicating if the start condition was met.

@author Mike
'''
def clean(cemetery, letter, namePath, cemPath, initialCount, isFirstFile, start, startFlag):
    pdfFiles = sorted(os.listdir(namePath))
    letters = sorted([folder for folder in os.listdir(cemPath) if os.path.isdir(os.path.join(cemPath, folder))])
    try:
        currentLetterIndex = letters.index(letter)
    except ValueError:
        return
    folderBeforeIndex = max(0, currentLetterIndex - 1)
    folderBefore = letters[folderBeforeIndex]
    pdfFilesBefore = sorted([file for file in os.listdir(os.path.join(cemPath, folderBefore)) if file.lower().endswith('.pdf')])
    maxCounter = 0
    for file in pdfFilesBefore:
        match = re.search(r'\d+', file)
        if match:
            number = int(match.group())
            maxCounter = max(maxCounter, number)
    counter = maxCounter + 1
    for x in pdfFiles:
        if counter != start and startFlag:
            isFirstFile = False
            if "a" in x[-5:]:
                pass
            else:
                counter += 1
            continue
        startFlag = False
        if isFirstFile:
            counter = initialCount 
        if "a.pdf" in x:
            newName = f"{cemetery}{letter}{counter:05d}a.pdf"
            os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
            counter -= 1
        elif "b.pdf"  in x:
            newName = f"{cemetery}{letter}{counter:05d}b.pdf"
            os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
        else:    
            newName = f"{cemetery}{letter}{counter:05d}.pdf"
            os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
        counter += 1
        isFirstFile = False
    return counter, startFlag


'''
Initiates the cleaning process for image files in all cemeteries, starting from a specific 
file ID. Adjusts the starting point for renaming based on the given ID and processes each 
cemetery found in the base path.

@param goodID (int) - The file ID from which to start the cleaning process.

- Adjusts image file names by filling in the gap when found, ID is decremented by 1
  for every file name after the gap

@author Mike
'''
def cleanImages(goodID):
    baseCemPath = r"\\ucclerk\pgmdoc\Veterans\Cemetery"
    cemeterys = [d for d in os.listdir(baseCemPath) if os.path.isdir(os.path.join(baseCemPath, d))]
    initialCount = 1
    start = goodID - 200
    startFlag = True
    for cemetery in cemeterys:
        if cemetery in ["Jewish", "Misc"]:
            subCemPath = os.path.join(baseCemPath, cemetery)
            subCemeteries = [d for d in os.listdir(subCemPath) if os.path.isdir(os.path.join(subCemPath, d))]
            for subCemetery in subCemeteries:
                print(f"Processing {cemetery} - {subCemetery}")
                initialCount, startFlag = processCemetery(subCemetery, subCemPath, initialCount, start, startFlag)
        else:
            print(f"Processing {cemetery}")
            initialCount, startFlag = processCemetery(cemetery, baseCemPath, initialCount, start, startFlag)
      
        
'''
Decrements the numerical part of file names in a directory and its subdirectories, 
starting from a specified index. This is used to maintain a sequential order after 
removing files. Files with a number greater than or equal to badID will have 
their number decremented by 1.

@param badID (str) - The path to the directory containing the files to be renamed.

- Decrements ID in redacted file name by 1, starting from specified ID

@author Mike
'''      
def cleanRedacted(badID):
    baseCemPath = r"\\ucclerk\pgmdoc\Veterans\Cemetery - Redacted"
    startIndex = badID + 1 # The image file name gets decremented starting at this ID number
    for root, dirs, files in os.walk(baseCemPath):
        for file in sorted(files):
            match = re.search(r'(\d+)', file)
            if match:
                currentIndex = int(match.group())
                if currentIndex >= startIndex:
                    newIndex = currentIndex - 1
                    newFileName = re.sub(r'\d+', f"{newIndex:05}", file, 1)  
                    os.rename(os.path.join(root, file), os.path.join(root, newFileName))
                    print(f"Renamed '{file}' to '{newFileName}'")
  
       
'''
Adjusts hyperlinks in an Excel file that contain numerical IDs, decrementing those IDs by 1
starting from a specified index. This function is used when IDs in a series are modified and 
corresponding hyperlinks need to reflect these changes. Hyperlinks with IDs greater than or 
equal to startIndex will have their ID decremented by 1.

@param badRow (int) - The row number from which to start adjusting hyperlinks.
@param cemetery (str) - The name of current cemetery, for file pathing

- Decrements ID in hyperlink address by 1 starting from specified row

@author Mike
'''     
def cleanHyperlink(badRow, cemetery):
    def decrementWithRollover(path):
        pattern = re.compile(fr"({cemetery})([A-Z])(\d+)(\s*redacted\.pdf)")
        def handleDecrement(match):
            prefix, letter, num, suffix = match.groups()
            newNum = int(num) - 2
            if newNum < 0:
                newNum = 9999  
            newNumberFormatted = f"{newNum:0{len(num)}d}"
            return f"{prefix}{letter}{newNumberFormatted}{suffix}"
        return pattern.sub(handleDecrement, path)
    excelFilePath = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"
    columnWithHyperlinks = 'O'
    startIndex = badRow 
    workbook = openpyxl.load_workbook(excelFilePath)
    sheet = workbook[cemetery]
    for row in sheet.iter_rows(min_col=openpyxl.utils.column_index_from_string(columnWithHyperlinks),
                               max_col=openpyxl.utils.column_index_from_string(columnWithHyperlinks),
                               min_row=startIndex):
        cell = row[0]
        if cell.hyperlink:
            url = cell.hyperlink.target
            url = url.replace("%20", " ")
            newURL = decrementWithRollover(url)
            if newURL != url:
                cell.hyperlink.target = newURL
                print(f"Updated hyperlink from {os.path.basename(url)} to {os.path.basename(newURL)}")
    workbook.save(excelFilePath)
    print("Hyperlinks updated.")


'''
Performs deletion of specified files and updates records in an Excel spreadsheet accordingly. 
This function is vital for keeping digital records accurate and up to date, especially after 
removing redundant or incorrect files.

@param cemetery (str) - The name of the cemetery where the deletion is to take place.
@param badID (int) - The identifier of the file to be deleted.
@param badRow (int) - The Excel row corresponding to the file to be deleted.

- Deletes badID redacted file
- Deletes badID row, shifting rows below up 1
- Adjusts ID column to correct for gap
- Openpyxl bug fixing:
    - Adjusts hyperlink ref and ID to coincide with new cell row and ID
    - Adjusts last row hyperlink and deletes empty link in unused row below final record

@author Mike
'''
def cleanDelete(cemetery, badID, badRow):
    baseRedacPath = r"\\ucclerk\pgmdoc\Veterans\Cemetery - Redacted"
    for dirPath, dirNames, fileNames in os.walk(baseRedacPath):
        for fileName in fileNames:
            if f"{badID:05d}" in fileName:
                filePath = os.path.join(dirPath, fileName)
                os.remove(filePath)
                print(f"{fileName} deleted successfully.")
    excelFilePath = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"
    workbook = openpyxl.load_workbook(excelFilePath)
    worksheet = workbook[cemetery]
    worksheet.delete_rows(badRow)
    print(f"Deleted row: {badRow} successfully.")
    for row in range(badRow, worksheet.max_row + 1):
        worksheet[f'A{row}'].value = worksheet[f'A{row}'].value - 1
        print(f"{worksheet[f'A{row}'].value + 1} changed to {worksheet[f'A{row}'].value}")
        nextRow = row + 1
        cellRef = f'O{row}'
        nextCellRef = f'O{nextRow}'
        if nextRow <= worksheet.max_row and worksheet[nextCellRef].hyperlink:
            tempHyperlink = worksheet[nextCellRef].hyperlink
            worksheet[cellRef].hyperlink = Hyperlink(ref=cellRef, target=tempHyperlink.target, display=tempHyperlink.display)
            worksheet[cellRef].value = worksheet[nextCellRef].value
            print(f"Hyperlink and value from row {nextRow} moved to row {row}.")
    if worksheet[f'O{worksheet.max_row - 1}'].hyperlink:
        lastRow = worksheet.max_row
        print(f"Adjusting the last hyperlink in row {lastRow}.")
        secondLastHyperlink = worksheet[f'O{lastRow - 1}'].hyperlink
        newTarget = re.sub(r'(\d+)(?=%\d+)', lambda m: f"{int(m.group(1)) + 1:05d}", secondLastHyperlink.target)
        worksheet[f'O{lastRow}'].hyperlink = Hyperlink(ref=f'O{lastRow}', target=newTarget, display=secondLastHyperlink.display)
        print(f"Updated hyperlink in row {lastRow} to {newTarget}.")
    workbook.save(excelFilePath)
       

'''
Adjusts the naming and organization of files when a duplicate or erroneous entry is found. 
This involves renaming and moving files as necessary to correct any issues with the digital 
records.

@param cemetery (str) - The cemetery's name where the file adjustment will occur.
@param goodID (int) - The identifier of the correct file.
@param badID (int) - The identifier of the duplicate or erroneous file to adjust.
@param goodRow (int) - The row in the Excel spreadsheet corresponding to the good file.

- Adds "a" suffix to goodID
- Changes badID to goodID, moves folders if necessary, adds "b" suffix
- Clears data in goodID row

@author Mike
'''     
def adjustImageName(cemetery, goodID, badID, goodRow):
    baseCemPath = r"\\ucclerk\pgmdoc\Veterans\Cemetery"
    goodIDFound = False
    badIDFound = False
    goodIDFilePath = ""
    badIDFilePath = ""
    for dirPath, dirNames, fileNames in os.walk(baseCemPath):
        for fileName in fileNames:
            if f"{goodID:05d}" in fileName:
                goodIDFilePath = os.path.join(dirPath, fileName)
                goodIDFound = True
            elif f"{badID:05d}" in fileName:
                badIDFilePath = os.path.join(dirPath, fileName)
                badIDFound = True
            if goodIDFound and badIDFound:
                break
        if goodIDFound and badIDFound:
            break
    if goodIDFound and badIDFound:
        goodIDPrefix = os.path.basename(os.path.dirname(goodIDFilePath))
        badIDPrefix = os.path.basename(os.path.dirname(badIDFilePath))
        newBadIDFilename = os.path.basename(badIDFilePath).replace(f"{badID:05d}", f"{goodID:05d}b").replace(badIDPrefix, goodIDPrefix)
        newBadIDFilePath = os.path.join(os.path.dirname(goodIDFilePath), newBadIDFilename)
        if not "a.pdf" in os.path.basename(goodIDFilePath):
            newGoodIDFilename = os.path.basename(goodIDFilePath).replace(".pdf", "a.pdf")
            newGoodIDFilePath = os.path.join(os.path.dirname(goodIDFilePath), newGoodIDFilename)
            os.rename(goodIDFilePath, newGoodIDFilePath)
            print(f"{os.path.basename(goodIDFilePath)} renamed to {newGoodIDFilename}")
        shutil.move(badIDFilePath, newBadIDFilePath)
        print(f"{os.path.basename(badIDFilePath)} moved and renamed to {newBadIDFilename}")
    else:
        print("GoodID or BadID file not found. Please check the IDs and directories.")
    excelFilePath = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"
    workbook = openpyxl.load_workbook(excelFilePath)
    worksheet = workbook[cemetery]
    startColumn = 'A'
    endColumn = 'O'
    startColumnIndex = openpyxl.utils.column_index_from_string(startColumn)
    endColumnIndex = openpyxl.utils.column_index_from_string(endColumn)
    for colIndex in range(startColumnIndex, endColumnIndex + 1):
        worksheet.cell(row=goodRow, column=colIndex).value = None
    worksheet[f'O{goodRow}'].hyperlink = None
    workbook.save(excelFilePath)
    print(f"Record {goodID} data from row {goodID + 1} cleared successfully.")

       
if __name__ == "__main__":
    cemetery = "Fairview"
    goodID = 6634   
    badID = 6643
    excelFilePath = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"
    workbook = openpyxl.load_workbook(excelFilePath)
    worksheet = workbook[cemetery]
    for row in range(1, worksheet.max_row + 1):
        if worksheet[f'A{row}'].value == goodID:
            goodRow = row
        if worksheet[f'A{row}'].value == badID:
            badRow = row
    adjustImageName(cemetery, goodID, badID, goodRow)
    cleanDelete(cemetery, badID, badRow)
    microsoftOCR.main(True, cemetery, "K") # Change this to match duplicate last name letter
    cleanHyperlink(badRow, cemetery)
    cleanRedacted(badID)
    cleanImages(goodID)
    duplicates.main()
