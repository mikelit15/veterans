import os
import re
from openpyxl.worksheet.hyperlink import Hyperlink
import openpyxl

def cleanDelete(badWorkSheet, badID, badRow):
    # baseRedacPath = r"\\ucclerk\pgmdoc\Veterans\Cemetery - Redacted"
    # for dirpath, dirnames, filenames in os.walk(baseRedacPath):
    #     for filename in filenames:
    #         if f"{badID:05d}" in filename:
    #             filePath = os.path.join(dirpath, filename)
    #             os.remove(filePath)
    #             print(f"\nRedacted file, {filename}, deleted successfully.")
    badWorkSheet.delete_rows(badRow)
    print(f"Row {badRow} deleted successfully.")
    for row in range(badRow, badWorkSheet.max_row + 1):
        nextRow = row + 1
        cellRef = f'O{row}'
        nextCellRef = f'O{nextRow}'
        if nextRow <= badWorkSheet.max_row and badWorkSheet[nextCellRef].hyperlink:
            tempHyperlink = badWorkSheet[nextCellRef].hyperlink
            badWorkSheet[cellRef].hyperlink = Hyperlink(ref=cellRef, target=tempHyperlink.target, display=tempHyperlink.display)
            badWorkSheet[cellRef].value = badWorkSheet[nextCellRef].value
            print(f"Hyperlink and value from row {nextRow} moved to row {row}.")
    print(badWorkSheet.max_row - 1)
    if badWorkSheet[f'O{badWorkSheet.max_row - 1}'].hyperlink:
        lastRow = badWorkSheet.max_row
        print(f"Adjusting the hyperlink in the last row, row {lastRow}.")
        secondLastHyperlink = badWorkSheet[f'O{lastRow - 1}'].hyperlink
        newTarget = re.sub(r'(\d+)(?=%\d+)', lambda m: f"{int(m.group(1)) + 1:05d}", secondLastHyperlink.target)
        badWorkSheet[f'O{lastRow}'].hyperlink = Hyperlink(ref=f'O{lastRow}', target=newTarget, display=secondLastHyperlink.display)
        print(f"Updated hyperlink in row {lastRow} to new target, {newTarget}.\n")

excelFilePath = r"\\ucclerk\pgmdoc\Veterans\Veterans2.xlsx"
workbook = openpyxl.load_workbook(excelFilePath)
sheet = workbook["Graceland"]
cleanDelete(sheet, 10370, 2130)
workbook.save(excelFilePath)