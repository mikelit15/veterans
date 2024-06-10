import openpyxl
import os
import shutil
import re
import sys
from openpyxl.worksheet.hyperlink import Hyperlink
import duplicates
sys.path.append(r'C:\workspace\veterans')
from microsoftOCR import microsoftOCR


'''
Initiates the file name cleaning process. This function goes through each cemetery folder
and then calls processCemetery to handle the letter subfolders. The function starts calling 
processCemetery when the ID matches "start." This is used to significantly speed up
the process of this function, especially for further cemeteries. Special handling is 
required if a cemetery folder contains more cemetery subfolders. This is the case for 
Jewish and Misc folders as they contain more, much smaller, cemeteries.

@param goodID (int) - A file ID used to determine when to start processing
@param redac (str) - The redacted folder name extension
@param redac2 (str) - The redacted file name extension

- Processes all cemeteries in the veternas cemetery folder, or the cemetery redacted
  folder.
- Starts processing file names at goodID - 600
@author Mike
'''
def cleanImages(goodID, redac, redac2):
    baseCemPath = fr"\\ucclerk\pgmdoc\Veterans\Cemetery{redac}"
    cemeterys = [d for d in os.listdir(baseCemPath) if os.path.isdir(os.path.join(baseCemPath, d))]
    initialCount = 1
    start = goodID - 600
    startFlag = True
    for cemetery in cemeterys:
        if cemetery in ["Jewish", "Misc"]:
            subCemPath = os.path.join(baseCemPath, cemetery)
            subCemeteries = [d for d in os.listdir(subCemPath) if os.path.isdir(os.path.join(subCemPath, d))]
            for subCemetery in subCemeteries:
                print(f"Processing {cemetery} - {subCemetery}")
                initialCount, startFlag = processCemetery(subCemetery, subCemPath, initialCount, start, startFlag, redac2)
        else:
            print(f"Processing {cemetery}")
            initialCount, startFlag = processCemetery(cemetery, baseCemPath, initialCount, start, startFlag, redac2)
      

'''
Iterates through alphabetically-based letter subfolders within a specified cemetery 
directory, calling clean() to adjust image names for each letter subfolder.

@param cemeteryName (str) - The name of the cemetery being processed.
@param basePath (str) - The root directory path where cemetery folders are located.
@param initialCount (int) - The starting point for file numbering within each alphabetical 
                            folder.
@param start (int) - A specific starting counter value used when deciding whether to process 
                     a file.
@param startFlag (bool) - A flag indicating whether the process should start renaming from 
                          the 'start' parameter.
@param redac2 (str) - The redacted file name extension

@return initialCount (int) - The starting point for file numbering within each alphabetical 
                             folder.
@return startFlag (bool) - A flag indicating whether the process should start renaming from 
                           the 'start' parameter.

- Iterates through A-Z folder, calling clean() for each letter
- Returns updated initialCount and startFlag

@author Mike
'''
def processCemetery(cemeteryName, basePath, initialCount, start, startFlag, redac2):
    uppercaseAlphabet = [chr(i) for i in range(ord('A'), ord('Z') + 1)]
    isFirstFile = True
    for letter in uppercaseAlphabet:
        namePath = os.path.join(basePath, cemeteryName, letter)
        try:
            print(f"Processing {cemeteryName} - {letter}")
            initialCount, startFlag = clean(letter, namePath, os.path.join(basePath, cemeteryName), \
                initialCount, isFirstFile, start, startFlag, redac2)
            isFirstFile = False
        except FileNotFoundError:
            continue
    return initialCount, startFlag

'''
Iterates through every PDF file in the given letter subfolder, renaming them to ensure 
sequential ordering and removes extra "a" and "b" redacted files. The function maintaines
continuity and handles special cases, such as the redacted files should not increment the 
count for the numbering sequence.

@param letter (str) - The current letter folder being processed.
@param namePath (str) - The path to the current folder where PDF files are located.
@param cemPath (str) - The path to the cemetery's root directory containing alphabetical 
                       folders.
@param initialCount (int) - The starting numbering for files in the current alphabetical 
                            sequence.
@param isFirstFile (bool) - Indicates if the current file is the first in its alphabetical 
                            sequence.
@param start (int) - The numbering value at which to begin processing files, if startFlag 
                     is True.
@param startFlag (bool) - A flag that, if True, delays file processing until the 'start' 
                          value is reached.
@param redac2 (str) - The redacted file name extension

@return counter (int) - The next starting count for continued processing.
@return startFlag (bool) - State of startFlag for continued processing.

- Creates a list for every PDF in the subfolder
- Finds the index of the last PDF file, which is in the previous letter subfolder
- Counter is set to that last index
- Adjusts the image file name to match the current index

@author Mike
'''
def clean(letter, namePath, cemPath, initialCount, isFirstFile, start, startFlag, redac2):
    pdfFiles = sorted(os.listdir(namePath))
    letters = sorted([folder for folder in os.listdir(cemPath) if os.path.isdir(os.path.join(cemPath, folder))])
    try:
        currentLetterIndex = letters.index(letter)
    except ValueError:
        return initialCount, startFlag
    folderBeforeIndex = max(0, currentLetterIndex - 1)
    folderBefore = letters[folderBeforeIndex]
    pdfFilesBefore = sorted([file for file in os.listdir(os.path.join(cemPath, folderBefore)) if file.lower().endswith('.pdf')])
    if letter == "A" and folderBefore == "A":
        cemeteryFolders = sorted(os.listdir(os.path.dirname(cemPath)))
        currentCemeteryIndex = cemeteryFolders.index(os.path.basename(os.path.normpath(cemPath)))
        if currentCemeteryIndex > 0:
            previousCemeteryIndex = currentCemeteryIndex - 1
            previousCemeteryFolder = cemeteryFolders[previousCemeteryIndex]
            previousCemeteryPath = os.path.join(os.path.dirname(cemPath), previousCemeteryFolder)
            previousCemeteryLetters = sorted([folder for folder in os.listdir(previousCemeteryPath) if os.path.isdir(os.path.join(previousCemeteryPath, folder))])
            folderBefore = previousCemeteryLetters[-1]  
            pdfFilesBefore = sorted([file for file in os.listdir(os.path.join(previousCemeteryPath, folderBefore)) if file.lower().endswith('.pdf')])
        else:
            return initialCount, startFlag
    maxCounter = 0
    for file in pdfFilesBefore:
        match = re.search(r'\d+', file)
        if match:
            number = int(match.group())
            maxCounter = max(maxCounter, number)
    counter = maxCounter + 1
    for x in pdfFiles:
        if counter != start and startFlag:
            isFirstFile = False
            if "a" in x[-5:]:
                pass
            else:
                counter += 1
            continue
        startFlag = False
        if isFirstFile:
            counter = initialCount 
        if redac2:
            if "a redacted.pdf" in x:
                counter -= 1
                os.remove(os.path.join(namePath, x))
                print(f"Removed {x}")
            elif "b redacted.pdf" in x:
                counter -= 1
                os.remove(os.path.join(namePath, x))
                print(f"Removed {x}")
            else:
                newName = re.sub(r'\d+', f"{counter:05d}", x)
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
        else:
            if "a.pdf" in x:
                newName = re.sub(r'\d+', f"{counter:05d}", x)
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
                counter -= 1
            elif "b.pdf"  in x:
                newName = re.sub(r'\d+', f"{counter:05d}", x)
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
            else:   
                newName = re.sub(r'\d+', f"{counter:05d}", x)
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
        counter += 1
        isFirstFile = False
    return counter, startFlag


'''
Adjusts the filenames and paths for image files associated with goodID and badID. GoodID 
receives a "a.pdf" suffix and badID receives a "b.pdf" suffix. badID is also adjusted to 
match the same letter and folder as the goodID if needed. The data in goodRow is then cleared
and the hyperlink forcefully cleared to allow for microsoftOCR to locate and append the new
merged data.

@param goodID (int): The correct identifier associated with the proper record.
@param badID (int): The incorrect identifier that needs to be adjusted to match the good ID.
@param goodRow (int): The row in the Excel spreadsheet where the good ID is located.

- Searches for files matching goodID and badID
- Adds the suffix, "a.pdf", to goodID
- Adds the suffix, "b.pdf", to badID. Changes letter in file name and moves to correct
  letter subfolder is necessary.
- Clears the data in Excel row associated with the goodID

@author Mike
'''
def adjustImageName(goodID, badID, goodRow, goodWorkSheet):
    baseCemPath = r"\\ucclerk\pgmdoc\Veterans\Cemetery"
    goodIDFound = False
    badIDFound = False
    goodIDFilePath = ""
    badIDFilePath = ""
    for dirPath, dirNames, fileNames in os.walk(baseCemPath):
        for fileName in fileNames:
            if f"{goodID:05d}" in fileName:
                goodIDFilePath = os.path.join(dirPath, fileName)
                goodIDFound = True
            elif f"{badID:05d}" in fileName:
                badIDFilePath = os.path.join(dirPath, fileName)
                badIDFound = True
            if goodIDFound and badIDFound:
                break
        if goodIDFound and badIDFound:
            break
    if goodIDFound and badIDFound:
        goodIDPrefix = os.path.basename(os.path.dirname(goodIDFilePath))
        badIDPrefix = os.path.basename(os.path.dirname(badIDFilePath))
        newBadIDFilename = os.path.basename(goodIDFilePath).replace(f"{goodID:05d}", f"{goodID:05d}b").replace(badIDPrefix, goodIDPrefix)
        newBadIDFilePath = os.path.join(os.path.dirname(goodIDFilePath), newBadIDFilename)
        if not "a.pdf" in os.path.basename(goodIDFilePath):
            newGoodIDFilename = os.path.basename(goodIDFilePath).replace(".pdf", "a.pdf")
            newGoodIDFilePath = os.path.join(os.path.dirname(goodIDFilePath), newGoodIDFilename)
            os.rename(goodIDFilePath, newGoodIDFilePath)
            print(f"{os.path.basename(goodIDFilePath)} renamed to {newGoodIDFilename}")
        shutil.move(badIDFilePath, newBadIDFilePath)
        print(f"{os.path.basename(badIDFilePath)} moved and renamed to {newBadIDFilename}")
    elif goodIDFound:
        print("BadID file not found. Please check the IDs and directories.")
        quit()
    elif badIDFound:
        print("GoodID file not found. Please check the IDs and directories.")
        quit()
    else:
        print("GoodID and BadID files not found. Please check the IDs and directories.")
        quit()
    startColumn = 'A'
    endColumn = 'O'
    startColumnIndex = openpyxl.utils.column_index_from_string(startColumn)
    endColumnIndex = openpyxl.utils.column_index_from_string(endColumn)
    for colIndex in range(startColumnIndex, endColumnIndex + 1):
        goodWorkSheet.cell(row=goodRow, column=colIndex).value = None
    goodWorkSheet[f'O{goodRow}'].hyperlink = None
    print(f"Record {goodID} data from row {goodRow} cleared successfully.")


'''
Deletes redacted image file associated with a badID and updates an Excel spreadsheet to 
remove the corresponding row. Due to a bug with the openpyxl module, hyperlinks do not 
move up a row with the rest of the cell data when calling .delete_rows(). So the hyperlinks 
are forced up one row and the hyperlink id and ref are adjusted to make the hyperlinks 
active in order to adjust them automatically later in the process. 

@param cemetery (str): The name of the cemetery where the file is located.
@param badID (int): The ID of the file to be deleted.
@param badRow (int): The row in the Excel spreadsheet corresponding to badID.

- Deletes the redacted file with badID ID number
- Deletes the Excel row corresponding to the bad ID
- Hyperlinks are adjusted starting from badRow

@author Mike
'''
def cleanDelete(badWorkSheet, badID, badRow):
    baseRedacPath = r"\\ucclerk\pgmdoc\Veterans\Cemetery - Redacted"
    for dirPath, dirNames, fileNames in os.walk(baseRedacPath):
        for fileName in fileNames:
            if f"{badID:05d}" in fileName:
                filePath = os.path.join(dirPath, fileName)
                os.remove(filePath)
                print(f"\nRedacted file, {fileName}, deleted successfully.")
    badWorkSheet.delete_rows(badRow)
    print(f"Row {badRow} deleted successfully.")
    for row in range(badRow, badWorkSheet.max_row + 1):
        nextRow = row + 1
        cellRef = f'O{row}'
        nextCellRef = f'O{nextRow}'
        if nextRow <= badWorkSheet.max_row and badWorkSheet[nextCellRef].hyperlink:
            tempHyperlink = badWorkSheet[nextCellRef].hyperlink
            badWorkSheet[cellRef].hyperlink = Hyperlink(ref=cellRef, target=tempHyperlink.target, display=tempHyperlink.display)
            badWorkSheet[cellRef].value = badWorkSheet[nextCellRef].value
            print(f"Hyperlink and value from row {nextRow} moved to row {row}.")
    if badWorkSheet[f'O{badWorkSheet.max_row - 1}'].hyperlink:
        lastRow = badWorkSheet.max_row
        print(f"Adjusting the hyperlink in the last row, row {lastRow}.")
        secondLastHyperlink = badWorkSheet[f'O{lastRow - 1}'].hyperlink
        newTarget = re.sub(r'(\d+)(?=%\d+)', lambda m: f"{int(m.group(1)) + 1:05d}", secondLastHyperlink.target)
        badWorkSheet[f'O{lastRow}'].hyperlink = Hyperlink(ref=f'O{lastRow}', target=newTarget, display=secondLastHyperlink.display)
        print(f"Updated hyperlink in row {lastRow} to new target, {newTarget}.\n")


'''
Updates the record ID numbers in column A to fix fill any gaps and continue going in 
sequential order. Updates hyperlink references in an Excel spreadsheet starting from a 
specified index. This function adjusts the numeric ID within each hyperlink to ensure 
they accurately reflect the current record IDs following modifications.

@param startIndex (int): The row index from which to start updating hyperlinks in the 
                         spreadsheet.

- Adjusts record ID in column A 
- Takes the hyperlink in the cell and changes the ID portion of the file name to match 
  the ID in column A for each row starting at startIndex

@author Mike
'''
def cleanHyperlinks(badWorksheet, startIndex, newID):
    basePath = fr'\\ucclerk\pgmdoc\Veterans\Cemetery'
    fileDirectoryMap = {}
    for dirPath, dirNames, fileNames in os.walk(basePath):
        for fileName in fileNames:
            fileID = ''.join(filter(str.isdigit, fileName))[:5]
            if fileID:
                fileDirectoryMap[fileID] = dirPath
    for row in range(startIndex, badWorksheet.max_row + 1):
        badWorksheet[f'A{row}'].value = newID
        cellRef = f'O{row}'
        if badWorksheet[cellRef].hyperlink:
            origTarget = badWorksheet[cellRef].hyperlink.target.replace("%20", " ")
            formattedID = f"{newID:05d}"
            folderName = fileDirectoryMap.get(formattedID)[-1]
            modifiedString = re.sub(r'([\\/])[A-Z]([\\/])', f'\\1{folderName}\\2', origTarget)
            modifiedString = re.sub(fr'({badWorksheet.title}[A-Z])\d{{5}}', f'{badWorksheet.title}{folderName}{formattedID}', modifiedString)
            badWorksheet[cellRef].hyperlink = Hyperlink(ref=cellRef, target=modifiedString, display="PDF Image")
            badWorksheet[cellRef].value = "PDF Image"
            print(f"Updated hyperlink from {origTarget} to {modifiedString} in row {row}.")
        newID += 1
    return newID


'''
Compares the letters in the folder path and file name of hyperlinks within an Excel 
workbook to ensure they match. The function identifies rows where the folder letter 
and file letter in the hyperlink do not match and prints these discrepancies.

@param workbookPath (str): The path to the Excel workbook to be processed.

- Iterates through each worksheet and each row starting from row 2.
- Extracts the hyperlink from column O (15th column) and compares the letters in the 
  folder path and file name.
- Prints hyperlinks where the letters do not match.

@author Mike
'''
def compareHyperlinkLetters(workbookPath):
    workbook = openpyxl.load_workbook(workbookPath)
    pattern = re.compile(
        r'Cemetery - Redacted[\\/][^\\/]+ - Redacted[\\/](?P<folderLetter>[A-Z])[\\/][^\\/]+(?P<folderLetter>[A-Z])\d+ redacted\.pdf'
    )
    for sheetName in workbook.sheetnames:
        worksheet = workbook[sheetName]
        for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column):
            cell = row[14]  
            if cell.hyperlink:
                hyperlink = cell.hyperlink.target.replace("%20", " ")
                match = pattern.search(hyperlink)
                if match:
                    folderLetter = match.group(1)
                    fileLetter = match.group(2)
                    if folderLetter == fileLetter:
                        pass
                    else:
                        print(f"Not matching: {hyperlink} in row {cell.row}")


'''
Main function to process and update Excel sheets containing good and bad IDs. It adjusts
image names, updates hyperlinks, cleans up records, and compares hyperlink letters to 
ensure consistency.

@param goodIDs (list): List of good IDs to be processed.
@param badIDs (list): List of bad IDs to be processed.

- Loads the Excel workbook and identifies sheets and rows containing good and bad IDs.
- Adjusts image names and updates records based on good and bad IDs.
- Cleans and deletes records with bad IDs.
- Processes images to update records and hyperlinks.
- Compares hyperlink letters to ensure consistency.

The function uses various helper functions such as `adjustImageName`, `microsoftOCR.main`,
`cleanDelete`, `cleanImages`, `cleanHyperlinks`, and `compareHyperlinkLetters`.

The Excel file is saved after each major operation to ensure data consistency.

@author Mike
'''        
def main(goodIDs, badIDs):
    excelFilePath = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"
    goodWorkSheet = None
    badWorksheet = None
    for x in range(0, len(goodIDs)):
        workbook = openpyxl.load_workbook(excelFilePath)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in range(2, worksheet.max_row + 1):
                if worksheet[f'A{row}'].value == goodIDs[x]:
                    goodRow = row
                    goodSheet = sheet
                if worksheet[f'A{row}'].value == badIDs[x]:
                    badRow = row
                    badSheet = sheet    
        goodWorkSheet = workbook[goodSheet] 
        if goodWorkSheet.title == "Extra":
            letter = "A"
        else:
            letter = goodWorkSheet[f'B{goodRow}'].value[0]
        adjustImageName(goodIDs[x], badIDs[x], goodRow, goodWorkSheet)
        workbook.save(excelFilePath)
        microsoftOCR.main(True, goodWorkSheet.title, letter)
        workbook = openpyxl.load_workbook(excelFilePath)
        badWorksheet = workbook[badSheet]
        cleanDelete(badWorksheet, badIDs[x], badRow)
        workbook.save(excelFilePath)
    miniG = min(goodIDs)
    if min(badIDs) < miniG:
        miniG = min(badIDs) - 1
    miniB = min(badIDs)
    cleanImages(miniG, "", "")
    cleanImages(miniG, " - Redacted", " redacted")
    workbook = openpyxl.load_workbook(excelFilePath)
    for x in range(0, len(goodIDs)):
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in range(2, worksheet.max_row + 1):
                if worksheet[f'A{row}'].value == miniB-1:
                    startIndex = row
                    startSheet = sheet 
    badWorksheet = workbook[startSheet]
    newID = badWorksheet[f'A{startIndex}'].value
    newID = cleanHyperlinks(badWorksheet, startIndex, newID)
    workbook.save(excelFilePath)
    breakFlag = False
    workbook = openpyxl.load_workbook(excelFilePath)
    for x in range(0, len(workbook.sheetnames)):
        if breakFlag:
            badWorksheet = workbook[workbook.sheetnames[x]]
            newID = cleanHyperlinks(badWorksheet, 2, newID)
        if workbook.sheetnames[x] == badWorksheet.title:
            breakFlag = True
    workbook.save(excelFilePath)
    compareHyperlinkLetters(excelFilePath)
    duplicates.main()

if __name__ == "__main__": 
    goodIDs = [10677] 
    badIDs = [10678]
    main(goodIDs, badIDs)