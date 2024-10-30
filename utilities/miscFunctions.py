import openpyxl
import os
import pandas as pd



### Changes slashes in hyperlinks

# workbook_path = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"
# wb = openpyxl.load_workbook(workbook_path)

# # Function to fix the hyperlink slashes
# def fix_hyperlink_slashes(sheet):
#     for row in sheet.iter_rows(min_col=15, max_col=15, min_row=1, values_only=False):
#         for cell in row:
#             if cell.hyperlink:  # Check if the cell has a hyperlink
#                 corrected_link = cell.hyperlink.target.replace("\\", "/")  # Replace backslashes with forward slashes
#                 cell.hyperlink.target = corrected_link  # Assign the corrected hyperlink back

# # Iterate through each sheet in the workbook
# for sheet in wb.worksheets:
#     fix_hyperlink_slashes(sheet)

# # Save the changes
# wb.save(r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx")  # Save the workbook with a new name



### Changes prefix in hyperlink

# workbook_path = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"
# wb = openpyxl.load_workbook(workbook_path)

# old_prefix = r"Cemetery - Redacted"
# new_prefix = r"Cemetery - Redacted"
# wb.save(workbook_path)
# for sheet in wb.worksheets:
#     for row in range(2, sheet.max_row + 1):
#         cell = sheet[f'O{row}']
#         if cell.hyperlink:
#             old_link = cell.hyperlink.target
#             if old_link.startswith(old_prefix):
#                 new_link = old_link.replace(old_prefix, new_prefix, 1)
#                 cell.hyperlink.target = new_link
#                 cell.value = new_link
                
# wb.save(r"\\ucclerk\pgmdoc\Veterans\VeteransWeb.xlsx")



### Shows full hyperlink value instead of cover text

workbook_path = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"
wb = openpyxl.load_workbook(workbook_path)
for sheet in wb.worksheets:
    for row in range(2, sheet.max_row + 1):
        cell = sheet[f'O{row}']
        if cell.hyperlink:
            cell.value = cell.hyperlink.target
wb.save(r"\\ucclerk\pgmdoc\Veterans\VeteransWeb.xlsx")


### Finds all cemeteries that do not have a folder

# workbook_path = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"  
# wb = openpyxl.load_workbook(workbook_path)
# cemetery_folder = r"\\ucclerk\pgmdoc\Veterans\Cemetery"
# cemetery_folder1 = r"\\ucclerk\pgmdoc\Veterans\Cemetery\Jewish"
# cemetery_folder2 = r"\\ucclerk\pgmdoc\Veterans\Cemetery\Misc"
# cemetery_names = set()
# for foldername in os.listdir(cemetery_folder):
#     folder_path = os.path.join(cemetery_folder, foldername)
#     if os.path.isdir(folder_path):  # Ensure it's a folder
#         cemetery_names.add(foldername)
# for foldername in os.listdir(cemetery_folder1):
#     folder_path = os.path.join(cemetery_folder1, foldername)
#     if os.path.isdir(folder_path):  # Ensure it's a folder
#         cemetery_names.add(foldername)
# for foldername in os.listdir(cemetery_folder2):
#     folder_path = os.path.join(cemetery_folder2, foldername)
#     if os.path.isdir(folder_path):  # Ensure it's a folder
#         cemetery_names.add(foldername)
### Function to find values not in the cemetery name set
# def find_all_invalid_cemetery_names(workbook):
#     invalid_names = set()  # Use a set to store unique invalid names
#     for sheet in workbook.worksheets:
#         for row in sheet.iter_rows(min_col=14, max_col=14, min_row=2, values_only=True):  # Column N is the 14th column
#             for cell_value in row:
#                 if cell_value and cell_value not in cemetery_names: 
#                     invalid_names.add(cell_value)  
#     return invalid_names
# invalid_cemeteries = find_all_invalid_cemetery_names(wb)
# if invalid_cemeteries:
#     print("The following cemetery names are not in the cemetery names set:")
#     for cemetery in sorted(invalid_cemeteries):
#         print(cemetery)
# else:
#     print("All values in column N match the cemetery names set.")


### Finds all unique values in a column

# file_path = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"
# excel_data = pd.read_excel(file_path, sheet_name=None)  
# unique_values_set = set()
# for sheet_name, df in excel_data.items():
#     if df.shape[1] >= 11:  
#         values = df.iloc[:, 10].dropna().unique()
#         unique_values_set.update(values)

# all_unique_values = list(unique_values_set)
# print(f"Unique values in column 'K' across all sheets:")
# for value in unique_values_set:
#     print(value)
