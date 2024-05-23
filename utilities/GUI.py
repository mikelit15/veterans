from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, \
     QHBoxLayout, QPushButton, QLineEdit, QLabel, QGroupBox, QFormLayout, \
     QTextEdit, QScrollArea, QComboBox, QMessageBox, QPlainTextEdit
from PyQt6.QtGui import QPixmap, QFont, QIcon
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from openpyxl.styles import Font, PatternFill
import traceback
import os
import openpyxl
import re
from openpyxl.worksheet.hyperlink import Hyperlink
import shutil
import sys
import pandas as pd
import duplicates
sys.path.append(r'C:\workspace\veterans')
from microsoftOCR import microsoftOCR
import qdarktheme
from tabulate import tabulate


'''
Dark Mode QMessageBox Button Styling
'''
DarkB = ("""
    QPushButton {
        background-color: #0078D4; 
        color: #E4E7EB;           
        border: 1px solid #8A8A8A; 

    }
    QPushButton:hover {
        background-color: #669df2; 
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                    stop:0 #80CFFF, stop:1 #004080);
    }
    QPushButton:pressed {
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                    stop:0 #004080, stop:1 #001B3D);
    }
    QPushButton:disabled {
        background-color: #1A1A1C; 
        border: 1px solid #3B3B3B;
        color: #3B3B3B;   
    }
    QTextEdit {
        background-color: #3F4042; 
        color: #FFFFFF;           
        border: 1px solid #8A8A8A; 
    }
    QTextEdit:disabled {
        background-color: #1A1A1C; 
        border: 1px solid #3B3B3B; 
    }
    QPlainTextEdit {
        background-color: #3F4042; 
        color: #FFFFFF;  
        border: none;         
    }
    QPlainTextEdit:disabled {
        background-color: #1A1A1C; 
    }
    QScrollArea {
        background-color: #3F4042;  
        color: #FFFFFF;            
        border: 1px solid #8A8A8A; 
    }
    QScrollArea:disabled {
        background-color: #1A1A1C;  
        border: 1px solid #3B3B3B; 
    }
""")

'''
Light Mode QMessageBox Button Styling
'''
LightB = ("""
    QPushButton {
        background-color: #70c5ff;  
        color: #111111;            
        border: 1px solid #111111; 
    }
    QPushButton:hover {
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 #70C5FF, stop:1 #0078D4);
    }
    QPushButton:pressed {
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                    stop:0 #0078D4, stop:1 #004080);
    }
    QPushButton:disabled {
        background-color: #DBDBDB; 
        color: #686868; 
        border: 1px solid #686868;
    }
    QTextEdit {
        background-color: #FFFFFF; 
        color: #111111; 
        border: 1px solid #111111; 
    }
    QTextEdit:disabled {
        background-color: #DBDBDB; 
        color: #686868; 
        border: 1px solid #686868;
    }
    QPlainTextEdit {
        background-color: #FFFFFF; 
        color: #111111; 
        border: none;
    }
    QPlainTextEdit:disabled {
        background-color: #DBDBDB; 
        color: #686868; 
    }
    QScrollArea {
        background-color: #FFFFFF;  
        color: #111111;            
        border: 1px solid #111111; 
    }
    QScrollArea:disabled {
        background-color: #DBDBDB;  
        color: #686868;            
        border: 1px solid #686868;
    }
""")

'''
Dark Mode Styling
'''
Dark = qdarktheme.load_stylesheet(
            theme="dark",
            custom_colors=
            {
                "[dark]": 
                {
                    "primary": "#0078D4",
                    "background": "#202124",
                    "border": "#8A8A8A",
                    "background>popup": "#252626",
                }
            },
        ) + """
            QMessageBox QLabel {
                color: #E4E7EB;
            }
            QDialog {
                background-color: #252626;
            }
        """

'''
Light Mode Styling
'''
Light = qdarktheme.load_stylesheet(
            theme="light",
            custom_colors=
            {
                "[light]": 
                {
                    "primary": "#0078D4",
                    "input.background": "#FFFFFF",
                    "background>textarea": "#FFFFFF",
                    "foreground": "#111111",
                    "background": "#F0F0F0",
                    "border": "#111111",
                    "background>popup": "#C9C9C9",
                }
            },
        ) + """
            QMessageBox QLabel {
                color: #111111;
            }
            QDialog {
                background-color: #C9C9C9;
            }
        """

class Worker(QThread):
    updateLabelSignal = pyqtSignal(str)  
    checkButtonSignal = pyqtSignal(bool)
    cleanButtonSignal = pyqtSignal(bool)
    statusSignal = pyqtSignal(str)
    '''
    The main processing logic that runs when the "Run" button is pressed and all the required
    fields are filled out. 

    @param self - the main window
    @param cemetery (str): The name of the cemetery being processed
    @param goodIDs (list): List of good ID numbers
    @param badIDs (list): List of bad ID numbers to be corrected

    @author Mike
    '''
    def __init__(self, cemetery, goodIDs, badIDs):
        super().__init__()
        self.cemetery = cemetery
        self.goodIDs = goodIDs
        self.badIDs = badIDs
        self.paused = False

    def runOCR(self, singleFlag, singleCem, singleLetter):
        global cemSet, miscSet, jewishSet
        networkFolder = r"\\ucclerk\pgmdoc\Veterans"
        os.chdir(networkFolder)
        cemeterys = []
        for x in os.listdir(r"Cemetery"):
            cemeterys.append(x)
        miscs = []
        for x in os.listdir(r"Cemetery\Misc"):
            miscs.append(x)
        jewishs = []
        for x in os.listdir(r"Cemetery\Jewish"):
            jewishs.append(x)
        cemSet, miscSet, jewishSet = set(cemeterys), set(miscs), set(jewishs)
        workbook = openpyxl.load_workbook('Veterans.xlsx')
        cemetery = singleCem 
        cemPath = os.path.join(networkFolder, fr"Cemetery\{cemetery}")
        letter = singleLetter 
        namePath = letter 
        namePath = os.path.join(cemPath, namePath)
        pathA = ""
        rowIndex = 2
        worksheet = workbook[cemetery]
        warFlag = False
        pdfFiles = sorted(os.listdir(namePath))
        breakFlag = False
        for y in range(len(pdfFiles)):
            try:
                warFlag = False
                filePath = os.path.join(namePath, pdfFiles[y])
                rowIndex = microsoftOCR.find_next_empty_row(worksheet)
                try:
                    id = worksheet[f'{"A"}{rowIndex-1}'].value + 1
                except Exception:
                    id = int(pdfFiles[0][:-4].split(letter)[-1].lstrip('0'))
                string = pdfFiles[y][:-4].split(letter)[-1].lstrip('0')
                if "a" not in string and "b" not in string:
                    if id != int(string.replace("a", "").replace("b", "")):
                        continue
                    self.updateLabelSignal.emit(f"\nImage loaded: {id}")
                    vals, warFlag, nameCoords, serialCoords, printedKVS, kinLast = microsoftOCR.createRecord(filePath, id, cemetery, "")
                    self.updateLabelSignal.emit(printedKVS)
                    redactedFile = microsoftOCR.redact(filePath, cemetery, letter, nameCoords, serialCoords)
                    worksheet.cell(row=rowIndex, column=15).value = "PDF Image"
                    worksheet.cell(row=rowIndex, column=15).font = Font(underline="single", color="0563C1")
                    worksheet.cell(row=rowIndex, column=15).hyperlink = redactedFile
                    counter = 1
                    worksheet.cell(row=rowIndex, column=counter, value=id)
                    counter += 1
                    for x in vals:
                        worksheet.cell(row=rowIndex, column=counter, value=x)
                        counter += 1
                    microsoftOCR.highlightSingle(worksheet, cemetery, letter, warFlag, rowIndex, kinLast)
                    id += 1
                    rowIndex += 1
                else:
                    if id != int(string.replace("a", "").replace("b", "")):
                        continue
                    if "a" in string:
                        if (filePath.replace("a.pdf", "") in pdfFiles):
                            continue
                        pathA = filePath
                        self.updateLabelSignal.emit(f"\nImage loaded: {id} A")
                        vals1, warFlag, nameCoords, serialCoords, printedKVS, kinLast = microsoftOCR.createRecord(filePath, id, cemetery, "A")
                        self.updateLabelSignal.emit(printedKVS)
                        redactedFile = microsoftOCR.redact(filePath, cemetery, letter, nameCoords, serialCoords)
                    if "b" in string:
                        if (filePath.replace("b.pdf", "") in pdfFiles):
                            continue
                        self.updateLabelSignal.emit(f"Image loaded: {id} B")
                        vals2, warFlagB, nameCoords, serialCoords, printedKVS, kinLast = microsoftOCR.createRecord(filePath, id, cemetery, "B")
                        self.updateLabelSignal.emit(printedKVS)
                        if not warFlag or not warFlagB:
                            warFlag = False
                        else:
                            warFlag = True
                        redactedFile = microsoftOCR.redact(filePath, cemetery, letter, nameCoords, serialCoords)
                        microsoftOCR.mergeRecords(worksheet, vals1, vals2, rowIndex, id, warFlag, cemetery, letter)
                        microsoftOCR.mergeImages(pathA, filePath, cemetery, letter)
                        link_text = "PDF Image"
                        worksheet.cell(row=rowIndex, column=15).value = link_text
                        worksheet.cell(row=rowIndex, column=15).font = Font(underline="single", color="0563C1")
                        worksheet.cell(row=rowIndex, column=15).hyperlink = redactedFile.replace("b redacted.pdf", " redacted.pdf")
                        id += 1
                        rowIndex += 1
                        if singleFlag:
                            breakFlag = True
            except Exception as e:
                errorTraceback = traceback.format_exc()
                print(f"An error occurred: {e}")
                self.updateLabelSignal.emit(f"An error occurred: {e}")
                errorMessage = f"SKIPPED DUE TO ERROR : {errorTraceback}"
                worksheet.cell(row=rowIndex, column=1, value=id)
                worksheet.cell(row=rowIndex, column=2, value=errorMessage)
                highlightColor = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                for colIndex in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=rowIndex, column=colIndex)
                    cell.fill = highlightColor
                error_file_path = fr'Errors/{cemetery}{letter}{str(id).zfill(5)} Error.txt' 
                with open(error_file_path, 'a') as error_file:
                    error_file.write(f'{printedKVS} \n\n {errorTraceback}')
                id += 1
                rowIndex += 1
            extension1 = str(id-1).zfill(5)
            extension2 = ""
            if "a.pdf" in filePath:
                extension1 = str(id).zfill(5)
                extension2 = "a"
            elif "b.pdf" in filePath:
                extension1 = str(id-1).zfill(5)
                extension2 = "b"
            logFilePath = fr'Logs/{cemetery}{letter}{extension1}{extension2} Extracted.txt' 
            with open(logFilePath, 'w', encoding='utf-8') as logFile:
                logFile.write(printedKVS)
                logFile.write("\n")
            workbook.save('Veterans.xlsx')
            if breakFlag:
                break
    
    def cleanImages(self, goodID, redac, redac2):
            baseCemPath = fr"\\ucclerk\pgmdoc\Veterans\Cemetery{redac}"
            cemeterys = [d for d in os.listdir(baseCemPath) if os.path.isdir(os.path.join(baseCemPath, d))]
            initialCount = 1
            start = goodID - 600
            startFlag = True
            for cemetery in cemeterys:
                if cemetery in ["Jewish", "Misc"]:
                    subCemPath = os.path.join(baseCemPath, cemetery)
                    subCemeteries = [d for d in os.listdir(subCemPath) if os.path.isdir(os.path.join(subCemPath, d))]
                    for subCemetery in subCemeteries:
                        self.updateLabelSignal.emit(f"Processing {cemetery} - {subCemetery}")
                        print(f"Processing {cemetery} - {subCemetery}")
                        initialCount, startFlag = self.process_cemetery(subCemetery, subCemPath, initialCount, start, startFlag, redac2)
                else:
                    self.updateLabelSignal.emit(f"Processing {cemetery}")
                    print(f"Processing {cemetery}")
                    initialCount, startFlag = self.process_cemetery(cemetery, baseCemPath, initialCount, start, startFlag, redac2)
            
    def process_cemetery(self, cemeteryName, basePath, initialCount, start, startFlag, redac2):
        uppercaseAlphabet = [chr(i) for i in range(ord('A'), ord('Z') + 1)]
        isFirstFile = True
        for letter in uppercaseAlphabet:
            namePath = os.path.join(basePath, cemeteryName, letter)
            try:
                self.updateLabelSignal.emit(f"Processing {cemeteryName} - {letter}")
                print(f"Processing {cemeteryName} - {letter}")
                initialCount, startFlag = self.clean(letter, namePath, os.path.join(basePath, cemeteryName), \
                    initialCount, isFirstFile, start, startFlag, redac2)
                isFirstFile = False
            except FileNotFoundError:
                continue
        return initialCount, startFlag

    def clean(self, letter, namePath, cemPath, initialCount, isFirstFile, start, startFlag, redac2):
        pdfFiles = sorted(os.listdir(namePath))
        letters = sorted([folder for folder in os.listdir(cemPath) if os.path.isdir(os.path.join(cemPath, folder))])
        try:
            currentLetterIndex = letters.index(letter)
        except ValueError:
            return
        folderBeforeIndex = max(0, currentLetterIndex - 1)
        folderBefore = letters[folderBeforeIndex]
        pdfFilesBefore = sorted([file for file in os.listdir(os.path.join(cemPath, folderBefore)) if file.lower().endswith('.pdf')])
        maxCounter = 0
        for file in pdfFilesBefore:
            match = re.search(r'\d+', file)
            if match:
                number = int(match.group())
                maxCounter = max(maxCounter, number)
        counter = maxCounter + 1
        for x in pdfFiles:
            if counter != start and startFlag:
                isFirstFile = False
                if "a" in x[-5:]:
                    pass
                else:
                    counter += 1
                continue
            startFlag = False
            if isFirstFile:
                counter = initialCount 
            if redac2:
                if "a redacted.pdf" in x:
                    counter -= 1
                    os.remove(os.path.join(namePath, x))
                    self.updateLabelSignal.emit(f"Removed {x}")
                    print(f"Removed {x}")
                elif "b redacted.pdf" in x:
                    counter -= 1
                    os.remove(os.path.join(namePath, x))
                    self.updateLabelSignal.emit(f"Removed {x}")
                    print(f"Removed {x}")
                else:
                    newName = re.sub(r'\d+', f"{counter:05d}", x)
                    os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
            else:
                if "a.pdf" in x:
                    newName = re.sub(r'\d+', f"{counter:05d}", x)
                    os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
                    counter -= 1
                elif "b.pdf"  in x:
                    newName = re.sub(r'\d+', f"{counter:05d}", x)
                    os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
                else:   
                    newName = re.sub(r'\d+', f"{counter:05d}", x)
                    os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
            counter += 1
            isFirstFile = False
        return counter, startFlag

    def adjustImageName(self, goodID, badID, goodRow):
        global excelFilePath
        baseCemPath = r"\\ucclerk\pgmdoc\Veterans\Cemetery"
        goodIDFound = False
        badIDFound = False
        goodIDFilePath = ""
        badIDFilePath = ""
        for dirPath, dirNames, fileNames in os.walk(baseCemPath):
            for fileName in fileNames:
                if f"{goodID:05d}" in fileName:
                    goodIDFilePath = os.path.join(dirPath, fileName)
                    goodIDFound = True
                elif f"{badID:05d}" in fileName:
                    badIDFilePath = os.path.join(dirPath, fileName)
                    badIDFound = True
                if goodIDFound and badIDFound:
                    break
            if goodIDFound and badIDFound:
                break
        if goodIDFound and badIDFound:
            goodIDPrefix = os.path.basename(os.path.dirname(goodIDFilePath))
            badIDPrefix = os.path.basename(os.path.dirname(badIDFilePath))
            newBadIDFilename = os.path.basename(badIDFilePath).replace(f"{badID:05d}", f"{goodID:05d}b").replace(badIDPrefix, goodIDPrefix)
            newBadIDFilePath = os.path.join(os.path.dirname(goodIDFilePath), newBadIDFilename)
            if not "a.pdf" in os.path.basename(goodIDFilePath):
                newGoodIDFilename = os.path.basename(goodIDFilePath).replace(".pdf", "a.pdf")
                newGoodIDFilePath = os.path.join(os.path.dirname(goodIDFilePath), newGoodIDFilename)
                os.rename(goodIDFilePath, newGoodIDFilePath)
                self.updateLabelSignal.emit(f"{os.path.basename(goodIDFilePath)} renamed to {newGoodIDFilename}")
                print(f"{os.path.basename(goodIDFilePath)} renamed to {newGoodIDFilename}")
            shutil.move(badIDFilePath, newBadIDFilePath)
            self.updateLabelSignal.emit(f"{os.path.basename(badIDFilePath)} moved and renamed to {newBadIDFilename}")
            print(f"{os.path.basename(badIDFilePath)} moved and renamed to {newBadIDFilename}")
        else:
            self.updateLabelSignal.emit("GoodID or BadID file not found. Please check the IDs and directories.")
            print("GoodID or BadID file not found. Please check the IDs and directories.")
            workbook.save(excelFilePath)
            self.checkButtonSignal.emit(False)
            self.cleanButtonSignal.emit(False)
            quit()
        startColumn = 'A'
        endColumn = 'O'
        startColumnIndex = openpyxl.utils.column_index_from_string(startColumn)
        endColumnIndex = openpyxl.utils.column_index_from_string(endColumn)
        for colIndex in range(startColumnIndex, endColumnIndex + 1):
            worksheet.cell(row= goodRow, column= colIndex).value = None
        worksheet[f'O{goodRow}'].hyperlink = None
        self.updateLabelSignal.emit(f"Record {goodID} data from row {goodRow} cleared successfully.")
        print(f"Record {goodID} data from row {goodRow} cleared successfully.")

    def cleanDelete(self, cemetery, badID, badRow):
        baseRedacPath = r"\\ucclerk\pgmdoc\Veterans\Cemetery - Redacted"
        for dirPath, dirNames, fileNames in os.walk(baseRedacPath):
            for fileName in fileNames:
                if f"{badID:05d}" in fileName:
                    filePath = os.path.join(dirPath, fileName)
                    os.remove(filePath)
                    self.updateLabelSignal.emit(f"\nRedacted file, {fileName}, deleted successfully.")
                    print(f"Redacted file, {fileName}, deleted successfully.")
        worksheet = workbook[cemetery]
        worksheet.delete_rows(badRow)
        self.updateLabelSignal.emit(f"Row {badRow} deleted successfully.")
        print(f"Row {badRow} deleted successfully.\n")
        for row in range(badRow, worksheet.max_row + 1):
            nextRow = row + 1
            cellRef = f'O{row}'
            nextCellRef = f'O{nextRow}'
            if nextRow <= worksheet.max_row and worksheet[nextCellRef].hyperlink:
                tempHyperlink = worksheet[nextCellRef].hyperlink
                worksheet[cellRef].hyperlink = Hyperlink(ref=cellRef, target=tempHyperlink.target, display=tempHyperlink.display)
                worksheet[cellRef].value = worksheet[nextCellRef].value
                self.updateLabelSignal.emit(f"Hyperlink and value from row {nextRow} moved to row {row}.")
                print(f"Hyperlink and value from row {nextRow} moved to row {row}.")
        if worksheet[f'O{worksheet.max_row - 1}'].hyperlink:
            lastRow = worksheet.max_row
            self.updateLabelSignal.emit(f"\nAdjusting the hyperlink in the last row, row {lastRow}.")
            print(f"\nAdjusting the hyperlink in the last row, row {lastRow}.")
            secondLastHyperlink = worksheet[f'O{lastRow - 1}'].hyperlink
            newTarget = re.sub(r'(\d+)(?=%\d+)', lambda m: f"{int(m.group(1)) + 1:05d}", secondLastHyperlink.target)
            worksheet[f'O{lastRow}'].hyperlink = Hyperlink(ref=f'O{lastRow}', target=newTarget, display=secondLastHyperlink.display)
            self.updateLabelSignal.emit(f"Updated hyperlink in row {lastRow} to new target, {newTarget}.\n")
            print(f"Updated hyperlink in row {lastRow} to new target, {newTarget}.\n")

    def cleanHyperlinks(self, cemetery, startIndex):
        baseBath = fr'\\ucclerk\pgmdoc\Veterans\Cemetery\{cemetery}'
        fileDirectoryMap = {}
        for dirPath, dirNames, fileNames in os.walk(baseBath):
            for fileName in fileNames:
                fileID = ''.join(filter(str.isdigit, fileName))[:5]
                if fileID:
                    fileDirectoryMap[fileID] = dirPath
        newID = worksheet[f'A{startIndex}'].value
        for row in range(startIndex, worksheet.max_row + 1):
            worksheet[f'A{row}'].value = newID
            cellRef = f'O{row}'
            if worksheet[cellRef].hyperlink:
                origTarget = worksheet[cellRef].hyperlink.target.replace("%20", " ")
                formattedID = f"{newID:05d}"
                folderName = fileDirectoryMap.get(formattedID, "UnknownFolder")[-1]
                modifiedString = re.sub(fr'(\\)([A-Z])(\\)', f'\\1{folderName}\\3', origTarget)
                modifiedString = re.sub(fr'({cemetery}[A-Z])\d{{5}}', f'{cemetery}{folderName}{formattedID}', modifiedString)
                worksheet[cellRef].hyperlink = Hyperlink(ref=cellRef, target=modifiedString, display="PDF Image")
                worksheet[cellRef].value = "PDF Image"
                self.updateLabelSignal.emit(f"Updated hyperlink from {origTarget} to {modifiedString} in row {row}.")
                print(f"Updated hyperlink from {origTarget} to {modifiedString} in row {row}.")
            newID += 1

    '''
    The main processing logic that runs when the "Clean Duplicates" button is pressed and all the required
    fields are filled out. 

    @param self - the main window

    @author Mike
    '''
    def run(self):
        
        global excelFilePath
        excelFilePath = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"
        global workbook
        workbook = openpyxl.load_workbook(excelFilePath)
        global worksheet
        worksheet = workbook[self.cemetery]
        for x in range(0, len(self.goodIDs)):
            workbook = openpyxl.load_workbook(excelFilePath)
            worksheet = workbook[self.cemetery]
            for row in range(1, worksheet.max_row + 1):
                if worksheet[f'A{row}'].value == self.goodIDs[x]:
                    goodRow = row
                if worksheet[f'A{row}'].value == self.badIDs[x]:
                    badRow = row
            letter = worksheet[f'B{goodRow}'].value[0]
            self.adjustImageName(self.goodIDs[x], self.badIDs[x], goodRow)
            workbook.save(excelFilePath)
            self.runOCR(True, self.cemetery, letter)
            workbook = openpyxl.load_workbook(excelFilePath)
            worksheet = workbook[self.cemetery]
            self.cleanDelete(self.cemetery, self.badIDs[x], badRow)
            workbook.save(excelFilePath)
        mini = min(self.goodIDs)
        if min(self.badIDs) < mini:
            mini = min(self.badIDs) - 1
        self.cleanImages(mini, "", "")
        print("\n")
        self.updateLabelSignal.emit("\n")
        self.cleanImages(mini, " - Redacted", " redacted")
        workbook = openpyxl.load_workbook(excelFilePath)
        worksheet = workbook[self.cemetery]
        for row in range(1, worksheet.max_row + 1):
            if worksheet[f'A{row}'].value == mini:
                startIndex = row
        self.cleanHyperlinks(self.cemetery, startIndex)
        workbook.save(excelFilePath)
        print("\n")
        columnWidths = {
        "VID": 8,
        "VLNAME": 15,
        "VFNAME": 15,
        "VMNAME": 5,
        "VSUF": 5,
        "VDOB": 15,
        "VDOBY": 10,
        "VDOD": 15,
        "VDODY": 10
        }
        df = duplicates.main(self.cemetery)
        string = df.to_string(index= False, header= True, formatters= {key: f'{{:>{width}}}'.format for key, width in columnWidths.items()})
        self.updateLabelSignal.emit("\nCurrent Duplicates:")
        self.updateLabelSignal.emit(string)
        self.checkButtonSignal.emit(False)
        self.cleanButtonSignal.emit(False)  
        self.statusSignal.emit(" ")  
                
class MainWindow(QMainWindow):
    updateLabelSignal = pyqtSignal(str)
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Veteran Duplicate Cleaner")
        self.setGeometry(400, 150, 1050, 150)
        self.worker = None  
        self.mainLayout()
        self.updateLabelSignal.connect(self.updateScroll) 


    '''
    Saves display mode to display_mode.txt

    @param mode - the name of the display mode that is being used, and therefore saved

    @author Mike
    '''
    def saveDisplayMode(self, mode):
        parentPath = os.path.dirname(os.getcwd())
        with open(f"{parentPath}/_internal/veteranData/display_mode.txt", "w") as file:
            file.write(mode)


    '''
    Loads display mode name from display_mode.txt

    @return string - the name of the display mode that was saved

    @author Mike
    '''
    def loadDisplayMode(self):
        parentPath = os.path.dirname(os.getcwd())
        try:
            with open(f"{parentPath}/_internal/veteranData/display_mode.txt", "r") as file:
                return file.read().strip()
        except FileNotFoundError:
            return "Light"
        
        
    '''
    Updates the display mode anytime the selection is changed within the app through
    the display mode selection box. Saves the mode selection to a local .txt file. 
    Calls functions to manually updates the styling of the buttons and the text widgets.

    @param self - the main window that is getting the styling adjustment
    @param checkButton - the QPushButton widget that calls checkDuplicates()
    @param pauseButton - the QPushButton widget that pauses the code
    @param cleanButton - the QPushButton widget that calls the main worker
    @param goodText - the QTextEdit widget holds the list of Good IDs
    @param badText - the QTextEdit widget holds the list of Bad IDs
    @param details1 - the QScrollArea widget that displays the processing details
    @param displayMode - the name of the display mode selected

    @author Mike
    '''
    def changeDisplayStyle(self, checkButton, pauseButton, cleanButton, goodText, badText, details, displayModeBox, displayMode):
        if displayMode == "Dark":
            displayModeBox.setStyleSheet("""            
                QComboBox QAbstractItemView {
                border: 1px solid #8A8A8A;
                border-radius: 1px;
                selection-background-color: #669df2;
                }
                QListView::item {
                border-radius: 0px;
                border-bottom: 1px solid #8A8A8A;
                }
                QListView::item:last {
                    border-bottom: none;
                }
                """)
        else:
            displayModeBox.setStyleSheet("""            
                QComboBox QAbstractItemView {
                border: 1px solid #111111;
                selection-background-color: #cfcfd1;
                }
                QListView::item {
                border-radius: 0px;
                border-bottom: 1px solid #111111;
                }
                QListView::item:last {
                    border-bottom: none;
                }
                """)
        self.setStyleSheet(globals()[displayMode])
        checkButton.setStyleSheet(globals()[f"{displayMode}B"])
        pauseButton.setStyleSheet(globals()[f"{displayMode}B"])
        cleanButton.setStyleSheet(globals()[f"{displayMode}B"])
        goodText.setStyleSheet(globals()[f"{displayMode}B"])
        badText.setStyleSheet(globals()[f"{displayMode}B"])
        details.setStyleSheet(globals()[f"{displayMode}B"])
        self.saveDisplayMode(displayMode)    
    
    
    '''
    Creates the main window UI. Places all the widgets in their respective places
    and adjusts their size and other properties.

    @param self - the main window

    @author Mike
    '''    
    def mainLayout(self):
        centralWidget = QWidget(self)
        centralWidget.setObjectName("mainCentralWidget")
        self.setCentralWidget(centralWidget)
        layout1 = QVBoxLayout(centralWidget)
        
        # Create the top container
        topContainer = QWidget()
        topLayout = QHBoxLayout()
        logoLabel = QLabel() 
        parentPath = os.path.dirname(os.getcwd())
        logoImage = QPixmap(f"{parentPath}/_internal/veteranData/ucLogo.png") 
        logoLabel.setPixmap(logoImage)
        programName = QLabel("Union County Clerk's Office\n\n Veteran Duplicate Cleaner")
        font = QFont()
        font.setBold(True)
        font.setPointSize(26)
        programName.setFont(font)
        topLayout.addWidget(logoLabel)
        topLayout.addWidget(programName)
        topContainer.setLayout(topLayout)

        # Create a group box for the middle top left
        middleTopLeftGroupBox = QGroupBox("Parameters")
        middleTopLeftGroupBox.setFixedSize(375, 425)
        middleTopLeftLayout = QFormLayout()
        self.cemeteryBox = QLineEdit()
        self.cemeteryBox.setFixedWidth(140)
        self.goodIDBox = QTextEdit()
        self.goodIDBox.setFixedWidth(250)
        self.goodIDBox.setFixedHeight(75)
        self.goodIDBox.setLineWrapColumnOrWidth(15)
        self.badIDBox = QTextEdit()
        self.badIDBox.setFixedWidth(250)
        self.badIDBox.setFixedHeight(75)
        self.badIDBox.setLineWrapColumnOrWidth(15)
        font = QFont("Monterchi Sans Book", 8)
        noteLabel1 = QLabel("Note: This is the cemetery name based on its folder.")
        noteLabel2 = QLabel("Note: This is a list of all the original IDs.")
        noteLabel3 = QLabel("Note: This is a list of all the duplicate IDs.")
        noteLabel1.setFont(font)
        noteLabel2.setFont(font)
        noteLabel3.setFont(font)
        self.displayModeBox = QComboBox()
        self.displayModeBox.setFixedWidth(75)
        modes = ["Light", "Dark"]
        self.displayModeBox.addItems(modes)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftLayout.addRow("Display Mode: ", self.displayModeBox)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftLayout.addRow("Cemetery :   ", self.cemeteryBox)
        middleTopLeftLayout.addRow(None, noteLabel1)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftLayout.addRow("Good IDs :", self.goodIDBox)
        middleTopLeftLayout.addRow(None, noteLabel2)
        middleTopLeftLayout.addRow(" ", None)
        middleTopLeftLayout.addRow("Bad IDs :", self.badIDBox)
        middleTopLeftLayout.addRow(None, noteLabel3)
        middleTopLeftGroupBox.setLayout(middleTopLeftLayout)
        self.goodIDBox.setDisabled(True)
        self.badIDBox.setDisabled(True)
        
        # Create a group box for the middle top right
        self.middleTopRightGroupBox = QGroupBox("Results")
        self.middleTopRightGroupBox.setFixedSize(775, 425)
        self.middleTopRightLayout = QFormLayout()
        self.detailsLabel = QPlainTextEdit()
        self.detailsLabel.setReadOnly(True)
        self.scrollArea = QScrollArea()
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setWidget(self.detailsLabel)
        self.middleTopRightLayout.addRow(" ", None)
        self.middleTopRightLayout.addRow("Details :", self.scrollArea)
        self.middleTopRightLayout.addRow(" ", None)
        self.middleTopRightGroupBox.setLayout(self.middleTopRightLayout)

        # Create container for middle top left and top right group boxes
        middleTopContainer = QWidget()
        middleTopLayout = QHBoxLayout()
        middleTopLayout.addWidget(middleTopLeftGroupBox)
        middleTopLayout.addWidget(self.middleTopRightGroupBox)
        middleTopContainer.setLayout(middleTopLayout)
        
        # Create the bottom container with a button
        bottomContainer = QWidget()
        bottomLayout = QHBoxLayout()
        self.checkButton = QPushButton("Check Duplicates")
        self.checkButton.clicked.connect(self.updateDuplicates)
        self.checkButton.setFixedWidth(150)
        self.pauseButton = QPushButton("Pause Code")
        self.pauseButton.clicked.connect(self.togglePauseResume)
        self.pauseButton.setFixedWidth(150)
        self.cleanButton = QPushButton("Clean Duplicates")
        self.cleanButton.clicked.connect(self.startProcessing)
        self.cleanButton.setFixedWidth(150)
        self.pauseButton.setDisabled(True)
        self.cleanButton.setDisabled(True)
        self.scrollArea.setDisabled(True)
        bottomLayout.addWidget(self.checkButton)
        bottomLayout.addWidget(self.pauseButton)
        bottomLayout.addWidget(self.cleanButton)
        self.displayModeBox.currentTextChanged.connect(\
            lambda: self.changeDisplayStyle(self.checkButton, \
                self.pauseButton, self.cleanButton, self.goodIDBox, \
                self.badIDBox, self.scrollArea, self.displayModeBox, \
                self.displayModeBox.currentText()))
        self.displayModeBox.setCurrentIndex(modes.index(self.loadDisplayMode()))
        bottomContainer.setLayout(bottomLayout)
        
        # Create a main layout to arrange the containers
        mainLayout = QVBoxLayout()
        mainLayout.addWidget(topContainer)
        mainLayout.addWidget(middleTopContainer)
        mainLayout.addWidget(bottomContainer)
        self.status = QLabel("             Status :  Idle\n")
        self.status.setFixedWidth(150)
        mainLayout.addWidget(self.status, 0, alignment= Qt.AlignmentFlag.AlignCenter)
        layout1.addLayout(mainLayout)
        self.changeDisplayStyle(self.checkButton, self.pauseButton, self.cleanButton, \
            self.goodIDBox, self.badIDBox, self.scrollArea, self.displayModeBox, self.displayModeBox.currentText())
    
    
    '''
    Calls a QMessageBox from the main run function from the Worker class. Passes in data needed for popup message.
    Resets the window status. 

    @param self - the main window
    @param type - the icon needed for the popup
    @param text - the text needed for the popup

    @author Mike
    '''        
    def updateStatus(self, type, text):
        if type == "Critical":
            msgBox = QMessageBox(QMessageBox.Icon.Critical, 'Error', text, QMessageBox.StandardButton.Ok, window)
            msgBox.setStyleSheet(globals()[f"{self.loadDisplayMode()}B"])
            msgBox.exec()
        elif type == "Warning":
            msgBox = QMessageBox(QMessageBox.Icon.Warning, 'Warning', text, QMessageBox.StandardButton.Ok, window)
            msgBox.setStyleSheet(globals()[f"{self.loadDisplayMode()}B"])
            msgBox.exec()
        elif type == "Info":
            msgBox = QMessageBox(QMessageBox.Icon.Information, 'Information', text, QMessageBox.StandardButton.Ok, window)
            msgBox.setStyleSheet(globals()[f"{self.loadDisplayMode()}B"])
            msgBox.exec()
        self.checkButton.setDisabled(False)
        self.cleanButton.setDisabled(False)
        self.status.setText("              Status :  Idle")

    
    '''
    Function that changes the text of the Pause/Resume button and updates the 
    app status accordingly. 

    @param self - the main window

    @author Mike
    ''' 
    def togglePauseResume(self):
        if self.worker:
            if self.worker.paused:
                self.worker.paused = False
                self.pauseButton.setText("Pause")
                self.status.setText("          Status : Running...\n")
            else:
                self.worker.paused = True
                self.pauseButton.setText("Resume")
                self.status.setText("          Status : Paused\n" )
    
    
    '''
    Function that scrolls to the bottom of the entire scroll area. Allows for new lines
    that are added to always be at the bottom the scroll area display.

    @param self - the main window
    @param newText - the new line to be displayed at the bottom of the display

    @author Mike
    '''     
    def updateScroll(self, newText):
        self.detailsLabel.appendPlainText(newText)
        self.scrollArea.verticalScrollBar().setValue(self.scrollArea.verticalScrollBar().maximum())
    
    
    '''
    Function that clears the layout of ScrollArea1. This allows for a new layout to 
    be implemented that is deisgned to show the details of the cleaning process.

    @param self - the main window

    @author Mike
    ''' 
    def updateDuplicates(self):
        self.detailsLabel.clear()
        columnWidths = {
        "VID": 8,
        "VLNAME": 15,
        "VFNAME": 15,
        "VMNAME": 15,
        "VSUF": 5,
        "VDOB": 11,
        "VDOBY": 6,
        "VDOD": 11,
        "VDODY": 6,
        "VWARREC": 15,
        "VCLNWAR": 15
        }
        try:
            df = duplicates.main(self.cemeteryBox.text())
        except Exception:
            self.updateStatus("Warning", "Cemetery field not filled out or misspelled.")
            return

        # Fill NaN values with "NaN" for better visibility in the output
        df = df.fillna("NaN")
        
        # Function to format a single column value
        def format_value(value, width):
            return f"{str(value):<{width}}"

        # Create a formatted string for the DataFrame
        formatted_lines = []
        
        # Format the header
        header = " ".join([format_value(col, columnWidths[col]) for col in df.columns])
        formatted_lines.append(header)
        
        # Format the rows
        for index, row in df.iterrows():
            formatted_row = " ".join([format_value(row[col], columnWidths[col]) for col in df.columns])
            formatted_lines.append(formatted_row)
        
        string = "\n".join(formatted_lines)
        self.scrollArea.setDisabled(False)
        self.detailsLabel.appendPlainText("Current Duplicates:")
        self.detailsLabel.appendPlainText(f"{string}\n")
        self.scrollArea.verticalScrollBar().setValue(self.scrollArea.verticalScrollBar().maximum())
        self.cleanButton.setDisabled(False)
        self.goodIDBox.clear()
        self.goodIDBox.clear()
        try:
            self.goodIDBox.setDisabled(False)
            self.badIDBox.setDisabled(False)
        except Exception:
            pass
          
    def startProcessing(self):
        try:
            self.goodIDBox = [int(numericString) for numericString in self.goodIDBox.toPlainText().split(", ")]
        except Exception:
            self.updateStatus("Warning", "Good ID field is empty.")
            return
        try:
            self.badIDBox = [int(numericString) for numericString in self.badIDBox.toPlainText().split(", ")]
        except Exception:
            self.updateStatus("Warning", "Bad ID field is empty.")
            return
        self.worker = Worker(self.cemeteryBox.text(), self.goodIDBox, self.badIDBox)
        self.worker.updateLabelSignal.connect(lambda text: self.updateScroll(text))
        self.worker.checkButtonSignal.connect(lambda condition: self.checkButton.setDisabled(condition))
        self.worker.cleanButtonSignal.connect(lambda condition: self.cleanButton.setDisabled(condition))
        self.worker.statusSignal.connect(lambda status: self.status.setText("              Status :  Idle"))
        self.worker.start()
        self.status.setText("          Status:  Running...\n")
        self.cleanButton.setDisabled(True)
        self.checkButton.setDisabled(True)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    parentPath = os.path.dirname(os.getcwd())
    window.setWindowIcon(QIcon(f"{parentPath}/_internal/veteranData/veteranLogo.png"))
    window.show()
    sys.exit(app.exec())