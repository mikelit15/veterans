import openpyxl
import os
import shutil
import re
import time
from PyPDF2 import PdfMerger
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import PatternFill

'''
Initiates the file name cleaning process. This function goes through each cemetery folder
and then calls processCemetery to handle the letter subfolders. The function starts calling 
processCemetery when the ID matches "start." This is used to significantly speed up
the process of this function, especially for further cemeteries. Special handling is 
required if a cemetery folder contains more cemetery subfolders. This is the case for 
Jewish and Misc folders as they contain more, much smaller, cemeteries.

@param fullClean (bool) - A flag indicating whether the full cleaning process should be done
@param goodID (int) - A file ID used to determine when to start processing
@param redac (str) - The redacted folder name extension
@param redac2 (str) - The redacted file name extension

- Processes all cemeteries in the veterans' cemetery folder or the cemetery redacted
  folder.
- Starts processing file names at goodID - 600

@author Mike
'''
def cleanImages(fullClean, goodID, redac, redac2):
    baseCemPath = fr"\\ucclerk\pgmdoc\Veterans\Cemetery{redac}"
    cemeterys = [d for d in os.listdir(baseCemPath) if os.path.isdir(os.path.join(baseCemPath, d))]
    initialCount = 1
    start = goodID
    startFlag = not fullClean
    for cemetery in cemeterys:
        if cemetery in [f"Jewish{redac}", f"Misc{redac}"]:
            subCemPath = os.path.join(baseCemPath, cemetery)
            subCemeteries = [d for d in os.listdir(subCemPath) if os.path.isdir(os.path.join(subCemPath, d))]
            for subCemetery in subCemeteries:
                print(f"Processing {cemetery} - {subCemetery}")
                initialCount, startFlag = processCemetery(subCemetery, subCemPath, initialCount, start, startFlag, redac, redac2)
        else:
            print(f"Processing {cemetery}")
            initialCount, startFlag = processCemetery(cemetery, baseCemPath, initialCount, start, startFlag, redac, redac2)
      
'''
Iterates through alphabetically-based letter subfolders within a specified cemetery 
directory, calling clean() to adjust image names for each letter subfolder.

@param cemeteryName (str) - The name of the cemetery being processed.
@param basePath (str) - The root directory path where cemetery folders are located.
@param initialCount (int) - The starting point for file numbering within each alphabetical 
                            folder.
@param start (int) - A specific starting counter value used when deciding whether to process 
                     a file.
@param startFlag (bool) - A flag indicating whether the process should start renaming from 
                          the 'start' parameter.
@param redac (str) - The redacted folder name extension
@param redac2 (str) - The redacted file name extension

@return initialCount (int) - The starting point for file numbering within each alphabetical 
                             folder.
@return startFlag (bool) - A flag indicating whether the process should start renaming from 
                           the 'start' parameter.

- Iterates through A-Z folders, calling clean() for each letter
- Returns updated initialCount and startFlag

@author Mike
'''
def processCemetery(cemeteryName, basePath, initialCount, start, startFlag, redac, redac2):
    uppercaseAlphabet = [chr(i) for i in range(ord('A'), ord('Z') + 1)]
    isFirstFile = True
    for letter in uppercaseAlphabet:
        namePath = os.path.join(basePath, cemeteryName, letter)
        try:
            print(f"Processing {cemeteryName} - {letter}")
            initialCount, startFlag = clean(letter, namePath, os.path.join(basePath, cemeteryName), \
                initialCount, isFirstFile, start, startFlag, redac, redac2)
            isFirstFile = False
        except FileNotFoundError:
            continue
    return initialCount, startFlag

'''
Iterates through every PDF file in the given letter subfolder, renaming them to ensure 
sequential ordering and removes extra "a" and "b" redacted files. The function maintains
continuity and handles special cases, such as the redacted files should not increment the 
count for the numbering sequence.

@param letter (str) - The current letter folder being processed.
@param namePath (str) - The path to the current folder where PDF files are located.
@param cemPath (str) - The path to the cemetery's root directory containing alphabetical 
                       folders.
@param initialCount (int) - The starting numbering for files in the current alphabetical 
                            sequence.
@param isFirstFile (bool) - Indicates if the current file is the first in its alphabetical 
                            sequence.
@param start (int) - The numbering value at which to begin processing files, if startFlag 
                     is True.
@param startFlag (bool) - A flag that, if True, delays file processing until the 'start' 
                          value is reached.
@param redac (str) - The redacted folder name extension
@param redac2 (str) - The redacted file name extension

@return counter (int) - The next starting count for continued processing.
@return startFlag (bool) - State of startFlag for continued processing.

- Creates a list for every PDF in the subfolder
- Finds the index of the last PDF file, which is in the previous letter subfolder
- Counter is set to that last index
- Adjusts the image file name to match the current index

@author Mike
'''
def clean(letter, namePath, cemPath, initialCount, isFirstFile, start, startFlag, redac, redac2):
    global cemSet, miscSet, jewishSet
    pdfFiles = sorted(os.listdir(namePath))
    letters = sorted([folder for folder in os.listdir(cemPath) if os.path.isdir(os.path.join(cemPath, folder))])
    try:
        currentLetterIndex = letters.index(letter)
    except ValueError:
        return initialCount, startFlag
    folderBeforeIndex = max(0, currentLetterIndex - 1)
    folderBefore = letters[folderBeforeIndex]
    
    if currentLetterIndex == 0 or folderBefore == letter: 
        pdfFilesBefore = []
    else:
        pdfFilesBefore = sorted([file for file in os.listdir(os.path.join(cemPath, folderBefore)) if file.lower().endswith('.pdf')])

    if letter == "A" and folderBefore == "A":
        if os.path.basename(os.path.normpath(cemPath)).replace(" - Redacted", "") in cemSet:
            cemeteryFolders = sorted(os.listdir(os.path.dirname(cemPath)))
            currentCemeteryIndex = cemeteryFolders.index(os.path.basename(os.path.normpath(cemPath)))
        elif os.path.basename(os.path.normpath(cemPath)).replace(" - Redacted", "") in jewishSet:
            cemPath = os.path.dirname(cemPath)
            cemeteryFolders = sorted(os.listdir(os.path.dirname(cemPath)))
            currentCemeteryIndex = cemeteryFolders.index(f"Jewish{redac}")
        elif os.path.basename(os.path.normpath(cemPath)).replace(" - Redacted", "") in miscSet:
            cemPath = os.path.dirname(cemPath)
            cemeteryFolders = sorted(os.listdir(os.path.dirname(cemPath)))
            currentCemeteryIndex = cemeteryFolders.index(f"Misc{redac}")
        if currentCemeteryIndex > 0:
            previousCemeteryIndex = currentCemeteryIndex - 1
            previousCemeteryFolder = cemeteryFolders[previousCemeteryIndex]
            previousCemeteryPath = os.path.join(os.path.dirname(cemPath), previousCemeteryFolder)
            previousCemeteryLetters = sorted([folder for folder in os.listdir(previousCemeteryPath) if os.path.isdir(os.path.join(previousCemeteryPath, folder))])
            folderBefore = previousCemeteryLetters[-1]
            pdfFilesBefore = sorted([file for file in os.listdir(os.path.join(previousCemeteryPath, folderBefore)) if file.lower().endswith('.pdf')])
    maxCounter = 0
    for file in pdfFilesBefore:
        match = re.search(r'\d+', file)
        if match:
            number = int(match.group())
            maxCounter = max(maxCounter, number)
    counter = maxCounter + 1
    for x in pdfFiles:
        if counter != start and startFlag:
            isFirstFile = False
            if "a" in x[-5:]:
                pass
            else:
                counter += 1
            continue
        startFlag = False
        if isFirstFile:
            counter = initialCount
        if redac2:
            if "a redacted.pdf" in x:
                counter -= 1
                os.remove(os.path.join(namePath, x))
                print(f"Removed {x}")
            elif "b redacted.pdf" in x:
                counter -= 1
                os.remove(os.path.join(namePath, x))
                print(f"Removed {x}")
            else:
                newName = re.sub(r'\d+', f"{counter:05d}", x)
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
        else:
            if "a.pdf" in x:
                newName = re.sub(r'\d+', f"{counter:05d}", x)
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
                counter -= 1
            elif "b.pdf" in x:
                newName = re.sub(r'\d+', f"{counter:05d}", x)
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
            else:
                newName = re.sub(r'\d+', f"{counter:05d}", x)
                os.rename(os.path.join(namePath, x), os.path.join(namePath, newName))
        counter += 1
        isFirstFile = False

    return counter, startFlag

'''
Updates the record ID numbers in column A to fix fill any gaps and continue going in 
sequential order. Updates hyperlink references in an Excel spreadsheet starting from a 
specified index. This function adjusts the numeric ID within each hyperlink to ensure 
they accurately reflect the current record IDs following modifications.

@param startIndex (int): The row index from which to start updating hyperlinks in the 
                         spreadsheet.
@param newID (int): The new record ID to assign starting from startIndex.
@param fileDirectoryMap (dict): A dictionary mapping file IDs to their respective directories.

- Adjusts record ID in column A 
- Takes the hyperlink in the cell and changes the ID portion of the file name to match 
  the ID in column A for each row starting at startIndex

@return newID (int): The updated record ID after processing.

@author Mike
'''
def cleanHyperlinks(badWorksheet, startIndex, newID, fileDirectoryMap):
    for row in range(startIndex, badWorksheet.max_row + 1):
        badWorksheet[f'A{row}'].value = newID
        cellRef = f'O{row}'
        if badWorksheet[cellRef].hyperlink:
            origTarget = badWorksheet[cellRef].hyperlink.target.replace("%20", " ")
            formattedID = f"{newID:05d}"
            try:
                folderName = fileDirectoryMap.get(formattedID)[-1]
                modifiedString = re.sub(r'([\\/])[A-Z]([\\/])', f'\\1{folderName}\\2', origTarget)
                if badWorksheet.title in ["Jewish", "Misc"]:
                    match = re.search(r'[\\/][A-Z][\\/](.*?)([A-Z])\d+', origTarget)
                    extracted_value = match.group(1)  
                    modifiedString = re.sub(fr'({extracted_value}[A-Z])\d{{5}}', f'{extracted_value}{folderName}{formattedID}', modifiedString)
                else:
                    modifiedString = re.sub(fr'({badWorksheet.title}[A-Z])\d{{5}}', f'{badWorksheet.title}{folderName}{formattedID}', modifiedString)
                badWorksheet[cellRef].hyperlink = Hyperlink(ref=cellRef, target=modifiedString, display="PDF Image")
                badWorksheet[cellRef].value = "PDF Image"
            except Exception:
                pass
            print(f"Updated hyperlink from {origTarget} to {modifiedString} in row {row}.")
        newID += 1
    return newID


'''
Compares the letters in the folder path and file name of hyperlinks within an Excel 
workbook to ensure they match. The function identifies rows where the folder letter 
and file letter in the hyperlink do not match and prints these discrepancies.

@param workbookPath (str): The path to the Excel workbook to be processed.

- Iterates through each worksheet and each row starting from row 2.
- Extracts the hyperlink from column O (15th column) and compares the letters in the 
  folder path and file name.
- Prints hyperlinks where the letters do not match.

@author Mike
'''
def compareHyperlinkLetters(workbookPath):
    print("Checking hyperlinks...")
    workbook = openpyxl.load_workbook(workbookPath)
    pattern = re.compile(
        r'Cemetery - Redacted[\\/][^\\/]+ - Redacted[\\/](?P<folderLetter>[A-Z])[\\/][^\\/]+(?P<fileLetter>[A-Z])\d+ redacted\.pdf'
    )
    for sheetName in workbook.sheetnames:
        worksheet = workbook[sheetName]
        flag = False
        for sheetName in workbook.sheetnames:
            worksheet = workbook[sheetName]
            for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column):
                cell = row[14]  
                if cell.hyperlink:
                    hyperlink = cell.hyperlink.target.replace("%20", " ")
                    match = pattern.search(hyperlink)
                    if match:
                        folderLetter = match.group(1)
                        fileLetter = match.group(2)
                        if folderLetter == fileLetter:
                            pass
                        else:
                            flag = True
                            print(f"Not matching: {hyperlink} in row {cell.row}\n")
        if not flag:
            print("All hyperlinks are correct.")

'''
Main function to process and update Excel sheets containing good and bad IDs. It adjusts
image names, updates hyperlinks, cleans up records, and compares hyperlink letters to 
ensure consistency.

@param fullClean (bool): A flag indicating whether the full cleaning process should be done
@param goodIDs (list): List of good IDs to be processed.
@param badIDs (list): List of bad IDs to be processed.

- Loads the Excel workbook and identifies sheets and rows containing good and bad IDs.
- Adjusts image names and updates records based on good and bad IDs.
- Cleans and deletes records with bad IDs.
- Processes images to update records and hyperlinks.
- Compares hyperlink letters to ensure consistency.

The function uses various helper functions such as `adjustImageName`, `microsoftOCR.main`,
`cleanDelete`, `cleanImages`, `cleanHyperlinks`, and `compareHyperlinkLetters`.

The Excel file is saved after each major operation to ensure data consistency.

@author Mike
'''      
def main(fullClean, goodIDs, badIDs):
    cemeterys = [x for x in os.listdir(r"\\ucclerk\pgmdoc\Veterans\Cemetery")]
    miscs = [x for x in os.listdir(r"\\ucclerk\pgmdoc\Veterans\Cemetery\Misc")]
    jewishs = [x for x in os.listdir(r"\\ucclerk\pgmdoc\Veterans\Cemetery\Jewish")]
    global cemSet, miscSet, jewishSet
    cemSet, miscSet, jewishSet = set(cemeterys), set(miscs), set(jewishs)
    excelFilePath = r"\\ucclerk\pgmdoc\Veterans\Veterans.xlsx"
    basePath = fr'\\ucclerk\pgmdoc\Veterans\Cemetery'
    redactedBasePath = fr'\\ucclerk\pgmdoc\Veterans\Cemetery - redacted'
    workbook = openpyxl.load_workbook(excelFilePath)
    for goodID, badID in zip(goodIDs, badIDs):
        print(f"Processing Good ID: {goodID}, Bad ID: {badID}...")
        time.sleep(5)
        goodRow, badRow = None, None
        goodSheet, badSheet = None, None
        goodWorkSheet, badWorksheet = None, None

        # Find goodID and badID in sheets
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in range(2, worksheet.max_row + 1):
                if worksheet[f'A{row}'].value == goodID:
                    goodRow = row
                    goodSheet = sheet
                    print(f"Good ID {goodID} found in sheet: {goodSheet}, row: {goodRow}")
                if worksheet[f'A{row}'].value == badID:
                    badRow = row
                    badSheet = sheet
                    print(f"Bad ID {badID} found in sheet: {badSheet}, row: {badRow}")
                if goodRow and badRow:
                    break
            if goodRow and badRow:
                break
        if goodRow is None or badRow is None:
            print(f"Failed to find either Good ID {goodID} or Bad ID {badID}, skipping...")
            continue
        
        # Assign worksheets
        goodWorkSheet = workbook[goodSheet]
        badWorksheet = workbook[badSheet]

        # Get row data for both goodID and badID
        vals1 = [goodWorkSheet.cell(row=goodRow, column=col).value for col in range(2, 16)]
        vals2 = [badWorksheet.cell(row=badRow, column=col).value for col in range(2, 16)]
        print(f"Retrieved values from Good ID row {goodRow}:\n", end="")
        print(' | '.join(str(value) for value in vals1 if value is not None))
        print(f"Retrieved values from Bad ID row {badRow}:\n", end="")
        print(' | '.join(str(value) for value in vals2 if value is not None))
        time.sleep(8)
        
        # Clear the goodRow data (starting from column B) and delete badRow
        for col in range(2, goodWorkSheet.max_column + 1):
            goodWorkSheet.cell(row=goodRow, column=col).value = None
        print(f"Cleared data from Good ID row {goodRow}.")
        badWorksheet.delete_rows(badRow)
        print(f"Row {badRow} deleted successfully.")
        for row in range(badRow, badWorksheet.max_row + 1):
            nextRow = row + 1
            cellRef = f'O{row}'
            nextCellRef = f'O{nextRow}'
            if nextRow <= badWorksheet.max_row and badWorksheet[nextCellRef].hyperlink:
                tempHyperlink = badWorksheet[nextCellRef].hyperlink
                badWorksheet[cellRef].hyperlink = Hyperlink(ref=cellRef, target=tempHyperlink.target, display=tempHyperlink.display)
                badWorksheet[cellRef].value = badWorksheet[nextCellRef].value
                print(f"Hyperlink and value from row {nextRow} moved to row {row}.")
        if badWorksheet[f'O{badWorksheet.max_row - 1}'].hyperlink:
            lastRow = badWorksheet.max_row
            print(f"Adjusting the hyperlink in the last row, row {lastRow}.")
            secondLastHyperlink = badWorksheet[f'O{lastRow - 1}'].hyperlink
            newTarget = re.sub(r'(\d+)(?=%\d+)', lambda m: f"{int(m.group(1)) + 1:05d}", secondLastHyperlink.target)
            badWorksheet[f'O{lastRow}'].hyperlink = Hyperlink(ref=f'O{lastRow}', target=newTarget, display=secondLastHyperlink.display)
            print(f"Updated hyperlink in row {lastRow} to new target, {newTarget}.\n")
        
        # Merge values but keep goodID's hyperlink and other values for columns A and O
        mergedArray = []
        for col in range(len(vals1)):
            if col == 0:  # Column A (1st column, Good ID)
                mergedArray.append(vals1[col])
            elif col == 14:  # Column O (15th column, hyperlink)
                mergedArray.append(vals1[col])
            else:
                mergedArray.append(max(vals1[col], vals2[col], key=lambda x: len(str(x)) if x else 0))
        print(f"Merged values into Good ID row {goodRow}.")        
        counter = 2
        for value in mergedArray:
            goodWorkSheet.cell(row=goodRow, column=counter, value=value)
            counter += 1

        # Highlight cells in GoodRow, columns B - N and P
        highlight_fill = PatternFill(start_color="001AFF", end_color="001AFF", fill_type="solid")
        for col in range(2, 15):  # Columns B (2) to N (14)
            goodWorkSheet.cell(row=goodRow, column=col).fill = highlight_fill
        # Highlight column P (16)
        goodWorkSheet.cell(row=goodRow, column=16).fill = highlight_fill
        print(f"Highlighted cells in GoodRow {goodRow}, columns B - N and P with color #001AFF.")
        
        workbook.save(excelFilePath)
        
        # File handling: Apply the suffix "a" and "b" to goodFilePath and badFilePath and move if necessary
        goodIDFilePath, badIDFilePath = None, None
        for dirPath, _, fileNames in os.walk(basePath):
            for fileName in fileNames:
                if f"{goodID:05d}" in fileName:
                    goodIDFilePath = os.path.join(dirPath, fileName)
                if f"{badID:05d}" in fileName:
                    badIDFilePath = os.path.join(dirPath, fileName)
                if goodIDFilePath and badIDFilePath:
                    break
        if goodIDFilePath and badIDFilePath:
            goodIDPrefix = os.path.basename(os.path.dirname(goodIDFilePath))
            badIDPrefix = os.path.basename(os.path.dirname(badIDFilePath))
            newBadIDFilename = os.path.basename(goodIDFilePath).replace(f"{goodID:05d}", f"{goodID:05d}b").replace(badIDPrefix, goodIDPrefix)
            newBadIDFilePath = os.path.join(os.path.dirname(goodIDFilePath), newBadIDFilename)
            if not "a.pdf" in os.path.basename(goodIDFilePath):
                newGoodIDFilename = os.path.basename(goodIDFilePath).replace(".pdf", "a.pdf")
                newGoodIDFilePath = os.path.join(os.path.dirname(goodIDFilePath), newGoodIDFilename)
                os.rename(goodIDFilePath, newGoodIDFilePath)
                print(f"{os.path.basename(goodIDFilePath)} renamed to {newGoodIDFilename}")
            shutil.move(badIDFilePath, newBadIDFilePath)
            print(f"{os.path.basename(badIDFilePath)} moved and renamed to {newBadIDFilename}\n")

        # Redacted file handling: Combine redacted versions of the good and bad files
        redactedGoodFilePath, redactedBadFilePath = None, None
        for dirPath, _, fileNames in os.walk(redactedBasePath):
            for fileName in fileNames:
                if f"{goodID:05d}" in fileName:
                    redactedGoodFilePath = os.path.join(dirPath, fileName)
                if f"{badID:05d}" in fileName:
                    redactedBadFilePath = os.path.join(dirPath, fileName)
                if redactedGoodFilePath and redactedBadFilePath:
                    break
        if redactedGoodFilePath and redactedBadFilePath:
            print(f"Combining redacted files for Good ID {goodID} and Bad ID {badID}...")
            merger = PdfMerger()
            with open(redactedGoodFilePath, 'rb') as f1, open(redactedBadFilePath, 'rb') as f2:
                merger.append(f1)
                merger.append(f2)
                tempRedactedFilePath = redactedGoodFilePath.replace(".pdf", "_temp.pdf")
                with open(tempRedactedFilePath, 'wb') as tempFile:
                    merger.write(tempFile)
            merger.close()
            os.remove(redactedGoodFilePath)
            os.remove(redactedBadFilePath)
            os.rename(tempRedactedFilePath, redactedGoodFilePath)
            print(f"Redacted files combined and renamed back to {redactedGoodFilePath}\n")

    cleanImages(fullClean, 1, "", "")
    cleanImages(fullClean, 1, " - Redacted", " redacted")
    fileDirectoryMap = {}
    for dirPath, _, fileNames in os.walk(basePath):
        for fileName in fileNames:
            fileID = ''.join(filter(str.isdigit, fileName))[:5]
            if fileID:
                fileDirectoryMap[fileID] = dirPath
    newID = 1
    for sheetName in workbook.sheetnames:
        badWorksheet = workbook[sheetName]
        newID = cleanHyperlinks(badWorksheet, 2, newID, fileDirectoryMap)
    workbook.save(excelFilePath)
    compareHyperlinkLetters(excelFilePath)

if __name__ == "__main__": 
    fullClean = True

    goodIDs = []
    badIDs = []
    
    main(fullClean, goodIDs, badIDs)
